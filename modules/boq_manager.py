# modules/boq_manager.py

import streamlit as st
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional, Tuple

class BOQManager:
    """Complete BOQ Management System"""
    
    def __init__(self, db):
        self.db = db
    
    # ========== BOQ CRUD Operations ==========
    
    def create_boq(self, tender_id: int, company_id: int, 
                   rate_source: str, zone: str, notes: str = None) -> Tuple[Optional[int], str]:
        """Create a new BOQ"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # Get tender details
            cursor.execute("""
                SELECT tender_id, tender_title, procuring_entity, official_estimate
                FROM company_tenders
                WHERE id = ? AND company_id = ?
            """, (tender_id, company_id))
            tender = cursor.fetchone()
            
            if not tender:
                return None, "Tender not found"
            
            # Get rate version
            cursor.execute("""
                SELECT edition_year FROM rate_versions 
                WHERE source = ? AND is_active = 1
                LIMIT 1
            """, (rate_source,))
            version = cursor.fetchone()
            edition_year = version[0] if version else 2025
            
            cursor.execute("""
                INSERT INTO boq_generation_history (
                    user_id, company_id, tender_id, tender_title, procuring_entity,
                    selected_zone, rate_source, edition_year, status, notes, generated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'draft', ?, ?)
            """, (
                st.session_state.get('user_id', 0), company_id, tender[0], tender[1], tender[2],
                zone, rate_source, edition_year, notes, datetime.now()
            ))
            
            boq_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            self._log_activity(boq_id, 'create', f"BOQ created")
            
            return boq_id, "BOQ created successfully"
            
        except Exception as e:
            return None, str(e)
    
    def add_boq_item(self, boq_id: int, item_code: str, description: str, 
                     unit: str, quantity: float, unit_rate: float, 
                     is_custom: bool = False, notes: str = None) -> Tuple[bool, str]:
        """Add item to BOQ"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            total = quantity * unit_rate
            
            cursor.execute("""
                INSERT INTO boq_items (
                    boq_id, item_code, description, unit, quantity, unit_rate, total, is_custom, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (boq_id, item_code, description, unit, quantity, unit_rate, total, is_custom, notes))
            
            # Update BOQ totals
            cursor.execute("""
                UPDATE boq_generation_history 
                SET item_count = (SELECT COUNT(*) FROM boq_items WHERE boq_id = ?),
                    total_estimated_cost = (SELECT SUM(total) FROM boq_items WHERE boq_id = ?)
                WHERE id = ?
            """, (boq_id, boq_id, boq_id))
            
            conn.commit()
            conn.close()
            
            return True, "Item added successfully"
            
        except Exception as e:
            return False, str(e)
    
    def update_boq_item(self, item_id: int, quantity: float, unit_rate: float = None) -> Tuple[bool, str]:
        """Update BOQ item quantity or rate"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            if unit_rate:
                cursor.execute("""
                    UPDATE boq_items 
                    SET quantity = ?, unit_rate = ?, total = quantity * ?
                    WHERE id = ?
                """, (quantity, unit_rate, unit_rate, item_id))
            else:
                cursor.execute("""
                    UPDATE boq_items 
                    SET quantity = ?, total = quantity * unit_rate
                    WHERE id = ?
                """, (quantity, item_id))
            
            # Get boq_id to update totals
            cursor.execute("SELECT boq_id FROM boq_items WHERE id = ?", (item_id,))
            boq_id = cursor.fetchone()[0]
            
            cursor.execute("""
                UPDATE boq_generation_history 
                SET total_estimated_cost = (SELECT SUM(total) FROM boq_items WHERE boq_id = ?)
                WHERE id = ?
            """, (boq_id, boq_id))
            
            conn.commit()
            conn.close()
            
            return True, "Item updated successfully"
            
        except Exception as e:
            return False, str(e)
    
    def delete_boq_item(self, item_id: int) -> Tuple[bool, str]:
        """Delete item from BOQ"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT boq_id FROM boq_items WHERE id = ?", (item_id,))
            boq_id = cursor.fetchone()[0]
            
            cursor.execute("DELETE FROM boq_items WHERE id = ?", (item_id,))
            
            cursor.execute("""
                UPDATE boq_generation_history 
                SET item_count = (SELECT COUNT(*) FROM boq_items WHERE boq_id = ?),
                    total_estimated_cost = (SELECT SUM(total) FROM boq_items WHERE boq_id = ?)
                WHERE id = ?
            """, (boq_id, boq_id, boq_id))
            
            conn.commit()
            conn.close()
            
            return True, "Item deleted successfully"
            
        except Exception as e:
            return False, str(e)
    
    def get_boq(self, boq_id: int) -> Optional[Dict]:
        """Get complete BOQ with items"""
        try:
            conn = self.db.get_connection()
            
            # Get BOQ header
            cursor = conn.cursor()
            cursor.execute("""
                SELECT b.*, c.company_name, u.username as created_by_name
                FROM boq_generation_history b
                LEFT JOIN companies c ON b.company_id = c.id
                LEFT JOIN users u ON b.user_id = u.id
                WHERE b.id = ?
            """, (boq_id,))
            
            row = cursor.fetchone()
            if not row:
                return None
            
            columns = [col[1] for col in cursor.description]
            boq = dict(zip(columns, row))
            
            # Get BOQ items
            items = pd.read_sql_query("""
                SELECT * FROM boq_items
                WHERE boq_id = ?
                ORDER BY id
            """, conn, params=[boq_id])
            
            # Get approval history
            history = pd.read_sql_query("""
                SELECT * FROM boq_approval_history
                WHERE boq_id = ?
                ORDER BY created_at DESC
            """, conn, params=[boq_id])
            
            conn.close()
            
            return {
                'boq': boq,
                'items': items,
                'history': history
            }
            
        except Exception as e:
            print(f"Error getting BOQ: {e}")
            return None
    
    # ========== BOQ Workflow ==========
    
    def submit_boq(self, boq_id: int, comment: str = None) -> Tuple[bool, str]:
        """Submit BOQ for approval"""
        try:
            return self._update_boq_status(boq_id, 'submitted', comment, 'submitted_for_review')
        except Exception as e:
            return False, str(e)
    
    def approve_boq(self, boq_id: int, comment: str = None) -> Tuple[bool, str]:
        """Approve BOQ"""
        try:
            return self._update_boq_status(boq_id, 'approved', comment, 'approved')
        except Exception as e:
            return False, str(e)
    
    def reject_boq(self, boq_id: int, comment: str = None) -> Tuple[bool, str]:
        """Reject BOQ with comments"""
        try:
            return self._update_boq_status(boq_id, 'rejected', comment, 'rejected')
        except Exception as e:
            return False, str(e)
    
    def _update_boq_status(self, boq_id: int, status: str, comment: str, action: str) -> Tuple[bool, str]:
        """Internal method to update BOQ status"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE boq_generation_history 
                SET status = ?, updated_at = ?
                WHERE id = ?
            """, (status, datetime.now(), boq_id))
            
            # Log approval history
            cursor.execute("""
                INSERT INTO boq_approval_history (boq_id, action, comment, user_id, username, user_role)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (boq_id, action, comment, 
                  st.session_state.get('user_id', 0),
                  st.session_state.get('username', 'system'),
                  st.session_state.get('user_role', 'viewer')))
            
            conn.commit()
            conn.close()
            
            self._log_activity(boq_id, action, comment)
            
            return True, f"BOQ {status} successfully"
            
        except Exception as e:
            return False, str(e)
    
    def delete_boq(self, boq_id: int) -> Tuple[bool, str]:
        """Delete BOQ"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("DELETE FROM boq_items WHERE boq_id = ?", (boq_id,))
            cursor.execute("DELETE FROM boq_approval_history WHERE boq_id = ?", (boq_id,))
            cursor.execute("DELETE FROM boq_generation_history WHERE id = ?", (boq_id,))
            
            conn.commit()
            conn.close()
            
            return True, "BOQ deleted successfully"
            
        except Exception as e:
            return False, str(e)
    
    def copy_boq(self, boq_id: int) -> Tuple[Optional[int], str]:
        """Create a copy of existing BOQ"""
        try:
            original = self.get_boq(boq_id)
            if not original:
                return None, "Original BOQ not found"
            
            # Create new BOQ
            new_boq_id, msg = self.create_boq(
                tender_id=original['boq']['tender_id'],
                company_id=original['boq']['company_id'],
                rate_source=original['boq']['rate_source'],
                zone=original['boq']['selected_zone'],
                notes=f"Copy of BOQ #{boq_id}"
            )
            
            if not new_boq_id:
                return None, msg
            
            # Copy items
            for _, item in original['items'].iterrows():
                self.add_boq_item(
                    boq_id=new_boq_id,
                    item_code=item['item_code'],
                    description=item['description'],
                    unit=item['unit'],
                    quantity=item['quantity'],
                    unit_rate=item['unit_rate'],
                    is_custom=item.get('is_custom', False),
                    notes=item.get('notes')
                )
            
            return new_boq_id, "BOQ copied successfully"
            
        except Exception as e:
            return None, str(e)
    
    # ========== Reporting ==========
    
    def export_boq_to_excel(self, boq_id: int) -> bytes:
        """Export BOQ to Excel format"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        boq_data = self.get_boq(boq_id)
        if not boq_data:
            return None
        
        wb = Workbook()
        
        # Main BOQ sheet
        ws = wb.active
        ws.title = "BOQ"
        
        # Header styling
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        
        # Company header
        ws.merge_cells('A1:F1')
        ws['A1'] = f"BOQ Report - {boq_data['boq'].get('tender_title', 'N/A')}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Tender details
        ws['A3'] = "Tender ID:"
        ws['B3'] = boq_data['boq'].get('tender_id', 'N/A')
        ws['A4'] = "Rate Source:"
        ws['B4'] = boq_data['boq'].get('rate_source', 'N/A')
        ws['A5'] = "Zone:"
        ws['B5'] = boq_data['boq'].get('selected_zone', 'N/A')
        ws['A6'] = "Status:"
        ws['B6'] = boq_data['boq'].get('status', 'N/A').upper()
        
        # Items header
        headers = ['Sl No', 'Item Code', 'Description', 'Unit', 'Quantity', 'Unit Rate (BDT)', 'Total (BDT)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Items data
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        total = 0
        for idx, (_, item) in enumerate(boq_data['items'].iterrows(), 1):
            row = 8 + idx
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item['item_code'])
            ws.cell(row=row, column=3, value=item['description'][:100])
            ws.cell(row=row, column=4, value=item['unit'])
            ws.cell(row=row, column=5, value=item['quantity'])
            ws.cell(row=row, column=6, value=round(item['unit_rate'], 2))
            ws.cell(row=row, column=7, value=round(item['total'], 2))
            total += item['total']
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
        
        # Total row
        total_row = 9 + len(boq_data['items'])
        ws.merge_cells(f'A{total_row}:F{total_row}')
        ws.cell(row=total_row, column=1, value="GRAND TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=round(total, 2))
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        
        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 15
        ws.column_dimensions['C'].width = 50
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    # ========== Helper Methods ==========
    
        def _log_activity(self, boq_id: int, action: str, details: str):
            """Log BOQ activity"""
            try:
                conn = self.db.get_connection()
                cursor = conn.cursor()
                                
                # Table already exists in unified manager
                
                cursor.execute("""
                    INSERT INTO boq_activity_log (boq_id, action, details, user_id, username, user_role)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (boq_id, action, details,
                    st.session_state.get('user_id', 0),
                    st.session_state.get('username', 'system'),
                    st.session_state.get('user_role', 'viewer')))
                
                conn.commit()
                conn.close()
                
            except Exception as e:
                print(f"Error logging activity: {e}")

    
    def get_boq_list(self, company_id: int = None, status: str = None, limit: int = 100) -> pd.DataFrame:
        """Get list of BOQs with filters"""
        try:
            conn = self.db.get_connection()
            user_role = st.session_state.get('user_role', 'viewer')
            
            if user_role in ['admin', 'system_admin']:
                query = """
                    SELECT b.*, c.company_name, u.username as created_by_name
                    FROM boq_generation_history b
                    LEFT JOIN companies c ON b.company_id = c.id
                    LEFT JOIN users u ON b.user_id = u.id
                    WHERE 1=1
                """
                params = []
            else:
                if not company_id:
                    company_id = st.session_state.get('company_id')
                query = """
                    SELECT b.*, u.username as created_by_name
                    FROM boq_generation_history b
                    LEFT JOIN users u ON b.user_id = u.id
                    WHERE b.company_id = ?
                """
                params = [company_id]
            
            if status:
                query += " AND b.status = ?"
                params.append(status)
            
            query += " ORDER BY b.generated_at DESC LIMIT ?"
            params.append(limit)
            
            df = pd.read_sql_query(query, conn, params=params)
            conn.close()
            return df
            
        except Exception as e:
            print(f"Error getting BOQ list: {e}")
            return pd.DataFrame()
        
    def lock_boq(self, boq_id: int) -> Tuple[bool, str]:
        """Lock BOQ after bid submission (prevents further edits)"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE boq_generation_history 
                SET is_locked = 1, locked_at = ?, locked_by = ?
                WHERE id = ?
            """, (datetime.now(), st.session_state.get('user_id', 0), boq_id))
            
            conn.commit()
            conn.close()
            
            self._log_activity(boq_id, 'lock', "BOQ locked after bid submission")
            return True, "BOQ locked successfully"
            
        except Exception as e:
            return False, str(e)

    def unlock_boq(self, boq_id: int) -> Tuple[bool, str]:
        """Unlock BOQ (admin only)"""
        try:
            user_role = st.session_state.get('user_role', 'viewer')
            if user_role not in ['admin', 'system_admin']:
                return False, "Only administrators can unlock BOQs"
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE boq_generation_history 
                SET is_locked = 0, locked_at = NULL, locked_by = NULL
                WHERE id = ?
            """, (boq_id,))
            
            conn.commit()
            conn.close()
            
            self._log_activity(boq_id, 'unlock', "BOQ unlocked by admin")
            return True, "BOQ unlocked successfully"
            
        except Exception as e:
            return False, str(e)    