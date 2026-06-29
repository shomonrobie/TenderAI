# database/crud_operations.py
import streamlit as st
import sqlite3
import json
import logging
import hashlib
import secrets
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any, Tuple, Union
from contextlib import contextmanager
import bcrypt
import pandas as pd
from database.connection import db_connection
import re
import string
import os
from unittest import result
from typing import List, Union, Dict, Callable, Optional

logger = logging.getLogger(__name__)

class DatabaseCRUD:
    """Handles all database CRUD operations"""
    
    def __init__(self, db_path="data/tender_system.db"):
        self.db_path = db_path  # Kept for compatibility, but not used directly
        self.db_conn = db_connection
    
    def get_connection(self):
        """
        Returns raw database connection (works with BOTH patterns)
        - Direct: conn = db.get_connection()
        - Context: with db.get_connection() as conn:
        """
        return self.db_conn._connect_func()
    
    def get_db_type(self):
        """Get current database type"""
        return self.db_conn.db_type
    
    def _row_to_dict(self, row, cursor):
        """Convert database row to dictionary (works for all databases)"""
        if row is None:
            return None
        if isinstance(row, dict):
            return row
        columns = [description[0] for description in cursor.description]
        return dict(zip(columns, row))

    def execute(self, sql: str, params: tuple = None) -> int:
        """Execute SQL with parameter substitution"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            if params:
                cursor.execute(sql, params)
            else:
                cursor.execute(sql)
            return cursor.rowcount
    
    def query(self, sql: str, params: tuple = None):
        """Execute query and return results as list of dicts"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            if params:
                cursor.execute(sql, params)
            else:
                cursor.execute(sql)
            
            # Convert to dict based on database type
            if self.get_db_type() in ['postgresql', 'cockroachdb']:
                # Already dict from RealDictCursor
                return [dict(row) for row in cursor.fetchall()]
            else:
                # SQLite or MySQL - need manual conversion
                columns = [description[0] for description in cursor.description]
                rows = cursor.fetchall()
                return [dict(zip(columns, row)) for row in rows]
    
    def query_one(self, sql: str, params: tuple = None):
        """Execute query and return single result as dict"""
        results = self.query(sql, params)
        return results[0] if results else None
    
    # ==================== DATABASE-AGNOSTIC SCHEMA METHODS ====================
    
    def get_existing_tables(self):
        """Get list of existing tables (database-agnostic)"""
        db_type = self.get_db_type()
        
        if db_type == 'sqlite':
            sql = "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        elif db_type in ['postgresql', 'cockroachdb']:
            sql = """
                SELECT table_name FROM information_schema.tables 
                WHERE table_schema = 'public' AND table_type = 'BASE TABLE'
            """
        elif db_type == 'mysql':
            sql = "SHOW TABLES"
        else:
            return set()
        
        results = self.query(sql)
        if db_type == 'mysql':
            # MySQL returns dict with key like 'Tables_in_database'
            return {list(row.values())[0] for row in results}
        return {row['name'] if 'name' in row else list(row.values())[0] for row in results}
    
    def get_table_columns(self, table_name: str):
        """Get table columns (database-agnostic)"""
        db_type = self.get_db_type()
        
        if db_type == 'sqlite':
            sql = f"PRAGMA table_info({table_name})"
            results = self.query(sql)
            return {row['name']: {
                'type': row['type'],
                'notnull': bool(row['notnull']),
                'default': row['dflt_value'],
                'pk': bool(row['pk'])
            } for row in results}
        
        elif db_type in ['postgresql', 'cockroachdb']:
            sql = """
                SELECT column_name, data_type, is_nullable, column_default
                FROM information_schema.columns 
                WHERE table_name = %s
            """
            results = self.query(sql, (table_name,))
            return {row['column_name']: {
                'type': row['data_type'],
                'notnull': row['is_nullable'] == 'NO',
                'default': row['column_default'],
                'pk': False
            } for row in results}
        
        elif db_type == 'mysql':
            sql = f"DESCRIBE {table_name}"
            results = self.query(sql)
            return {row['Field']: {
                'type': row['Type'],
                'notnull': row['Null'] == 'NO',
                'default': row['Default'],
                'pk': row['Key'] == 'PRI'
            } for row in results}
        
        return {}
    
    def get_existing_indexes(self):
        """Get existing indexes (database-agnostic)"""
        db_type = self.get_db_type()
        
        if db_type == 'sqlite':
            sql = "SELECT name, tbl_name FROM sqlite_master WHERE type='index' AND name NOT LIKE 'sqlite_%'"
            results = self.query(sql)
            return {row['name']: row['tbl_name'] for row in results}
        
        elif db_type in ['postgresql', 'cockroachdb']:
            sql = """
                SELECT indexname, tablename 
                FROM pg_indexes 
                WHERE schemaname = 'public'
            """
            results = self.query(sql)
            return {row['indexname']: row['tablename'] for row in results}
        
        elif db_type == 'mysql':
            sql = "SHOW INDEX FROM information_schema.statistics WHERE table_schema = DATABASE()"
            results = self.query(sql)
            return {row['Index_name']: row['Table'] for row in results}
        
        return {}
    
    def column_exists(self, table_name: str, column_name: str) -> bool:
        """Check if column exists (database-agnostic)"""
        columns = self.get_table_columns(table_name)
        return column_name in columns
    
    def table_exists(self, table_name: str) -> bool:
        """Check if table exists (database-agnostic)"""
        tables = self.get_existing_tables()
        return table_name in tables
    
    def add_column_if_not_exists(self, table_name: str, column_name: str, column_type: str) -> bool:
        """Add column if it doesn't exist (database-agnostic)"""
        if self.column_exists(table_name, column_name):
            return False
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            db_type = self.get_db_type()
            
            if db_type == 'sqlite':
                sql = f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}"
            elif db_type in ['postgresql', 'cockroachdb']:
                sql = f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS {column_name} {column_type}"
            elif db_type == 'mysql':
                sql = f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS {column_name} {column_type}"
            else:
                return False
            
            cursor.execute(sql)
            return True
    # ==================== HELPER METHODS ====================
    
    def validate_bangladesh_mobile(self, mobile: str) -> bool:
        """Validate Bangladeshi mobile number"""
        if not mobile:
            return False
        mobile = re.sub(r'[\s\-+]', '', mobile)
        if mobile.startswith('88'):
            mobile = mobile[2:]
        pattern = r'^01[3-9]\d{8}$'
        return bool(re.match(pattern, mobile))
    
    def normalize_mobile(self, mobile: str) -> str:
        """Normalize mobile number to standard format"""
        if not mobile:
            return mobile
        mobile = re.sub(r'[\s\-+]', '', mobile)
        if mobile.startswith('+88'):
            mobile = mobile[3:]
        elif mobile.startswith('88'):
            mobile = mobile[2:]
        return mobile


    def create_user(self, company_id: int, user_data: Dict, created_by: int = None) -> tuple:
        """
        Create a new user with unique mobile number and email
        
        Args:
            company_id: Company ID (can be None for system users)
            user_data: Dict with keys: username, password, email, full_name, phone, mobile_number, role
            created_by: User ID of creator
        
        Returns:
            (success: bool, message: str or user_id: int)
        """
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Validate required fields
            if not user_data.get('username'):
                return False, "Username is required"
            if not user_data.get('password'):
                return False, "Password is required"
            if not user_data.get('email'):
                return False, "Email is required"
            if not user_data.get('mobile_number'):
                return False, "Mobile number is required"
            
            # Validate mobile number format
            mobile = self.normalize_mobile(user_data['mobile_number'])
            if not self.validate_bangladesh_mobile(mobile):
                return False, f"Invalid mobile number: {mobile}. Must be 11 digits starting with 01"
            
            # Check if mobile number already exists
            cursor.execute("SELECT id FROM users WHERE mobile_number = ?", (mobile,))
            if cursor.fetchone():
                return False, f"Mobile number {mobile} is already registered"
            
            # Check if email already exists
            cursor.execute("SELECT id FROM users WHERE email = ?", (user_data['email'],))
            if cursor.fetchone():
                return False, f"Email {user_data['email']} is already registered"
            
            # Check if username already exists
            cursor.execute("SELECT id FROM users WHERE username = ?", (user_data['username'],))
            if cursor.fetchone():
                return False, f"Username {user_data['username']} is already taken"
            
            # Hash password
            hashed = bcrypt.hashpw(user_data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            # Set default role if not provided
            role = user_data.get('role', 'viewer')
            
            # Insert user
            cursor.execute("""
                INSERT INTO users (
                    company_id, username, password, email, full_name, phone, mobile_number,
                    role, is_active, is_approved, created_by, created_at, account_type
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?)
            """, (
                company_id,
                user_data['username'],
                hashed,
                user_data['email'],
                user_data.get('full_name', ''),
                user_data.get('phone', ''),
                mobile,
                role,
                1,  # is_active
                1,  # is_approved (auto-approved for company users)
                created_by,
                'company'
            ))
            
            user_id = cursor.lastrowid
            return True, user_id

    def create_google_user(self, user_data: Dict) -> tuple:
        """
        Create a new user from Google OAuth (mobile number optional)
        
        Args:
            user_data: Dict with keys: username, email, full_name, google_id, phone (optional)
        
        Returns:
            (success: bool, message: str or user_id: int)
        """
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Validate required fields
            if not user_data.get('username'):
                return False, "Username is required"
            if not user_data.get('email'):
                return False, "Email is required"
            if not user_data.get('full_name'):
                return False, "Full name is required"
            
            # Generate a random password for Google users
            import secrets
            import bcrypt
            temp_password = secrets.token_urlsafe(16)
            hashed = bcrypt.hashpw(temp_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            # Check if email already exists
            cursor.execute("SELECT id, auth_provider FROM users WHERE email = ?", (user_data['email'],))
            existing = cursor.fetchone()
            if existing:
                # If user exists, update with Google info if it's a Google user
                user_id = existing[0]
                google_id = user_data.get('google_id')
                if google_id:
                    cursor.execute("""
                        UPDATE users 
                        SET google_id = ?, auth_provider = 'google', email_verified = 1, 
                            auth_provider_user_id = ?
                        WHERE id = ?
                    """, (google_id, google_id, user_id))
                    conn.commit()
                    return True, user_id
                return False, f"Email {user_data['email']} is already registered"
            
            # Check if username already exists
            cursor.execute("SELECT id FROM users WHERE username = ?", (user_data['username'],))
            if cursor.fetchone():
                return False, f"Username {user_data['username']} is already taken"
            
            # Mobile is optional for Google users
            mobile = user_data.get('mobile_number', '')
            if mobile:
                mobile = self.normalize_mobile(mobile)
                if self.validate_bangladesh_mobile(mobile):
                    cursor.execute("SELECT id FROM users WHERE mobile_number = ?", (mobile,))
                    if cursor.fetchone():
                        return False, f"Mobile number {mobile} is already registered"
                else:
                    mobile = ''
            
            google_id = user_data.get('google_id', '')
            
            # Insert user with OAuth columns
            cursor.execute("""
                INSERT INTO users (
                    username, password, email, full_name, phone, mobile_number,
                    role, is_active, is_approved, created_at, account_type,
                    auth_provider, google_id, google_email, email_verified,
                    auth_provider_user_id, google_picture
                ) VALUES (
                    ?, ?, ?, ?, ?, ?,
                    'individual', 1, 1, CURRENT_TIMESTAMP, 'individual',
                    'google', ?, ?, 1,
                    ?, ?
                )
            """, (
                user_data['username'],
                hashed,
                user_data['email'],
                user_data['full_name'],
                user_data.get('phone', ''),
                mobile,
                google_id,
                user_data['email'],
                google_id,  # auth_provider_user_id
                user_data.get('picture', '')  # google_picture
            ))
            
            user_id = cursor.lastrowid
            
            # Also insert into user_oauth table for extensibility
            if google_id:
                cursor.execute("""
                    INSERT INTO user_oauth (user_id, provider_name, provider_user_id, provider_email)
                    VALUES (?, 'google', ?, ?)
                """, (user_id, google_id, user_data['email']))
            
            conn.commit()
            return True, user_id

    def create_facebook_user(self, user_data: Dict) -> tuple:
        """
        Create a new user from Facebook OAuth (disabled for now)
        
        Args:
            user_data: Dict with keys: username, email, full_name, facebook_id, phone (optional)
        
        Returns:
            (success: bool, message: str or user_id: int)
        """
        # TODO: Implement Facebook OAuth
        logger.warning("Facebook OAuth is not yet implemented")
        return False, "Facebook OAuth is not yet implemented. Please use Google or traditional registration."

    def get_user_by_oauth_provider(self, provider_name: str, provider_user_id: str):
        """Get user by OAuth provider and provider user ID"""
        try:
            # First check user_oauth table
            result = self.query_one("""
                SELECT u.* FROM users u
                JOIN user_oauth o ON u.id = o.user_id
                WHERE o.provider_name = ? AND o.provider_user_id = ?
            """, (provider_name, provider_user_id))
            
            if result:
                return result
            
            # Fallback to direct column check for Google
            if provider_name == 'google':
                result = self.query_one("""
                    SELECT * FROM users WHERE google_id = ?
                """, (provider_user_id,))
                return result
            
            if provider_name == 'facebook':
                result = self.query_one("""
                    SELECT * FROM users WHERE facebook_id = ?
                """, (provider_user_id,))
                return result
            
            return None
        except Exception as e:
            logger.error(f"Error getting user by OAuth provider: {e}")
            return None           
    def authenticate_user(self, username_or_email: str, password: str) -> Optional[Dict]:
        """Authenticate user - supports both SHA256 and bcrypt hashes"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                SELECT id, username, email, full_name, role, company_id, 
                    is_active, is_approved, account_type, created_at, last_login,
                    password
                FROM users 
                WHERE (username = ? OR email = ?) AND is_active = 1
            """, (username_or_email, username_or_email))
            
            row = cursor.fetchone()
            
            if row:
                # Convert to dict if it's a tuple
                if isinstance(row, dict):
                    user_dict = row
                else:
                    # Get column names from cursor description
                    columns = [description[0] for description in cursor.description]
                    user_dict = dict(zip(columns, row))
                
                stored_password = user_dict.get('password', '')
                
                # ===== TRY SHA256 FIRST =====
                import hashlib
                hashed_input = hashlib.sha256(password.encode()).hexdigest()
                
                if stored_password == hashed_input:
                    # Update last login
                    cursor.execute(
                        "UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE id = ?",
                        (user_dict['id'],)
                    )
                    conn.commit()
                    del user_dict['password']
                    return user_dict
                
                # ===== TRY BCRYPT AS FALLBACK =====
                try:
                    import bcrypt
                    # Check if it looks like a bcrypt hash
                    if stored_password.startswith('$2'):
                        if bcrypt.checkpw(password.encode(), stored_password.encode()):
                            # Convert to SHA256 for future use
                            new_hash = hashlib.sha256(password.encode()).hexdigest()
                            cursor.execute(
                                "UPDATE users SET password = ? WHERE id = ?",
                                (new_hash, user_dict['id'])
                            )
                            conn.commit()
                            
                            # Update last login
                            cursor.execute(
                                "UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE id = ?",
                                (user_dict['id'],)
                            )
                            conn.commit()
                            
                            del user_dict['password']
                            return user_dict
                except:
                    pass
                
                return None
            
    def get_user_by_id(self, user_id: int) -> Optional[Dict]:
        """Get user by ID (without password)"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, username, email, mobile_number, full_name, role, 
                       company_id, is_active, created_at, last_login,
                       mobile_verified, email_verified
                FROM users 
                WHERE id = ?
            """, (user_id,))
            
            row = cursor.fetchone()
            return dict(row) if row else None
    
    def get_user_by_mobile(self, mobile_number: str) -> Optional[Dict]:
        """Get user by mobile number"""
        mobile_number = self.normalize_mobile(mobile_number)
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, username, email, mobile_number, full_name, role, 
                       company_id, is_active, created_at, last_login,
                       mobile_verified, email_verified
                FROM users 
                WHERE mobile_number = ?
            """, (mobile_number,))
            
            row = cursor.fetchone()
            return dict(row) if row else None
    
    def update_user_verification(self, user_id: int, contact_type: str, verified: bool = True) -> bool:
        """Update user verification status"""
        field = f"{contact_type}_verified"
        time_field = f"{contact_type}_verified_at"
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute(f"""
                UPDATE users
                SET {field} = ?, {time_field} = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (1 if verified else 0, user_id))
            
            return cursor.rowcount > 0
    
    def get_all_users(self, company_id=None, role=None):
        """Get all users as dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            query = '''
            SELECT u.id, u.username, u.email, u.full_name, u.phone,
                u.mobile_number, u.mobile_verified,  -- ✅ ADD THESE
                u.role, u.is_active, u.created_at, u.last_login, 
                c.company_name, u.is_approved
            FROM users u
            JOIN companies c ON u.company_id = c.id
            WHERE 1=1
            '''
            params = []
            
            if company_id:
                query += " AND u.company_id = ?"
                params.append(company_id)
            if role:
                query += " AND u.role = ?"
                params.append(role)
            
            query += " ORDER BY u.created_at DESC"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            # Convert to list of dictionaries
            users = []
            for row in rows:
                users.append({
                    'id': row.get('id'),
                    'username': row.get('username'),
                    'email': row.get('email'),
                    'full_name': row.get('full_name'),
                    'phone': row.get('phone', ''),
                    'mobile_number': row.get('mobile_number', ''),  # ✅ ADDED
                    'mobile_verified': row.get('mobile_verified', 0),  # ✅ ADDED
                    'role': row.get('role', 'viewer'),
                    'is_active': row.get('is_active', 1),
                    'created_at': row.get('created_at'),
                    'last_login': row.get('last_login'),
                    'company_name': row.get('company_name'),
                    'is_approved': row.get('is_approved', 1)
                })
            return users

    def get_all_users_filtered(self, company_id=None, search="", role="", status=None, limit=20, offset=0):
        """Get users filtered by company, search, role, status"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Build base query
            if company_id == -1:
                # Special case: get system users only (company_id IS NULL)
                query = """
                    SELECT u.id, u.username, u.email, u.full_name, u.phone, 
                        u.mobile_number, u.mobile_verified,
                        u.role, u.is_active, u.created_at, u.last_login, 
                        NULL as company_name, u.is_approved
                    FROM users u
                    WHERE u.company_id IS NULL
                """
                params = []
            elif company_id and company_id > 0:
                # Get users for a specific company
                query = """
                    SELECT u.id, u.username, u.email, u.full_name, u.phone,
                        u.mobile_number, u.mobile_verified,
                        u.role, u.is_active, u.created_at, u.last_login, 
                        c.company_name, u.is_approved
                    FROM users u
                    LEFT JOIN companies c ON u.company_id = c.id
                    WHERE u.company_id = ?
                """
                params = [company_id]
            else:
                # Get all users (including system users) - for system admin
                query = """
                    SELECT u.id, u.username, u.email, u.full_name, u.phone,
                        u.mobile_number, u.mobile_verified,
                        u.role, u.is_active, u.created_at, u.last_login, 
                        c.company_name, u.is_approved
                    FROM users u
                    LEFT JOIN companies c ON u.company_id = c.id
                    WHERE 1=1
                """
                params = []
            
            # Add search filter
            if search:
                query += " AND (u.username LIKE ? OR u.email LIKE ? OR u.full_name LIKE ?)"
                params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
            
            # Add role filter
            if role:
                query += " AND u.role = ?"
                params.append(role)
            
            # Add status filter
            if status is not None:
                query += " AND u.is_active = ?"
                params.append(1 if status else 0)
            
            # Create count query
            count_query = query.replace(
                "SELECT u.id, u.username, u.email, u.full_name, u.phone, u.mobile_number, u.mobile_verified, u.role, u.is_active, u.created_at, u.last_login, c.company_name, u.is_approved",
                "SELECT COUNT(*)"
            )
            
            # Execute count query
            try:
                cursor.execute(count_query, params)
                count_result = cursor.fetchone()
                total = count_result[0] if count_result else 0
            except Exception as e:
                print(f"Count query error: {e}")
                total = 0
            
            # Add pagination
            query += " ORDER BY u.created_at DESC LIMIT ? OFFSET ?"
            params.extend([limit, offset])
            
            # Execute main query
            try:
                cursor.execute(query, params)
                rows = cursor.fetchall()
            except Exception as e:
                print(f"Main query error: {e}")
                rows = []
            
            # Convert to list of dicts
            users = []
            for row in rows:
                users.append({
                    'id': row.get('id'),
                    'username': row.get('username'),
                    'email': row.get('email'),
                    'full_name': row.get('full_name'),
                    'phone': row.get('phone', ''),
                    'mobile_number': row.get('mobile_number', ''),  # ✅ ADDED
                    'mobile_verified': row.get('mobile_verified', 0),  # ✅ ADDED
                    'role': row.get('role', 'viewer'),
                    'is_active': row.get('is_active', 1),
                    'created_at': row.get('created_at'),
                    'last_login': row.get('last_login'),
                    'company_name': row.get('company_name'),
                    'is_approved': row.get('is_approved', 1)
                })
            
            return users, total

    # ==================== USER PROFILE METHODS ====================

    def get_user_profile(self, user_id: int) -> Optional[Dict]:
        """Get complete user profile with social links"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Get user data with new fields
            cursor.execute("""
                SELECT id, username, email, mobile_number, full_name, phone,
                    role, company_id, is_active, created_at, last_login,
                    mobile_verified, email_verified,
                    avatar_url, bio, location, website,
                    specialization, years_experience
                FROM users 
                WHERE id = ?
            """, (user_id,))
            
            user = cursor.fetchone()
            if not user:
                return None
            
            # Convert to dict
            if isinstance(user, dict):
                user_dict = user
            else:
                columns = [description[0] for description in cursor.description]
                user_dict = dict(zip(columns, user))
            
            # Get social links
            social_links = self.get_user_social_links(user_id)
            user_dict['social_links'] = social_links
            
            return user_dict
    
    def update_user_profile(self, user_id: int, **kwargs) -> bool:
        """Update user profile fields"""
        allowed_fields = [
            'full_name', 'phone', 'email', 'bio', 'location', 
            'website', 'avatar_url', 'specialization', 'years_experience'
        ]
        
        updates = {}
        for key, value in kwargs.items():
            if key in allowed_fields:
                updates[key] = value
        
        if not updates:
            return False
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            set_clause = ', '.join([f"{key} = ?" for key in updates.keys()])
            values = list(updates.values())
            values.append(user_id)
            
            cursor.execute(f"""
                UPDATE users 
                SET {set_clause}
                WHERE id = ?
            """, values)
            
            conn.commit()
            return cursor.rowcount > 0
    
    def update_user_avatar(self, user_id: int, avatar_url: str) -> bool:
        """Update user avatar URL"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                UPDATE users 
                SET avatar_url = ? 
                WHERE id = ?
            """, (avatar_url, user_id))
            conn.commit()
            return cursor.rowcount > 0
    
    # ==================== SOCIAL LINKS METHODS ====================
    
    def get_user_social_links(self, user_id: int) -> List[Dict]:
        """Get all social links for a user"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, platform, url, is_active, is_public, icon, display_order
                FROM social_links 
                WHERE user_id = ?
                ORDER BY display_order ASC, platform ASC
            """, (user_id,))
            
            rows = cursor.fetchall()
            
            # Convert to list of dicts
            links = []
            for row in rows:
                if isinstance(row, dict):
                    links.append(row)
                else:
                    columns = [description[0] for description in cursor.description]
                    links.append(dict(zip(columns, row)))
            
            return links
    
    def add_social_link(self, user_id: int, platform: str, url: str, 
                        is_public: bool = True, icon: str = None) -> bool:
        """Add a social link for a user"""
        # Validate platform
        valid_platforms = [
            'facebook', 'twitter', 'instagram', 'linkedin', 'github',
            'youtube', 'tiktok', 'pinterest', 'reddit', 'whatsapp',
            'telegram', 'discord', 'slack', 'medium', 'dev.to'
        ]
        
        if platform.lower() not in valid_platforms:
            return False
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Check if platform already exists
            cursor.execute("""
                SELECT id FROM social_links 
                WHERE user_id = ? AND platform = ?
            """, (user_id, platform.lower()))
            
            if cursor.fetchone():
                return False
            
            # Get max display order
            cursor.execute("""
                SELECT MAX(display_order) as max_order 
                FROM social_links 
                WHERE user_id = ?
            """, (user_id,))
            row = cursor.fetchone()
            max_order = row['max_order'] if row and row['max_order'] is not None else -1
            
            cursor.execute("""
                INSERT INTO social_links 
                (user_id, platform, url, is_public, icon, display_order)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (user_id, platform.lower(), url, 1 if is_public else 0, icon, max_order + 1))
            
            conn.commit()
            return True
    
    def update_social_link(self, link_id: int, **kwargs) -> bool:
        """Update a social link"""
        allowed_fields = ['url', 'is_active', 'is_public', 'icon', 'display_order']
        
        updates = {}
        for key, value in kwargs.items():
            if key in allowed_fields:
                updates[key] = value
        
        if not updates:
            return False
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            set_clause = ', '.join([f"{key} = ?" for key in updates.keys()])
            values = list(updates.values())
            values.append(link_id)
            
            cursor.execute(f"""
                UPDATE social_links 
                SET {set_clause}, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, values)
            
            conn.commit()
            return cursor.rowcount > 0
    
    def delete_social_link(self, link_id: int) -> bool:
        """Delete a social link"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("DELETE FROM social_links WHERE id = ?", (link_id,))
            conn.commit()
            return cursor.rowcount > 0
    
    def reorder_social_links(self, user_id: int, link_order: List[int]) -> bool:
        """Reorder social links for a user"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            for index, link_id in enumerate(link_order):
                cursor.execute("""
                    UPDATE social_links 
                    SET display_order = ? 
                    WHERE id = ? AND user_id = ?
                """, (index, link_id, user_id))
            
            conn.commit()
            return True
    
    # ==================== USER ACTIVITY LOG METHODS ====================
    
    def log_user_activity(self, user_id: int, action: str, details: str = None,
                        ip_address: str = None, user_agent: str = None) -> bool:
        """Log user activity"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                INSERT INTO user_activity_log 
                (user_id, action, details, ip_address, user_agent)
                VALUES (?, ?, ?, ?, ?)
            """, (user_id, action, details, ip_address, user_agent))
            conn.commit()
            return True
    
    def get_user_activities(self, user_id: int, limit: int = 20) -> List[Dict]:
        """Get user activity log"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, action, details, ip_address, user_agent, created_at
                FROM user_activity_log 
                WHERE user_id = ?
                ORDER BY created_at DESC
                LIMIT ?
            """, (user_id, limit))
            
            rows = cursor.fetchall()
            
            activities = []
            for row in rows:
                if isinstance(row, dict):
                    activities.append(row)
                else:
                    columns = [description[0] for description in cursor.description]
                    activities.append(dict(zip(columns, row)))
            
            return activities
    
    def get_user_activity_stats(self, user_id: int) -> Dict:
        """Get user activity statistics"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Total activities
            cursor.execute("""
                SELECT COUNT(*) as total_activities
                FROM user_activity_log 
                WHERE user_id = ?
            """, (user_id,))
            total = cursor.fetchone()
            
            # Activities by action
            cursor.execute("""
                SELECT action, COUNT(*) as count
                FROM user_activity_log 
                WHERE user_id = ?
                GROUP BY action
                ORDER BY count DESC
            """, (user_id,))
            
            actions = cursor.fetchall()
            
            # Last 7 days activity
            cursor.execute("""
                SELECT DATE(created_at) as date, COUNT(*) as count
                FROM user_activity_log 
                WHERE user_id = ? 
                AND created_at >= DATE('now', '-7 days')
                GROUP BY DATE(created_at)
                ORDER BY date DESC
            """, (user_id,))
            
            weekly = cursor.fetchall()
            
            return {
                'total': total['total_activities'] if total else 0,
                'by_action': [dict(row) if isinstance(row, dict) else dict(zip(['action', 'count'], row)) for row in actions],
                'weekly': [dict(row) if isinstance(row, dict) else dict(zip(['date', 'count'], row)) for row in weekly]
            }

    # ==================== SYSTEM CONFIG METHODS ====================
    
    def get_config(self, key: str, default: Any = None) -> Any:
        """Get system configuration value"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("SELECT value FROM system_config WHERE key = ?", (key,))
            row = cursor.fetchone()
            
            if row:
                value = row['value']
                # Parse boolean
                if value.lower() == 'true':
                    return True
                elif value.lower() == 'false':
                    return False
                return value
            
            return default
    
    def set_config(self, key: str, value: Any, updated_by: int = None) -> bool:
        """Set system configuration value"""
        # Convert boolean to string
        if isinstance(value, bool):
            value = 'true' if value else 'false'
        else:
            value = str(value)
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                INSERT OR REPLACE INTO system_config (key, value, updated_by, updated_at)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            """, (key, value, updated_by))
            
            return True
    


    
    def update_user(self, user_id: int, **kwargs) -> bool:
        """Update user information"""
        # Remove subscription_tier from allowed fields since it doesn't exist
        allowed_fields = ['email', 'full_name', 'role', 'company_id', 'is_active']
        
        updates = []
        values = []
        
        for key, value in kwargs.items():
            if key in allowed_fields:
                updates.append(f"{key} = ?")
                values.append(value)
        
        if not updates:
            return False
        
        values.append(user_id)
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute(f"""
                UPDATE users 
                SET {', '.join(updates)}
                WHERE id = ?
            """, values)
            
            return cursor.rowcount > 0


    
    def delete_user(self, user_id: int, hard_delete: bool = False) -> bool:
        """Delete or deactivate user"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            if hard_delete:
                cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
            else:
                cursor.execute("UPDATE users SET is_active = 0 WHERE id = ?", (user_id,))
            
            return cursor.rowcount > 0
    
    def reset_password(self, user_id: int, new_password: str) -> bool:
        """Reset user password using bcrypt"""
        hashed = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("UPDATE users SET password = ? WHERE id = ?", 
                        (hashed, user_id))
            return cursor.rowcount > 0

    def get_user_by_email(self, email: str) -> Optional[Dict]:
        """Get user by email address"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, username, email, full_name, role, company_id, is_active,
                    created_at, last_login
                FROM users WHERE email = ?
            """, (email,))
            
            row = cursor.fetchone()
            return dict(row) if row else None

    # ==================== QUERY METHODS ====================
    
    def query(self, sql: str, params: tuple = None) -> List[Dict]:
        """Execute SELECT query and return results as list of dicts"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            if params:
                cursor.execute(sql, params)
            else:
                cursor.execute(sql)
            return [dict(row) for row in cursor.fetchall()]
    
    def query_one(self, sql: str, params: tuple = None) -> Optional[Dict]:
        """Execute SELECT query and return single result as dict"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            if params:
                cursor.execute(sql, params)
            else:
                cursor.execute(sql)
            row = cursor.fetchone()
            return dict(row) if row else None
    
    def execute(self, sql: str, params: tuple = None) -> int:
        """Execute INSERT/UPDATE/DELETE query and return rowcount"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            if params:
                cursor.execute(sql, params)
            else:
                cursor.execute(sql)
            return cursor.rowcount
    
    @property
    def last_insert_id(self) -> int:
        """Get last inserted row ID"""
        with self.get_connection() as conn:
            return conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # ==================== SUBSCRIPTION METHODS ====================
    
    def get_user_subscription(self, user_id: int) -> Dict:
        """Get user's subscription details"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # ✅ Try user's own subscription first
            cursor.execute("""
                SELECT 
                    plan as subscription_tier,
                    status,
                    start_date,
                    end_date,
                    analyses_limit as max_projects,
                    analyses_used,
                    max_boq_generations,
                    max_bid_optimizations,
                    can_export_data,
                    can_edit_rates,
                    can_delete_rates,
                    can_create_versions,
                    can_manage_team
                FROM subscriptions 
                WHERE user_id = ? AND status = 'active'
                ORDER BY id DESC
                LIMIT 1
            """, (user_id,))
            
            row = cursor.fetchone()
            if row:
                return dict(row)
            
            # ✅ If no user subscription, check company subscription
            cursor.execute("""
                SELECT 
                    u.company_id,
                    s.plan as subscription_tier,
                    s.status,
                    s.start_date,
                    s.end_date,
                    s.analyses_limit as max_projects,
                    s.analyses_used,
                    s.max_boq_generations,
                    s.max_bid_optimizations,
                    s.can_export_data,
                    s.can_edit_rates,
                    s.can_delete_rates,
                    s.can_create_versions,
                    s.can_manage_team
                FROM users u
                LEFT JOIN subscriptions s ON u.company_id = s.company_id AND s.status = 'active'
                WHERE u.id = ?
                ORDER BY s.id DESC
                LIMIT 1
            """, (user_id,))
            
            row = cursor.fetchone()
            if row and row.get('subscription_tier'):
                return dict(row)
            
            return {
                'subscription_tier': 'free',
                'status': 'active',
                'max_projects': 5,
                'analyses_used': 0,
                'max_boq_generations': 5,
                'max_bid_optimizations': 5,
                'can_export_data': False,
                'can_edit_rates': False,
                'can_delete_rates': False,
                'can_create_versions': False,
                'can_manage_team': False
            }
    def get_company_subscription(self, company_id: int) -> Dict:
        """Get company's subscription details"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # ✅ Get the latest active subscription ordered by id DESC
            cursor.execute("""
                SELECT 
                    s.plan as subscription_tier,
                    s.status,
                    s.start_date,
                    s.end_date,
                    s.analyses_limit as max_projects,
                    s.analyses_used,
                    s.max_boq_generations,
                    s.max_bid_optimizations,
                    s.boq_used,
                    s.bid_optimizations_used,
                    s.can_export_data,
                    s.can_edit_rates,
                    s.can_delete_rates,
                    s.can_create_versions,
                    s.can_manage_team,
                    s.payment_method,
                    s.transaction_id,
                    s.updated_at,
                    sp.max_users,
                    sp.extension_auto_fills,
                    sp.plan_name
                FROM subscriptions s
                LEFT JOIN subscription_plans sp ON s.plan = sp.plan_name
                WHERE s.company_id = ? AND s.status = 'active'
                ORDER BY s.id DESC
                LIMIT 1
            """, (company_id,))
            
            row = cursor.fetchone()
            
            if row:
                result = dict(row)
                # ✅ Debug print to see what's being returned
                print(f"📊 get_company_subscription for company {company_id}: plan={result.get('subscription_tier')}")
                return result
            
            # ✅ If no active subscription, check if there's any subscription
            cursor.execute("""
                SELECT 
                    plan as subscription_tier,
                    status,
                    start_date,
                    end_date,
                    analyses_limit as max_projects,
                    analyses_used
                FROM subscriptions 
                WHERE company_id = ?
                ORDER BY id DESC
                LIMIT 1
            """, (company_id,))
            
            row = cursor.fetchone()
            if row:
                result = dict(row)
                print(f"⚠️ Found inactive subscription for company {company_id}: {result.get('subscription_tier')}")
                return result
            
            print(f"⚠️ No subscription found for company {company_id}")
            return {
                'subscription_tier': 'free',
                'status': 'active',
                'max_projects': 5,
                'max_users': 1,
                'analyses_used': 0,
                'max_boq_generations': 5,
                'max_bid_optimizations': 5,
                'boq_used': 0,
                'bid_optimizations_used': 0,
                'can_export_data': False,
                'can_edit_rates': False,
                'can_delete_rates': False,
                'can_create_versions': False,
                'can_manage_team': False,
                'extension_auto_fills': 5
            }
    
    # database/crud_operations.py - Update update_company_subscription

    def update_company_subscription(self, company_id: int, plan: str, 
                                    duration: str = 'monthly', 
                                    payment_method: str = 'admin', 
                                    transaction_id: str = None) -> bool:
        """Update or create subscription for a company"""
        from datetime import datetime, timedelta
        
        print("\n" + "=" * 60)
        print("📝 update_company_subscription() CALLED")
        print("=" * 60)
        print(f"   company_id: {company_id}")
        print(f"   plan: {plan}")
        print(f"   duration: {duration}")
        print(f"   payment_method: {payment_method}")
        print(f"   transaction_id: {transaction_id}")
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            start_date = datetime.now().date()
            print(f"   start_date: {start_date}")
            
            # Plan limits and features
            plan_limits = {
                'free': {'limit': 5, 'max_boq': 5, 'max_bid': 5},
                'basic': {'limit': 30, 'max_boq': 30, 'max_bid': 30},
                'professional': {'limit': -1, 'max_boq': 100, 'max_bid': 100},
                'enterprise': {'limit': -1, 'max_boq': -1, 'max_bid': -1}
            }
            
            plan_features = {
                'free': {'can_export': 0, 'can_edit_rates': 0, 'can_delete_rates': 0, 
                        'can_create_versions': 0, 'can_manage_team': 0},
                'basic': {'can_export': 1, 'can_edit_rates': 0, 'can_delete_rates': 0,
                        'can_create_versions': 0, 'can_manage_team': 0},
                'professional': {'can_export': 1, 'can_edit_rates': 1, 'can_delete_rates': 0,
                                'can_create_versions': 1, 'can_manage_team': 1},
                'enterprise': {'can_export': 1, 'can_edit_rates': 1, 'can_delete_rates': 1,
                            'can_create_versions': 1, 'can_manage_team': 1}
            }
            
            features = plan_features.get(plan, plan_features['free'])
            limits = plan_limits.get(plan, plan_limits['free'])
            
            if duration == 'monthly':
                end_date = start_date + timedelta(days=30)
            else:
                end_date = start_date + timedelta(days=365)
            print(f"   end_date: {end_date}")
            
            trans_id = transaction_id or f"ADMIN_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            # Check if subscription exists
            cursor.execute("""
                SELECT id, plan FROM subscriptions 
                WHERE company_id = ? AND company_id IS NOT NULL
                ORDER BY id DESC LIMIT 1
            """, (company_id,))
            
            existing = cursor.fetchone()
            
            if existing:
                print(f"   ✅ Found existing subscription ID: {existing['id']}, Current plan: {existing['plan']}")
                
                # UPDATE existing subscription
                cursor.execute("""
                    UPDATE subscriptions 
                    SET plan = ?, 
                        status = 'active', 
                        start_date = ?, 
                        end_date = ?,
                        analyses_limit = ?,
                        max_boq_generations = ?,
                        max_bid_optimizations = ?,
                        can_export_data = ?,
                        can_edit_rates = ?,
                        can_delete_rates = ?,
                        can_create_versions = ?,
                        can_manage_team = ?,
                        payment_method = ?,
                        transaction_id = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (
                    plan,
                    start_date,
                    end_date,
                    limits['limit'],
                    limits['max_boq'],
                    limits['max_bid'],
                    features['can_export'],
                    features['can_edit_rates'],
                    features['can_delete_rates'],
                    features['can_create_versions'],
                    features['can_manage_team'],
                    payment_method,
                    trans_id,
                    existing['id']
                ))
                print(f"   ✅ Updated subscription ID {existing['id']} to {plan}")
                
            else:
                print(f"   ⚠️ No existing subscription found, creating new one")
                
                # INSERT new subscription
                cursor.execute("""
                    INSERT INTO subscriptions (
                        company_id, plan, status, start_date, end_date,
                        analyses_limit, analyses_used,
                        max_boq_generations, max_bid_optimizations,
                        can_export_data, can_edit_rates, can_delete_rates,
                        can_create_versions, can_manage_team,
                        payment_method, transaction_id,
                        created_at, updated_at
                    ) VALUES (?, ?, 'active', ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """, (
                    company_id,
                    plan,
                    start_date,
                    end_date,
                    limits['limit'],
                    limits['max_boq'],
                    limits['max_bid'],
                    features['can_export'],
                    features['can_edit_rates'],
                    features['can_delete_rates'],
                    features['can_create_versions'],
                    features['can_manage_team'],
                    payment_method,
                    trans_id
                ))
                print(f"   ✅ Created new subscription for company {company_id} with plan {plan}")
            
            # Verify the update worked
            print("\n   🔍 Verifying update...")
            cursor.execute("""
                SELECT id, plan, status, start_date, end_date 
                FROM subscriptions 
                WHERE company_id = ? AND company_id IS NOT NULL
                ORDER BY id DESC LIMIT 1
            """, (company_id,))
            verify = cursor.fetchone()
            if verify:
                print(f"   ✅ Verification: Subscription ID {verify['id']} now has plan {verify['plan']}")
                return verify['plan'] == plan
            
            print("   ✅ Update completed successfully")
            return True
    
    def cancel_user_subscription(self, user_id: int) -> bool:
        """Cancel user subscription (set to free)"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                UPDATE subscriptions 
                SET plan = 'free',
                    status = 'active',
                    analyses_limit = 5,
                    max_boq_generations = 5,
                    max_bid_optimizations = 5,
                    can_export_data = 0,
                    can_edit_rates = 0,
                    can_delete_rates = 0,
                    can_create_versions = 0,
                    can_manage_team = 0,
                    updated_at = CURRENT_TIMESTAMP
                WHERE user_id = ? AND user_id IS NOT NULL
            """, (user_id,))
            
            return True
    def cancel_company_subscription(self, company_id: int) -> bool:
        """Cancel subscription for a company (set to free)"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                UPDATE subscriptions 
                SET plan = 'free',
                    status = 'active',
                    analyses_limit = 5,
                    max_boq_generations = 5,
                    max_bid_optimizations = 5,
                    can_export_data = 0,
                    can_edit_rates = 0,
                    can_delete_rates = 0,
                    can_create_versions = 0,
                    can_manage_team = 0,
                    updated_at = CURRENT_TIMESTAMP
                WHERE company_id = ?
            """, (company_id,))
            
            print(f"✅ Cancelled subscription for company {company_id}")
            return True

    

    def update_user_subscription(self, user_id: int, plan: str, 
                             duration: str = 'monthly',
                             payment_method: str = 'admin',
                             transaction_id: str = None) -> bool:
        """Update or create subscription for a user"""
        from datetime import datetime, timedelta
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            start_date = datetime.now().date()
            
            # Plan limits
            plan_limits = {
                'free': {'limit': 5, 'max_boq': 5, 'max_bid': 5},
                'basic': {'limit': 30, 'max_boq': 30, 'max_bid': 30},
                'professional': {'limit': -1, 'max_boq': 100, 'max_bid': 100},
                'enterprise': {'limit': -1, 'max_boq': -1, 'max_bid': -1}
            }
            
            plan_features = {
                'free': {'can_export': 0, 'can_edit_rates': 0, 'can_delete_rates': 0, 
                        'can_create_versions': 0, 'can_manage_team': 0},
                'basic': {'can_export': 1, 'can_edit_rates': 0, 'can_delete_rates': 0,
                        'can_create_versions': 0, 'can_manage_team': 0},
                'professional': {'can_export': 1, 'can_edit_rates': 1, 'can_delete_rates': 0,
                                'can_create_versions': 1, 'can_manage_team': 1},
                'enterprise': {'can_export': 1, 'can_edit_rates': 1, 'can_delete_rates': 1,
                            'can_create_versions': 1, 'can_manage_team': 1}
            }
            
            features = plan_features.get(plan, plan_features['free'])
            limits = plan_limits.get(plan, plan_limits['free'])
            
            if duration == 'monthly':
                end_date = start_date + timedelta(days=30)
            else:
                end_date = start_date + timedelta(days=365)
            
            trans_id = transaction_id or f"USER_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            # Check if subscription exists
            cursor.execute('SELECT id FROM subscriptions WHERE user_id = ? AND user_id IS NOT NULL', (user_id,))
            existing = cursor.fetchone()
            
            if existing:
                cursor.execute("""
                    UPDATE subscriptions 
                    SET plan = ?, 
                        status = 'active', 
                        start_date = ?, 
                        end_date = ?,
                        analyses_limit = ?,
                        max_boq_generations = ?,
                        max_bid_optimizations = ?,
                        can_export_data = ?,
                        can_edit_rates = ?,
                        can_delete_rates = ?,
                        can_create_versions = ?,
                        can_manage_team = ?,
                        payment_method = ?,
                        transaction_id = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE user_id = ?
                """, (
                    plan, start_date, end_date,
                    limits['limit'], limits['max_boq'], limits['max_bid'],
                    features['can_export'], features['can_edit_rates'],
                    features['can_delete_rates'], features['can_create_versions'],
                    features['can_manage_team'],
                    payment_method, trans_id, user_id
                ))
            else:
                cursor.execute("""
                    INSERT INTO subscriptions (
                        user_id, plan, status, start_date, end_date,
                        analyses_limit, analyses_used,
                        max_boq_generations, max_bid_optimizations,
                        can_export_data, can_edit_rates, can_delete_rates,
                        can_create_versions, can_manage_team,
                        payment_method, transaction_id,
                        created_at, updated_at
                    ) VALUES (?, ?, 'active', ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """, (
                    user_id, plan, start_date, end_date,
                    limits['limit'], limits['max_boq'], limits['max_bid'],
                    features['can_export'], features['can_edit_rates'],
                    features['can_delete_rates'], features['can_create_versions'],
                    features['can_manage_team'],
                    payment_method, trans_id
                ))
            
            return True
    
    
    
    def increment_api_usage(self, user_id: int) -> bool:
        """Increment API call count for user"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                UPDATE users 
                SET api_calls_used = api_calls_used + 1
                WHERE id = ? AND api_calls_used < api_calls_limit
            """, (user_id,))
            
            return cursor.rowcount > 0
    
    # ==================== TENDER ANALYSIS METHODS ====================
    
    def save_analysis(self, user_id: int, tender_id: str, analysis_data: Dict,
                     confidence_score: float = None) -> int:
        """Save tender analysis results"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                INSERT INTO tender_analyses 
                (user_id, tender_id, analysis_data, confidence_score, created_at)
                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (user_id, tender_id, json.dumps(analysis_data), confidence_score))
            
            return cursor.lastrowid
    
    # database/crud_operations.py - Fixed get_user_analyses

    def get_user_analyses(self, user_id: int, company_id: int = None, role: str = 'user', limit: int = 50):
        """Get user's tender analyses with role-based filtering"""
        
        # Use the correct context manager pattern
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            print(f"🔍 get_user_analyses called with: user_id={user_id}, company_id={company_id}, role={role}")
            
            # System admin can see all analyses across all companies
            if role in ['admin', 'system_admin']:
                cursor.execute('''
                    SELECT * FROM tender_analyses 
                    ORDER BY analysis_date DESC LIMIT ?
                ''', (limit,))
            # Company admin and manager can see all analyses for their company
            elif role in ['company_admin', 'manager'] and company_id:
                cursor.execute('''
                    SELECT * FROM tender_analyses 
                    WHERE company_id = ? 
                    ORDER BY analysis_date DESC LIMIT ?
                ''', (company_id, limit))
            # Regular users can only see their own analyses
            else:
                cursor.execute('''
                    SELECT * FROM tender_analyses 
                    WHERE user_id = ? AND company_id = ?
                    ORDER BY analysis_date DESC LIMIT ?
                ''', (user_id, company_id, limit))
            
            # Get column names
            columns = [description[0] for description in cursor.description]
            data = cursor.fetchall()
            
            print(f"🔍 Found {len(data)} analyses for role={role}, company_id={company_id}")
            
            if data:
                
                df = pd.DataFrame(data, columns=columns)
                
                # Parse competitor_bids JSON if present
                if 'competitor_bids' in df.columns:
                    import json
                    def parse_competitor_bids(value):
                        if value is None:
                            return []
                        if isinstance(value, str):
                            try:
                                if value and value != 'null':
                                    return json.loads(value)
                            except:
                                pass
                        return []
                    df['competitor_bids'] = df['competitor_bids'].apply(parse_competitor_bids)
                
                return df
            return pd.DataFrame()
    
    def get_company_stats(self, company_id: int) -> Dict:
        """Get analysis statistics for a company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Get user IDs for this company
            cursor.execute("SELECT id FROM users WHERE company_id = ?", (company_id,))
            user_ids = [row['id'] for row in cursor.fetchall()]
            
            stats = {
                'total_analyses': 0,
                'avg_confidence': 0,
                'total_users': len(user_ids),
                'win_rate': 0,
                'total_bids': 0,
                'total_wins': 0
            }
            
            if not user_ids:
                return stats
            
            placeholders = ','.join('?' * len(user_ids))
            
            # Get analysis stats
            cursor.execute(f"""
                SELECT COUNT(*) as total, AVG(confidence_score) as avg_conf,
                    SUM(CASE WHEN bid_status = 'won' THEN 1 ELSE 0 END) as wins,
                    COUNT(CASE WHEN bid_status IS NOT NULL THEN 1 END) as total_bids
                FROM tender_analyses
                WHERE user_id IN ({placeholders})
            """, user_ids)
            
            row = cursor.fetchone()
            
            if row:
                stats['total_analyses'] = row['total'] or 0
                stats['avg_confidence'] = row['avg_conf'] or 0
                stats['total_bids'] = row['total_bids'] or 0
                stats['total_wins'] = row['wins'] or 0
                
                if stats['total_bids'] > 0:
                    stats['win_rate'] = (stats['total_wins'] / stats['total_bids']) * 100
            
            return stats

    
    def toggle_favorite_analysis(self, analysis_id: int, user_id: int) -> bool:
        """Toggle favorite status of an analysis"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                UPDATE tender_analyses
                SET is_favorite = NOT is_favorite
                WHERE id = ? AND user_id = ?
            """, (analysis_id, user_id))
            
            return cursor.rowcount > 0
    
    # ==================== HISTORICAL TENDER METHODS ====================
    
    def save_historical_tender(self, tender_data: Dict) -> int:
        """Save historical tender data"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                INSERT OR REPLACE INTO historical_tenders
                (tender_id, title, procuring_entity, estimated_amount, awarded_amount,
                 award_date, contractor, contract_period, category)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                tender_data.get('tender_id'),
                tender_data.get('title'),
                tender_data.get('procuring_entity'),
                tender_data.get('estimated_amount'),
                tender_data.get('awarded_amount'),
                tender_data.get('award_date'),
                tender_data.get('contractor'),
                tender_data.get('contract_period'),
                tender_data.get('category')
            ))
            
            return cursor.lastrowid
    
    def get_historical_tenders(self, filters: Dict = None, limit: int = 100) -> List[Dict]:
        """Get historical tenders with optional filters"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            query = "SELECT * FROM historical_tenders WHERE 1=1"
            values = []
            
            if filters:
                if 'procuring_entity' in filters:
                    query += " AND procuring_entity LIKE ?"
                    values.append(f"%{filters['procuring_entity']}%")
                
                if 'category' in filters:
                    query += " AND category = ?"
                    values.append(filters['category'])
                
                if 'min_amount' in filters:
                    query += " AND awarded_amount >= ?"
                    values.append(filters['min_amount'])
                
                if 'max_amount' in filters:
                    query += " AND awarded_amount <= ?"
                    values.append(filters['max_amount'])
                
                if 'from_date' in filters:
                    query += " AND award_date >= ?"
                    values.append(filters['from_date'])
                
                if 'to_date' in filters:
                    query += " AND award_date <= ?"
                    values.append(filters['to_date'])
            
            query += " ORDER BY award_date DESC LIMIT ?"
            values.append(limit)
            
            cursor.execute(query, values)
            return [dict(row) for row in cursor.fetchall()]
    
    def get_winning_statistics_by_contractor(self, contractor_name: str = None) -> Dict:
        """Get winning statistics for a specific contractor"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            if contractor_name:
                cursor.execute("""
                    SELECT winning_competitor as contractor, COUNT(*) as wins, 
                        AVG(awarded_price) as avg_amount,
                        SUM(awarded_price) as total_amount,
                        MIN(award_date) as first_win,
                        MAX(award_date) as last_win
                    FROM historical_tenders
                    WHERE winning_competitor = ?
                    GROUP BY winning_competitor
                """, (contractor_name,))
            else:
                cursor.execute("""
                    SELECT winning_competitor as contractor, COUNT(*) as wins, 
                        AVG(awarded_price) as avg_amount,
                        SUM(awarded_price) as total_amount,
                        MIN(award_date) as first_win,
                        MAX(award_date) as last_win
                    FROM historical_tenders
                    WHERE winning_competitor IS NOT NULL
                    GROUP BY winning_competitor
                    ORDER BY wins DESC
                    LIMIT 50
                """)
            
            result = cursor.fetchone()
            return dict(result) if result else {}

    def get_winning_statistics(self, company_id: int, procurement_type: str = None) -> Dict:
        """Get winning statistics for analysis"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            query = '''
            SELECT 
                COUNT(*) as total_tenders,
                SUM(CASE WHEN winning_company_type = 'Our Company' THEN 1 ELSE 0 END) as our_wins,
                SUM(CASE WHEN winning_company_type = 'Competitor' THEN 1 ELSE 0 END) as competitor_wins,
                SUM(CASE WHEN winning_company_type = 'Unknown' THEN 1 ELSE 0 END) as unknown_wins,
                AVG(CASE WHEN winning_company_type = 'Our Company' THEN awarded_price ELSE NULL END) as avg_our_winning_price,
                AVG(CASE WHEN winning_company_type = 'Competitor' THEN awarded_price ELSE NULL END) as avg_competitor_winning_price,
                AVG(official_estimate) as avg_estimate,
                MIN(CASE WHEN winning_company_type = 'Our Company' THEN awarded_price ELSE NULL END) as min_our_winning_price,
                MAX(CASE WHEN winning_company_type = 'Our Company' THEN awarded_price ELSE NULL END) as max_our_winning_price,
                MIN(CASE WHEN winning_company_type = 'Competitor' THEN awarded_price ELSE NULL END) as min_competitor_winning_price,
                MAX(CASE WHEN winning_company_type = 'Competitor' THEN awarded_price ELSE NULL END) as max_competitor_winning_price
            FROM historical_tenders 
            WHERE company_id = ?
            '''
            params = [company_id]
            
            if procurement_type:
                query += " AND procurement_type = ?"
                params.append(procurement_type)
            
            cursor.execute(query, params)
            result = cursor.fetchone()
            
            if result:
                return {
                    'total_tenders': result['total_tenders'] or 0,
                    'our_wins': result['our_wins'] or 0,
                    'competitor_wins': result['competitor_wins'] or 0,
                    'unknown_wins': result['unknown_wins'] or 0,
                    'our_win_rate': (result['our_wins'] / result['total_tenders'] * 100) if result['total_tenders'] > 0 else 0,
                    'avg_our_winning_price': result['avg_our_winning_price'] or 0,
                    'avg_competitor_winning_price': result['avg_competitor_winning_price'] or 0,
                    'avg_estimate': result['avg_estimate'] or 0,
                    'min_our_winning_price': result['min_our_winning_price'] or 0,
                    'max_our_winning_price': result['max_our_winning_price'] or 0,
                    'min_competitor_winning_price': result['min_competitor_winning_price'] or 0,
                    'max_competitor_winning_price': result['max_competitor_winning_price'] or 0
                }
            return None

    
    # ==================== KNOWLEDGE REPOSITORY METHODS ====================
    
    # Personnel Management
    def add_personnel(self, company_id: int, name: str, designation: str,
                     expertise_areas: str, years_experience: int,
                     certifications: str = None, cv_path: str = None) -> int:
        """Add personnel to knowledge repository"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                INSERT INTO personnel (company_id, name, designation, expertise_areas,
                                     years_experience, certifications, cv_path)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (company_id, name, designation, expertise_areas,
                  years_experience, certifications, cv_path))
            
            return cursor.lastrowid
    
    def get_personnel(self, company_id: int, expertise: str = None) -> List[Dict]:
        """Get personnel list with optional expertise filter"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            query = "SELECT * FROM personnel WHERE company_id = ?"
            values = [company_id]
            
            if expertise:
                query += " AND expertise_areas LIKE ?"
                values.append(f"%{expertise}%")
            
            query += " ORDER BY years_experience DESC"
            
            cursor.execute(query, values)
            return [dict(row) for row in cursor.fetchall()]
    
    def update_personnel(self, personnel_id: int, **kwargs) -> bool:
        """Update personnel information"""
        allowed_fields = ['name', 'designation', 'expertise_areas', 
                         'years_experience', 'certifications', 'cv_path']
        
        updates = []
        values = []
        
        for key, value in kwargs.items():
            if key in allowed_fields:
                updates.append(f"{key} = ?")
                values.append(value)
        
        if not updates:
            return False
        
        values.append(personnel_id)
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute(f"""
                UPDATE personnel 
                SET {', '.join(updates)}
                WHERE id = ?
            """, values)
            
            return cursor.rowcount > 0
    
    # Equipment Management
    def add_equipment(self, company_id: int, name: str, type: str,
                     specifications: str, quantity: int, condition: str,
                     ownership: str, value: float = None) -> int:
        """Add equipment to knowledge repository"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                INSERT INTO equipment (company_id, name, type, specifications,
                                     quantity, condition, ownership, value)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (company_id, name, type, specifications,
                  quantity, condition, ownership, value))
            
            return cursor.lastrowid
    
    def get_equipment(self, company_id: int, equipment_type: str = None) -> List[Dict]:
        """Get equipment list with optional type filter"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            query = "SELECT * FROM equipment WHERE company_id = ?"
            values = [company_id]
            
            if equipment_type:
                query += " AND type = ?"
                values.append(equipment_type)
            
            query += " ORDER BY name"
            
            cursor.execute(query, values)
            return [dict(row) for row in cursor.fetchall()]
    
    # Experience Management
    def add_experience(self, company_id: int, project_name: str,
                      client_name: str, contract_value: float,
                      completion_date: str, project_type: str,
                      description: str = None) -> int:
        """Add project experience to knowledge repository"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                INSERT INTO experiences (company_id, project_name, client_name,
                                       contract_value, completion_date,
                                       project_type, description)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (company_id, project_name, client_name,
                  contract_value, completion_date, project_type, description))
            
            return cursor.lastrowid
    
    def get_experiences(self, company_id: int, project_type: str = None,
                       min_value: float = None) -> List[Dict]:
        """Get project experiences with filters"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            query = "SELECT * FROM experiences WHERE company_id = ?"
            values = [company_id]
            
            if project_type:
                query += " AND project_type = ?"
                values.append(project_type)
            
            if min_value:
                query += " AND contract_value >= ?"
                values.append(min_value)
            
            query += " ORDER BY completion_date DESC"
            
            cursor.execute(query, values)
            return [dict(row) for row in cursor.fetchall()]
    
    # Financial Management
    def add_financial_capacity(self, company_id: int, fiscal_year: str,
                              annual_turnover: float, net_profit: float,
                              liquid_assets: float, credit_limit: float) -> int:
        """Add financial capacity record"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                INSERT INTO financial_records (company_id, fiscal_year, annual_turnover,
                                             net_profit, liquid_assets, credit_limit)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (company_id, fiscal_year, annual_turnover,
                  net_profit, liquid_assets, credit_limit))
            
            return cursor.lastrowid
    
    def get_financial_records(self, company_id: int, 
                             fiscal_year: str = None) -> List[Dict]:
        """Get financial records for a company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            query = "SELECT * FROM financial_records WHERE company_id = ?"
            values = [company_id]
            
            if fiscal_year:
                query += " AND fiscal_year = ?"
                values.append(fiscal_year)
            
            query += " ORDER BY fiscal_year DESC"
            
            cursor.execute(query, values)
            return [dict(row) for row in cursor.fetchall()]
    
    # ==================== SEMANTIC SEARCH METHODS ====================
    
    def store_embedding(self, content_type: str, content_id: int,
                       content_text: str, embedding_blob: bytes) -> bool:
        """Store embedding vector for semantic search"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                INSERT OR REPLACE INTO embeddings
                (content_type, content_id, content_text, embedding, created_at)
                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (content_type, content_id, content_text, embedding_blob))
            
            return cursor.rowcount > 0
    
    def semantic_search(self, query_embedding: bytes, content_type: str = None,
                       limit: int = 10) -> List[Dict]:
        """
        Perform semantic search using cosine similarity
        Note: This is a simplified version. In production, use vector similarity
        """
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # For SQLite without vector extension, we'll retrieve candidates
            # and compute similarity in Python (not efficient for large datasets)
            query = "SELECT content_type, content_id, content_text FROM embeddings"
            values = []
            
            if content_type:
                query += " WHERE content_type = ?"
                values.append(content_type)
            
            cursor.execute(query, values)
            
            # Here you would implement cosine similarity calculation
            # This is a placeholder - implement actual similarity search
            results = []
            for row in cursor.fetchall():
                results.append({
                    'content_type': row['content_type'],
                    'content_id': row['content_id'],
                    'content_text': row['content_text'],
                    'similarity_score': 0.0  # Placeholder
                })
            
            # Sort by similarity and limit
            results.sort(key=lambda x: x['similarity_score'], reverse=True)
            return results[:limit]
    
    def hybrid_search(self, query_text: str, query_embedding: bytes = None,
                     content_type: str = None, limit: int = 10) -> List[Dict]:
        """Combine keyword and semantic search"""
        # Keyword search results
        keyword_results = self.keyword_search(query_text, content_type, limit*2)
        
        # Semantic search results if embedding provided
        semantic_results = []
        if query_embedding:
            semantic_results = self.semantic_search(query_embedding, content_type, limit*2)
        
        # Merge and deduplicate results (simplified)
        all_results = keyword_results + semantic_results
        seen = set()
        unique_results = []
        
        for result in all_results:
            key = f"{result.get('content_type')}_{result.get('content_id')}"
            if key not in seen:
                seen.add(key)
                unique_results.append(result)
        
        return unique_results[:limit]
    
    def keyword_search(self, query: str, content_type: str = None,
                      limit: int = 10) -> List[Dict]:
        """Perform keyword-based search across knowledge repository"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            results = []
            
            # Search in personnel
            if not content_type or content_type == 'personnel':
                cursor.execute("""
                    SELECT 'personnel' as content_type, id as content_id,
                           name as title, expertise_areas as content,
                           designation as extra_info
                    FROM personnel
                    WHERE name LIKE ? OR expertise_areas LIKE ? OR designation LIKE ?
                    LIMIT ?
                """, (f"%{query}%", f"%{query}%", f"%{query}%", limit))
                results.extend([dict(row) for row in cursor.fetchall()])
            
            # Search in equipment
            if not content_type or content_type == 'equipment':
                cursor.execute("""
                    SELECT 'equipment' as content_type, id as content_id,
                           name as title, specifications as content,
                           type as extra_info
                    FROM equipment
                    WHERE name LIKE ? OR specifications LIKE ? OR type LIKE ?
                    LIMIT ?
                """, (f"%{query}%", f"%{query}%", f"%{query}%", limit))
                results.extend([dict(row) for row in cursor.fetchall()])
            
            # Search in experiences
            if not content_type or content_type == 'experience':
                cursor.execute("""
                    SELECT 'experience' as content_type, id as content_id,
                           project_name as title, description as content,
                           client_name as extra_info
                    FROM experiences
                    WHERE project_name LIKE ? OR description LIKE ? OR client_name LIKE ?
                    LIMIT ?
                """, (f"%{query}%", f"%{query}%", f"%{query}%", limit))
                results.extend([dict(row) for row in cursor.fetchall()])
            
            return results[:limit]
    
    # ==================== AUTO-FILL METHODS ====================
    
    def get_auto_fill_data(self, company_id: int, field_type: str) -> List[Dict]:
        """Get auto-fill suggestions based on field type"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            if field_type == 'personnel':
                cursor.execute("""
                    SELECT id, name, designation, expertise_areas
                    FROM personnel
                    WHERE company_id = ?
                    ORDER BY years_experience DESC
                """, (company_id,))
                
            elif field_type == 'equipment':
                cursor.execute("""
                    SELECT id, name, type, specifications
                    FROM equipment
                    WHERE company_id = ?
                    ORDER BY name
                """, (company_id,))
                
            elif field_type == 'experience':
                cursor.execute("""
                    SELECT id, project_name, client_name, contract_value, project_type
                    FROM experiences
                    WHERE company_id = ?
                    ORDER BY completion_date DESC
                    LIMIT 20
                """, (company_id,))
                
            elif field_type == 'financial':
                cursor.execute("""
                    SELECT fiscal_year, annual_turnover, net_profit, liquid_assets
                    FROM financial_records
                    WHERE company_id = ?
                    ORDER BY fiscal_year DESC
                    LIMIT 5
                """, (company_id,))
            else:
                return []
            
            return [dict(row) for row in cursor.fetchall()]
    
    def search_knowledge_base(self, company_id: int, search_term: str,
                             category: str = None) -> List[Dict]:
        """Search across knowledge repository for auto-fill"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            results = []
            
            if not category or category == 'personnel':
                cursor.execute("""
                    SELECT 'personnel' as source, id, name, designation as info
                    FROM personnel
                    WHERE company_id = ? AND (name LIKE ? OR expertise_areas LIKE ?)
                    LIMIT 10
                """, (company_id, f"%{search_term}%", f"%{search_term}%"))
                results.extend([dict(row) for row in cursor.fetchall()])
            
            if not category or category == 'equipment':
                cursor.execute("""
                    SELECT 'equipment' as source, id, name, type as info
                    FROM equipment
                    WHERE company_id = ? AND (name LIKE ? OR type LIKE ?)
                    LIMIT 10
                """, (company_id, f"%{search_term}%", f"%{search_term}%"))
                results.extend([dict(row) for row in cursor.fetchall()])
            
            if not category or category == 'experience':
                cursor.execute("""
                    SELECT 'experience' as source, id, project_name as name,
                           client_name as info
                    FROM experiences
                    WHERE company_id = ? AND (project_name LIKE ? OR client_name LIKE ?)
                    LIMIT 10
                """, (company_id, f"%{search_term}%", f"%{search_term}%"))
                results.extend([dict(row) for row in cursor.fetchall()])
            
            return results
    
    # ==================== EXTENSION TRACKING METHODS ====================
    
    def get_extension_fill_usage(self, user_id: int, days: int = 30) -> Dict:
        """Get extension auto-fill usage statistics"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT field_type, COUNT(*) as usage_count,
                       MAX(used_at) as last_used
                FROM extension_fill_log
                WHERE user_id = ? AND used_at >= DATE('now', ?)
                GROUP BY field_type
            """, (user_id, f"-{days} days"))
            
            usage = {row['field_type']: row['usage_count'] for row in cursor.fetchall()}
            return usage
    
    def log_extension_fill(self, user_id: int, field_type: str,
                          source_type: str, source_id: int) -> bool:
        """Log extension auto-fill usage"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                INSERT INTO extension_fill_log (user_id, field_type, source_type, source_id)
                VALUES (?, ?, ?, ?)
            """, (user_id, field_type, source_type, source_id))
            
            return cursor.rowcount > 0
    
    # ==================== SCENARIO METHODS ====================
    
    def save_scenario(self, user_id: int, name: str, scenario_data: Dict,
                     is_shared: bool = False) -> int:
        """Save a tender scenario for future use"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            share_token = secrets.token_urlsafe(16) if is_shared else None
            
            cursor.execute("""
                INSERT INTO scenarios (user_id, name, scenario_data, is_shared, share_token)
                VALUES (?, ?, ?, ?, ?)
            """, (user_id, name, json.dumps(scenario_data), is_shared, share_token))
            
            return cursor.lastrowid
    
    def get_user_scenarios(self, user_id: int, include_shared: bool = False) -> List[Dict]:
        """Get scenarios for a user"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            if include_shared:
                cursor.execute("""
                    SELECT id, user_id, name, scenario_data, is_favorite,
                           is_shared, share_token, created_at, updated_at
                    FROM scenarios
                    WHERE user_id = ? OR is_shared = 1
                    ORDER BY is_favorite DESC, updated_at DESC
                """, (user_id,))
            else:
                cursor.execute("""
                    SELECT id, user_id, name, scenario_data, is_favorite,
                           is_shared, share_token, created_at, updated_at
                    FROM scenarios
                    WHERE user_id = ?
                    ORDER BY is_favorite DESC, updated_at DESC
                """, (user_id,))
            
            scenarios = []
            for row in cursor.fetchall():
                scenario = dict(row)
                scenario['scenario_data'] = json.loads(scenario['scenario_data'])
                scenarios.append(scenario)
            
            return scenarios
    
    def get_scenario_by_id(self, scenario_id: int, user_id: int = None) -> Optional[Dict]:
        """Get scenario by ID with optional user ownership check"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            if user_id:
                cursor.execute("""
                    SELECT id, user_id, name, scenario_data, is_favorite,
                           is_shared, share_token, created_at, updated_at
                    FROM scenarios
                    WHERE id = ? AND (user_id = ? OR is_shared = 1)
                """, (scenario_id, user_id))
            else:
                cursor.execute("""
                    SELECT id, user_id, name, scenario_data, is_favorite,
                           is_shared, share_token, created_at, updated_at
                    FROM scenarios
                    WHERE id = ?
                """, (scenario_id,))
            
            row = cursor.fetchone()
            if row:
                scenario = dict(row)
                scenario['scenario_data'] = json.loads(scenario['scenario_data'])
                return scenario
            return None
    
    def delete_scenario(self, scenario_id: int, user_id: int) -> bool:
        """Delete a scenario"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                DELETE FROM scenarios
                WHERE id = ? AND user_id = ?
            """, (scenario_id, user_id))
            
            return cursor.rowcount > 0
    
    def toggle_favorite_scenario(self, scenario_id: int, user_id: int) -> bool:
        """Toggle favorite status of a scenario"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                UPDATE scenarios
                SET is_favorite = NOT is_favorite
                WHERE id = ? AND user_id = ?
            """, (scenario_id, user_id))
            
            return cursor.rowcount > 0
    
    def update_scenario_name(self, scenario_id: int, user_id: int, new_name: str) -> bool:
        """Update scenario name"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                UPDATE scenarios
                SET name = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND user_id = ?
            """, (new_name, scenario_id, user_id))
            
            return cursor.rowcount > 0
    
    def share_scenario(self, scenario_id: int, user_id: int) -> Optional[str]:
        """Share a scenario and return share token"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            share_token = secrets.token_urlsafe(16)
            
            cursor.execute("""
                UPDATE scenarios
                SET is_shared = 1, share_token = ?
                WHERE id = ? AND user_id = ?
            """, (share_token, scenario_id, user_id))
            
            if cursor.rowcount > 0:
                return share_token
            return None
    
    def get_shared_scenario_by_token(self, share_token: str) -> Optional[Dict]:
        """Get shared scenario by token"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, user_id, name, scenario_data, created_at
                FROM scenarios
                WHERE share_token = ? AND is_shared = 1
            """, (share_token,))
            
            row = cursor.fetchone()
            if row:
                scenario = dict(row)
                scenario['scenario_data'] = json.loads(scenario['scenario_data'])
                return scenario
            return None
    
    # ==================== COMPANY MANAGEMENT METHODS ====================
    def validate_mobile_number(self, mobile: str) -> bool:
        """Validate Bangladeshi mobile number"""
        # Bangladesh mobile: 01XXXXXXXXX (11 digits)
        pattern = r'^01[3-9]\d{8}$'
        return bool(re.match(pattern, mobile))

    # database/crud_operations.py - Corrected for your actual companies table

    def get_all_companies_filtered(self, search: str = "", status: int = None, limit: int = 20, offset: int = 0):
        """Get all companies with pagination and filtering"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Build WHERE clause
            where_clauses = []
            params = []
            
            if search:
                where_clauses.append("(company_name LIKE ? OR email LIKE ? OR phone LIKE ?)")
                params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
            
            if status is not None:
                where_clauses.append("is_active = ?")
                params.append(1 if status else 0)
            
            where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
            
            # Get total count
            cursor.execute(f"SELECT COUNT(*) FROM companies WHERE {where_sql}", params)
            row = cursor.fetchone()
            if row:
                if hasattr(row, 'keys'):  # It's a dict
                    total = list(row.values())[0] if row else 0
                else:  # It's a tuple
                    total = row[0] if row else 0
            else:
                total = 0

            # Get paginated companies
            query = f"""
                SELECT id, company_name, email, phone, division, district, address,
                    registration_number, vat_number, created_at, is_active,
                    status, is_individual, mobile_number
                FROM companies
                WHERE {where_sql}
                ORDER BY created_at DESC
                LIMIT ? OFFSET ?
            """
            params.extend([limit, offset])
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            companies = []
            for row in rows:
                companies.append({
                    'id': row['id'],
                    'company_name': row['company_name'],
                    'email': row['email'],
                    'phone': row['phone'],
                    'division': row['division'],
                    'district': row['district'],
                    'address': row['address'],
                    'registration_number': row['registration_number'],
                    'vat_number': row['vat_number'],
                    'created_at': row['created_at'],
                    'is_active': row['is_active'],
                    'status': row.get('status', 'active'),
                    'is_individual': row.get('is_individual', 0),
                    'mobile_number': row.get('mobile_number', '')
                })
            
            return companies, total
    
    def create_company(self, company_data: Dict) -> tuple:
        """
        Create a new company with unique mobile number
        
        Args:
            company_data: Dict with keys: company_name, email, phone, mobile_number,
                        division, district, address, registration_number, vat_number
        
        Returns:
            (success: bool, message: str or company_id: int)
        """
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Validate required fields
            if not company_data.get('company_name'):
                return False, "Company name is required"
            
            # Validate mobile number if provided
            mobile = None
            if company_data.get('mobile_number'):
                mobile = self.normalize_mobile(company_data['mobile_number'])
                if not self.validate_bangladesh_mobile(mobile):
                    return False, f"Invalid mobile number: {mobile}"
                
                # Check if mobile already exists
                cursor.execute("SELECT id FROM companies WHERE mobile_number = ?", (mobile,))
                if cursor.fetchone():
                    return False, f"Mobile number {mobile} is already registered"
            
            # Check if company name already exists
            cursor.execute("SELECT id FROM companies WHERE company_name = ?", (company_data['company_name'],))
            if cursor.fetchone():
                return False, f"Company name '{company_data['company_name']}' already exists"
            
            # Insert company
            cursor.execute("""
                INSERT INTO companies (
                    company_name, email, phone, mobile_number, division, district,
                    address, registration_number, vat_number, is_active, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, CURRENT_TIMESTAMP)
            """, (
                company_data['company_name'],
                company_data.get('email', ''),
                company_data.get('phone', ''),
                mobile,
                company_data.get('division', ''),
                company_data.get('district', ''),
                company_data.get('address', ''),
                company_data.get('registration_number', ''),
                company_data.get('vat_number', '')
            ))
            
            company_id = cursor.lastrowid
            return True, company_id

    
    def get_company_by_id(self, company_id: int) -> Optional[Dict]:
        """Get company by ID"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, company_name, mobile_number, email, address, 
                    registration_number, phone, division, district, 
                    is_active, created_at
                FROM companies 
                WHERE id = ?
            """, (company_id,))
            
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_company_by_name(self, company_name: str) -> Optional[Dict]:
        """Get company by name"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, company_name, email, phone, mobile_number, division, district,
                    address, registration_number, vat_number, created_at, is_active
                FROM companies 
                WHERE company_name = ?
            """, (company_name,))
            
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_company_by_mobile(self, mobile_number: str) -> Optional[Dict]:
        """Get company by mobile number"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, company_name, mobile_number, email, address, 
                    registration_number, phone, division, district, 
                    is_active, created_at
                FROM companies 
                WHERE mobile_number = ?
            """, (mobile_number,))
            
            row = cursor.fetchone()
            return dict(row) if row else None
    
    
    def update_company(self, company_id: int, **kwargs) -> bool:
        """Update company information"""
        allowed_fields = ['name', 'registration_no', 'address', 'phone', 'email',
                         'subscription_tier', 'is_active', 'max_users', 'max_projects']
        
        updates = []
        values = []
        
        for key, value in kwargs.items():
            if key in allowed_fields:
                updates.append(f"{key} = ?")
                values.append(value)
        
        if not updates:
            return False
        
        values.append(company_id)
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute(f"""
                UPDATE companies 
                SET {', '.join(updates)}
                WHERE id = ?
            """, values)
            
            return cursor.rowcount > 0
    
    def delete_company(self, company_id: int, hard_delete: bool = False) -> bool:
        """Delete or deactivate company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            if hard_delete:
                cursor.execute("DELETE FROM companies WHERE id = ?", (company_id,))
            else:
                cursor.execute("UPDATE companies SET is_active = 0 WHERE id = ?", (company_id,))
            
            return cursor.rowcount > 0
    
    # ==================== ROLE/PERMISSION METHODS ====================
    
    def get_role_permissions(self, role: str) -> List[str]:
        """Get permissions for a role"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT permission
                FROM role_permissions
                WHERE role = ?
            """, (role,))
            
            return [row['permission'] for row in cursor.fetchall()]
    
    def get_all_roles(self) -> List[Dict]:
        """Get all roles with their permissions"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT role, GROUP_CONCAT(permissions) as permissions
                FROM role_permissions
                GROUP BY role
                ORDER BY role
            """)
            
            roles = []
            for row in cursor.fetchall():
                roles.append({
                    'role': row['role'],
                    'permissions': row['permissions'].split(',') if row['permissions'] else []
                })
            
            return roles
    
    def update_role_permissions(self, role: str, permissions: List[str]) -> bool:
        """Update permissions for a role"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Delete existing permissions
            cursor.execute("DELETE FROM role_permissions WHERE role = ?", (role,))
            
            # Insert new permissions
            for permission in permissions:
                cursor.execute("""
                    INSERT INTO role_permissions (role, permission)
                    VALUES (?, ?)
                """, (role, permission))
            
            return True
    
    # ---------- PWD Chapters ----------
    def get_pwd_chapters(self):
        """Get PWD chapters as DataFrame (original - for backward compatibility)"""
        with self.get_connection() as conn:
            return pd.read_sql_query("""
                SELECT chapter_number, chapter_name, description 
                FROM pwd_chapters 
                ORDER BY CAST(chapter_number AS INTEGER)
            """, conn)

    def get_pwd_chapters_dict(self):
        """Get PWD chapters as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT chapter_number, chapter_name, description 
                FROM pwd_chapters 
                ORDER BY CAST(chapter_number AS INTEGER)
            """)
            return [dict(row) for row in cursor.fetchall()]


    # ---------- PWD Parents ----------
    def get_pwd_parents(self, chapter_number: str = None):
        """Get PWD parents as DataFrame"""
        with self.get_connection() as conn:
            query = """
                SELECT pwd_code, description, chapter_number
                FROM pwd_parents
            """
            params = []
            if chapter_number:
                query += " WHERE chapter_number = ?"
                params.append(chapter_number)
            query += " ORDER BY pwd_code"
            return pd.read_sql_query(query, conn, params=params if params else None)

    def get_pwd_parents_dict(self, chapter_number: str = None):
        """Get PWD parents as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            query = """
                SELECT pwd_code, description, chapter_number
                FROM pwd_parents
            """
            params = []
            if chapter_number:
                query += " WHERE chapter_number = ?"
                params.append(chapter_number)
            query += " ORDER BY pwd_code"
            cursor.execute(query, params if params else None)
            return [dict(row) for row in cursor.fetchall()]


    # ---------- PWD Children ----------
    def get_pwd_children(self, parent_code: str = None, limit: int = 100):
        """Get PWD children as DataFrame"""
        with self.get_connection() as conn:
            if parent_code:
                return pd.read_sql_query("""
                    SELECT c.pwd_code, c.description, c.unit, c.parent_code,
                        r.zone_name, r.unit_rate
                    FROM pwd_children c
                    LEFT JOIN pwd_rates r ON c.pwd_code = r.pwd_code
                    WHERE c.parent_code = ?
                    ORDER BY c.pwd_code
                    LIMIT ?
                """, conn, params=[parent_code, limit])
            else:
                return pd.read_sql_query("""
                    SELECT c.pwd_code, c.description, c.unit, c.parent_code,
                        r.zone_name, r.unit_rate
                    FROM pwd_children c
                    LEFT JOIN pwd_rates r ON c.pwd_code = r.pwd_code
                    ORDER BY c.pwd_code
                    LIMIT ?
                """, conn, params=[limit])

    def get_pwd_children_dict(self, parent_code: str = None, limit: int = 100):
        """Get PWD children as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            if parent_code:
                cursor.execute("""
                    SELECT c.pwd_code, c.description, c.unit, c.parent_code,
                        r.zone_name, r.unit_rate
                    FROM pwd_children c
                    LEFT JOIN pwd_rates r ON c.pwd_code = r.pwd_code
                    WHERE c.parent_code = ?
                    ORDER BY c.pwd_code
                    LIMIT ?
                """, (parent_code, limit))
            else:
                cursor.execute("""
                    SELECT c.pwd_code, c.description, c.unit, c.parent_code,
                        r.zone_name, r.unit_rate
                    FROM pwd_children c
                    LEFT JOIN pwd_rates r ON c.pwd_code = r.pwd_code
                    ORDER BY c.pwd_code
                    LIMIT ?
                """, (limit,))
            return [dict(row) for row in cursor.fetchall()]


    # ---------- PWD Rates ----------
    def get_pwd_rates(self, pwd_code: str = None, zone: str = None):
        """Get PWD rates as DataFrame"""
        with self.get_connection() as conn:
            query = """
                SELECT pwd_code, zone_name, unit_rate, edition_year
                FROM pwd_rates
                WHERE 1=1
            """
            params = []
            if pwd_code:
                query += " AND pwd_code = ?"
                params.append(pwd_code)
            if zone:
                query += " AND zone_name = ?"
                params.append(zone)
            query += " ORDER BY pwd_code, zone_name"
            return pd.read_sql_query(query, conn, params=params if params else None)

    def get_pwd_rates_dict(self, pwd_code: str = None, zone: str = None):
        """Get PWD rates as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            query = """
                SELECT pwd_code, zone_name, unit_rate, edition_year
                FROM pwd_rates
                WHERE 1=1
            """
            params = []
            if pwd_code:
                query += " AND pwd_code = ?"
                params.append(pwd_code)
            if zone:
                query += " AND zone_name = ?"
                params.append(zone)
            query += " ORDER BY pwd_code, zone_name"
            cursor.execute(query, params if params else None)
            return [dict(row) for row in cursor.fetchall()]
    
    def get_pwd_stats_v1(self, chapter_code: str = None) -> Dict:
        """Get statistics for PWD schedule"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            query = """
                SELECT COUNT(*) as total_items,
                       AVG(rate_bdt) as avg_rate,
                       MIN(rate_bdt) as min_rate,
                       MAX(rate_bdt) as max_rate
                FROM pwd_chapters
                WHERE rate_bdt IS NOT NULL
            """
            values = []
            
            if chapter_code:
                query += " AND chapter_code = ?"
                values.append(chapter_code)
            
            cursor.execute(query, values)
            row = cursor.fetchone()
            
            return dict(row) if row else {}
 
    def save_pwd_hierarchy_enhanced(self, hierarchy, version_name, edition_year,
                                    effective_date=None, selected_chapters=None):
        """
        Save PWD hierarchy with robust version handling and safe inserts.
        """
        from datetime import date
        import json
        import datetime as dt

        conn = self.get_connection()
        cursor = conn.cursor()
       
        effective_date = effective_date or date.today()
       
        try:
            # === BULLETPROOF NEXT VERSION LOGIC ===
            cursor.execute("""
                SELECT MAX(version_number) 
                FROM rate_versions
                WHERE source = 'PWD' AND edition_year = ?
            """, (edition_year,))
           
            result = cursor.fetchone()
            print(f"DEBUG: MAX(version_number) returned: {result} (type: {type(result)})")

            # Handle tuple, sqlite3.Row, or dict return types
            if result is None:
                max_val = None
            elif isinstance(result, dict) or hasattr(result, 'get'):
                max_val = result.get('MAX(version_number)') or result.get(0)
            else:
                # traditional tuple
                max_val = result[0] if len(result) > 0 else None

            next_version = 1 if max_val is None else int(max_val) + 1

            print(f"📌 Creating Version {next_version} for PWD {edition_year}")

            # Deactivate current active version
            cursor.execute("""
                UPDATE rate_versions
                SET is_active = 0, updated_at = ?
                WHERE source = 'PWD' AND edition_year = ? AND is_active = 1
            """, (dt.datetime.now(), edition_year))

            # Create new version record
            cursor.execute("""
                INSERT INTO rate_versions (
                    source, version_name, edition_year, version_number,
                    effective_from, is_active, release_date, created_by,
                    has_sections, created_at
                ) VALUES ('PWD', ?, ?, ?, ?, 1, ?, ?, 0, ?)
            """, (version_name, edition_year, next_version, effective_date,
                  dt.datetime.now(), 'system', dt.datetime.now()))
           
            version_id = cursor.lastrowid
            print(f"✅ Created Version {next_version} (ID: {version_id})")

            # ====================== SAVE DATA ======================

            chapter_ids = {}
            chapters_saved = 0

            # Save chapters to rate_chapters
            if selected_chapters:
                for chapter_num, chapter_info in selected_chapters.items():
                    cursor.execute("""
                        INSERT INTO rate_chapters (
                            source, version_id, chapter_number,
                            chapter_name, description, display_order
                        ) VALUES ('PWD', ?, ?, ?, ?, ?)
                    """, (version_id, chapter_num, 
                          chapter_info.get('name', f'Chapter {chapter_num}'),
                          chapter_info.get('description', ''), 
                          int(str(chapter_num)) if str(chapter_num).isdigit() else 999))
                    chapter_ids[chapter_num] = cursor.lastrowid
                    chapters_saved += 1

            # Save Parents
            parents_saved = 0
            for parent in hierarchy.get('parents', []):
                ch_num = parent.get('chapter_number') or parent.get('chapter') or str(chapter_num)[:2]
                cursor.execute("""
                    INSERT OR REPLACE INTO pwd_parents 
                    (pwd_code, description, chapter_number, version_id)
                    VALUES (?, ?, ?, ?)
                """, (parent['code'], parent.get('description', ''), ch_num, version_id))
                parents_saved += 1

            # Save Children + Rates
            children_saved = 0
            rates_saved = 0
            current_chapter = None  # fallback

            for child in hierarchy.get('children', []):
                code = child.get('pwd_code') or child.get('code')
                if not code:
                    continue

                parent_code = child.get('parent_code') or code  # self-reference

                # Use chapter from child or fallback
                ch_num = child.get('chapter_number') or current_chapter

                cursor.execute("""
                    INSERT OR REPLACE INTO pwd_children (
                        pwd_code, parent_code, description, unit,
                        edition_year, version_id, chapter_number
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    code, 
                    parent_code, 
                    child.get('description', ''),
                    child.get('unit', ''), 
                    edition_year, 
                    version_id, 
                    ch_num
                ))
                children_saved += 1

                # Save rates
                for zone, rate in child.get('rates', {}).items():
                    cursor.execute("""
                        INSERT OR REPLACE INTO pwd_rates 
                        (pwd_code, zone_name, unit_rate, edition_year, version_id)
                        VALUES (?, ?, ?, ?, ?)
                    """, (code, zone, float(rate), edition_year, version_id))
                    rates_saved += 1

            # Update version statistics
            cursor.execute("""
                UPDATE rate_versions
                SET total_parents = ?, total_children = ?, total_rates = ?,
                    chapter_numbers = ?
                WHERE id = ?
            """, (parents_saved, children_saved, rates_saved,
                  json.dumps(list(chapter_ids.keys())), version_id))

            conn.commit()
           
            print(f"✅ Successfully saved PWD hierarchy → Version {next_version} | "
                  f"Chapters: {chapters_saved}, Parents: {parents_saved}, "
                  f"Children: {children_saved}, Rates: {rates_saved}")
           
            return version_id
           
        except Exception as e:
            conn.rollback()
            print(f"❌ Error saving PWD hierarchy: {e}")
            import traceback
            traceback.print_exc()
            raise
        finally:
            conn.close()
    def save_pwd_hierarchy_enhanced_bak(self, hierarchy, version_name, edition_year, 
                                effective_date=None, selected_chapters=None):
        """
        Save PWD hierarchy with chapter support (PWD doesn't have sections).
        Auto-increments version number for each new version.
        """
        from datetime import date
        import json
        
        conn = self.get_connection()
        cursor = conn.cursor()
        
        effective_date = effective_date or date.today()
        
        try:
            # Get the next version number
            cursor.execute("""
                SELECT MAX(version_number) FROM rate_versions 
                WHERE source = 'PWD' AND edition_year = ?
            """, (edition_year,))
            
            result = cursor.fetchone()
            next_version = (result[0] or 0) + 1
            
            # Deactivate current active version
            cursor.execute("""
                UPDATE rate_versions 
                SET is_active = 0, updated_at = ?
                WHERE source = 'PWD' AND edition_year = ? AND is_active = 1
            """, (datetime.now(), edition_year))
            
            # Create version record
            cursor.execute("""
                INSERT INTO rate_versions (
                    source, version_name, edition_year, version_number,
                    effective_from, is_active, release_date, created_by, 
                    has_sections, created_at
                ) VALUES ('PWD', ?, ?, ?, ?, 1, ?, ?, 0, ?)
            """, (version_name, edition_year, next_version, effective_date, 
                datetime.now(), 'system', datetime.now()))
            
            version_id = cursor.lastrowid
            print(f"✅ Created version {next_version} for PWD {edition_year}")
            
            chapter_ids = {}
            
            # Save chapters to rate_chapters
            chapters_saved = 0
            if selected_chapters:
                for chapter_num, chapter_info in selected_chapters.items():
                    cursor.execute("""
                        INSERT INTO rate_chapters (
                            source, version_id, chapter_number, 
                            chapter_name, description, display_order
                        ) VALUES ('PWD', ?, ?, ?, ?, ?)
                    """, (version_id, chapter_num, chapter_info.get('name', f'Chapter {chapter_num}'),
                        chapter_info.get('description', ''), int(chapter_num) if chapter_num.isdigit() else 999))
                    chapter_ids[chapter_num] = cursor.lastrowid
                    chapters_saved += 1
            
            # Save parents to pwd_parents
            parents_saved = 0
            for parent in hierarchy.get('parents', []):
                chapter_num = parent.get('chapter_number', parent.get('chapter', parent['code']))
                
                cursor.execute("""
                    INSERT INTO pwd_parents (pwd_code, description, chapter_number, version_id)
                    VALUES (?, ?, ?, ?)
                """, (parent['code'], parent.get('description', ''), chapter_num, version_id))
                parents_saved += 1
            
            # Save children to pwd_children - handle NOT NULL constraint
            children_saved = 0
            rates_saved = 0
            
            for child in hierarchy.get('children', []):
                code = child.get('pwd_code') or child.get('code')
                if not code:
                    continue
                
                # Handle parent_code NOT NULL constraint
                parent_code = child.get('parent_code')
                if parent_code is None or parent_code == '':
                    # For leaf items, use empty string (if allowed) or a default value
                    parent_code = ''  # Try empty string first
                    # If empty string fails, use: parent_code = code (self-reference)
                
                cursor.execute("""
                    INSERT INTO pwd_children (
                        pwd_code, parent_code, description, unit, 
                        edition_year, version_id
                    ) VALUES (?, ?, ?, ?, ?, ?)
                """, (code, parent_code, child.get('description', ''), 
                    child.get('unit', ''), edition_year, version_id))
                children_saved += 1
                
                # Save rates
                for zone, rate in child.get('rates', {}).items():
                    cursor.execute("""
                        INSERT INTO pwd_rates (pwd_code, zone_name, unit_rate, edition_year, version_id)
                        VALUES (?, ?, ?, ?, ?)
                    """, (code, zone, rate, edition_year, version_id))
                    rates_saved += 1
            
            # Update version record with statistics
            cursor.execute("""
                UPDATE rate_versions 
                SET total_parents = ?, total_children = ?, total_rates = ?,
                    chapter_numbers = ?
                WHERE id = ?
            """, (parents_saved, children_saved, rates_saved,
                json.dumps(list(chapter_ids.keys())), version_id))
            
            conn.commit()
            
            print(f"✅ Saved PWD hierarchy: {chapters_saved} chapters, "
                f"{parents_saved} parents, {children_saved} children, {rates_saved} rates")
            
            return version_id
            
        except Exception as e:
            conn.rollback()
            print(f"❌ Error saving PWD hierarchy: {e}")
            import traceback
            traceback.print_exc()
            raise e
        finally:
            conn.close()


    def save_pwd_hierarchy(self, hierarchy, edition_year, version_name=None):
        """
        Save PWD hierarchy to database with version number support.
        """
        from datetime import date
        
        self.init_pwd_hierarchical_tables()
        
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Create version if new
        if version_name is None:
            version_name = f"PWD Schedule {edition_year}"
        
        # Check if rate_versions table has all columns
        cursor.execute("PRAGMA table_info(rate_versions)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'source' not in columns:
            cursor.execute("ALTER TABLE rate_versions ADD COLUMN source TEXT DEFAULT 'PWD'")
        if 'total_parents' not in columns:
            cursor.execute("ALTER TABLE rate_versions ADD COLUMN total_parents INTEGER DEFAULT 0")
        if 'total_children' not in columns:
            cursor.execute("ALTER TABLE rate_versions ADD COLUMN total_children INTEGER DEFAULT 0")
        if 'total_rates' not in columns:
            cursor.execute("ALTER TABLE rate_versions ADD COLUMN total_rates INTEGER DEFAULT 0")
        if 'created_by' not in columns:
            cursor.execute("ALTER TABLE rate_versions ADD COLUMN created_by TEXT")
        if 'version_number' not in columns:
            cursor.execute("ALTER TABLE rate_versions ADD COLUMN version_number INTEGER DEFAULT 1")
        if 'created_at' not in columns:
            cursor.execute("ALTER TABLE rate_versions ADD COLUMN created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        if 'updated_at' not in columns:
            cursor.execute("ALTER TABLE rate_versions ADD COLUMN updated_at TIMESTAMP")
        
        # Get the next version number
        cursor.execute("""
            SELECT MAX(version_number) FROM rate_versions 
            WHERE source = 'PWD' AND edition_year = ?
        """, (edition_year,))
        
        result = cursor.fetchone()
        next_version = (result[0] or 0) + 1
        
        # Deactivate current active version
        cursor.execute("""
            UPDATE rate_versions 
            SET is_active = 0, updated_at = ?
            WHERE source = 'PWD' AND edition_year = ? AND is_active = 1
        """, (datetime.now(), edition_year))
        
        # Create version record
        cursor.execute("""
            INSERT INTO rate_versions (
                source, version_name, edition_year, version_number, 
                effective_from, is_active, release_date, created_by, created_at
            ) VALUES ('PWD', ?, ?, ?, ?, 1, ?, ?, ?)
        """, (version_name, edition_year, next_version, date.today(), 
            datetime.now(), 'system', datetime.now()))
        
        version_id = cursor.lastrowid
        
        # Clear existing data for this version
        cursor.execute("DELETE FROM pwd_rates WHERE version_id = ?", (version_id,))
        cursor.execute("DELETE FROM pwd_children WHERE version_id = ?", (version_id,))
        cursor.execute("DELETE FROM pwd_parents WHERE version_id = ?", (version_id,))
        
        # Save parents
        parents_saved = 0
        for parent in hierarchy.get('parents', []):
            cursor.execute("""
                INSERT INTO pwd_parents (pwd_code, description, chapter_number, version_id)
                VALUES (?, ?, ?, ?)
            """, (parent['code'], parent.get('description', ''), parent.get('chapter', parent['code']), version_id))
            parents_saved += 1
        
        # Save children and rates
        children_saved = 0
        rates_saved = 0
        
        for child in hierarchy.get('children', []):
            # Determine chapter from code
            code = child['pwd_code']
            chapter_num = code.split('.')[0] if '.' in code else code
            
            # Get chapter_id from rate_chapters if exists
            chapter_id = None
            cursor.execute("""
                SELECT id FROM rate_chapters 
                WHERE source = 'PWD' AND chapter_number = ? AND version_id = ?
            """, (chapter_num, version_id))
            chapter_row = cursor.fetchone()
            if chapter_row:
                chapter_id = chapter_row[0]
            
            cursor.execute("""
                INSERT INTO pwd_children (
                    pwd_code, parent_code, description, unit, 
                    edition_year, version_id, chapter_id, is_parent
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (code, child.get('parent_code', ''), child.get('description', ''), 
                child.get('unit', ''), edition_year, version_id, chapter_id, 0))
            children_saved += 1
            
            for zone, rate in child.get('rates', {}).items():
                cursor.execute("""
                    INSERT INTO pwd_rates (pwd_code, zone_name, unit_rate, edition_year, version_id)
                    VALUES (?, ?, ?, ?, ?)
                """, (code, zone, rate, edition_year, version_id))
                rates_saved += 1
        
        # Update version record with statistics
        cursor.execute("""
            UPDATE rate_versions 
            SET total_parents = ?, total_children = ?, total_rates = ?
            WHERE id = ?
        """, (parents_saved, children_saved, rates_saved, version_id))
        
        conn.commit()
        conn.close()
        
        print(f"✅ Saved PWD hierarchy: {parents_saved} parents, {children_saved} children, {rates_saved} rates")
        
        return {
            'parents': parents_saved, 
            'children': children_saved, 
            'rates': rates_saved,
            'version_id': version_id,
            'version_number': next_version
        }
    # =========================================================
    # LGED METHODS - Option 3 (Keep original + Add dict versions)
    # =========================================================

    # ---------- LGED Chapters ----------
    def get_lged_chapters(self):
        """Get LGED chapters as DataFrame"""
        with self.get_connection() as conn:
            return pd.read_sql_query("""
                SELECT chapter_number, chapter_name, description 
                FROM lged_chapters 
                ORDER BY CAST(chapter_number AS INTEGER)
            """, conn)

    def get_lged_chapters_dict(self):
        """Get LGED chapters as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT chapter_number, chapter_name, description 
                FROM lged_chapters 
                ORDER BY CAST(chapter_number AS INTEGER)
            """)
            return [dict(row) for row in cursor.fetchall()]


    # ---------- LGED Sections ----------
    def get_lged_sections(self, chapter_number: str = None):
        """Get LGED sections as DataFrame"""
        with self.get_connection() as conn:
            if chapter_number:
                return pd.read_sql_query("""
                    SELECT section_number, section_name, description
                    FROM lged_sections
                    WHERE chapter_number = ?
                    ORDER BY section_number
                """, conn, params=[chapter_number])
            else:
                return pd.read_sql_query("""
                    SELECT chapter_number, section_number, section_name, description
                    FROM lged_sections
                    ORDER BY chapter_number, section_number
                """, conn)

    def get_lged_sections_dict(self, chapter_number: str = None):
        """Get LGED sections as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            if chapter_number:
                cursor.execute("""
                    SELECT section_number, section_name, description
                    FROM lged_sections
                    WHERE chapter_number = ?
                    ORDER BY section_number
                """, (chapter_number,))
            else:
                cursor.execute("""
                    SELECT chapter_number, section_number, section_name, description
                    FROM lged_sections
                    ORDER BY chapter_number, section_number
                """)
            return [dict(row) for row in cursor.fetchall()]


    # ---------- LGED Parents ----------
    def get_lged_parents(self, chapter_number: str = None, section_number: str = None):
        """Get LGED parents as DataFrame"""
        with self.get_connection() as conn:
            query = """
                SELECT code, description, chapter_number, section_number, 
                    parent_type, has_children, unit
                FROM lged_parents
                WHERE 1=1
            """
            params = []
            if chapter_number:
                query += " AND chapter_number = ?"
                params.append(chapter_number)
            if section_number:
                query += " AND section_number = ?"
                params.append(section_number)
            query += " ORDER BY code"
            return pd.read_sql_query(query, conn, params=params if params else None)

    def get_lged_parents_dict(self, chapter_number: str = None, section_number: str = None):
        """Get LGED parents as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            query = """
                SELECT code, description, chapter_number, section_number, 
                    parent_type, has_children, unit
                FROM lged_parents
                WHERE 1=1
            """
            params = []
            if chapter_number:
                query += " AND chapter_number = ?"
                params.append(chapter_number)
            if section_number:
                query += " AND section_number = ?"
                params.append(section_number)
            query += " ORDER BY code"
            cursor.execute(query, params if params else None)
            return [dict(row) for row in cursor.fetchall()]


    # ---------- LGED Children ----------
    def get_lged_children(self, parent_code: str = None, chapter_number: str = None, limit: int = 100):
        """Get LGED children as DataFrame"""
        with self.get_connection() as conn:
            query = """
                SELECT code, parent_code, description, unit,
                    chapter_number, section_number,
                    zone_a, zone_b, zone_c, zone_d
                FROM lged_children
                WHERE 1=1
            """
            params = []
            if parent_code:
                query += " AND parent_code = ?"
                params.append(parent_code)
            if chapter_number:
                query += " AND chapter_number = ?"
                params.append(chapter_number)
            query += " ORDER BY code LIMIT ?"
            params.append(limit)
            return pd.read_sql_query(query, conn, params=params)

    def get_lged_children_dict(self, parent_code: str = None, chapter_number: str = None, limit: int = 100):
        """Get LGED children as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            query = """
                SELECT code, parent_code, description, unit,
                    chapter_number, section_number,
                    zone_a, zone_b, zone_c, zone_d
                FROM lged_children
                WHERE 1=1
            """
            params = []
            if parent_code:
                query += " AND parent_code = ?"
                params.append(parent_code)
            if chapter_number:
                query += " AND chapter_number = ?"
                params.append(chapter_number)
            query += " ORDER BY code LIMIT ?"
            params.append(limit)
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]


    # ---------- LGED Zone Rates ----------
    def get_lged_zone_rates(self, child_id: int = None):
        """Get LGED zone rates as DataFrame"""
        with self.get_connection() as conn:
            if child_id:
                return pd.read_sql_query("""
                    SELECT zone_name, unit_rate
                    FROM lged_zone_rates
                    WHERE child_id = ?
                    ORDER BY zone_name
                """, conn, params=[child_id])
            else:
                return pd.read_sql_query("""
                    SELECT child_id, zone_name, unit_rate
                    FROM lged_zone_rates
                    ORDER BY child_id, zone_name
                """, conn)

    def get_lged_zone_rates_dict(self, child_id: int = None):
        """Get LGED zone rates as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            if child_id:
                cursor.execute("""
                    SELECT zone_name, unit_rate
                    FROM lged_zone_rates
                    WHERE child_id = ?
                    ORDER BY zone_name
                """, (child_id,))
            else:
                cursor.execute("""
                    SELECT child_id, zone_name, unit_rate
                    FROM lged_zone_rates
                    ORDER BY child_id, zone_name
                """)
            return [dict(row) for row in cursor.fetchall()]


    # ---------- LGED Zone Mapping ----------
    def get_lged_zone_mapping(self):
        """Get LGED zone mapping as DataFrame"""
        with self.get_connection() as conn:
            return pd.read_sql_query("""
                SELECT zone_code, zone_name, divisions, accessibility_bonus, description
                FROM lged_zone_mapping
                ORDER BY zone_code
            """, conn)

    def get_lged_zone_mapping_dict(self):
        """Get LGED zone mapping as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT zone_code, zone_name, divisions, accessibility_bonus, description
                FROM lged_zone_mapping
                ORDER BY zone_code
            """)
            return [dict(row) for row in cursor.fetchall()]


    def get_rate_by_item_code(self, schedule_type: str, item_code: str) -> Optional[Dict]:
        """Get rate by item code from specified schedule"""
        table = 'pwd_chapters' if schedule_type == 'pwd' else 'lged_schedule'
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute(f"""
                SELECT item_code, description, unit, rate_bdt, chapter_code, chapter_name
                FROM {table}
                WHERE item_code = ?
            """, (item_code,))
            
            row = cursor.fetchone()
            return dict(row) if row else None
    
    
    def save_lged_hierarchy_enhanced(self, hierarchy, version_name, edition_year, 
                                  effective_date=None, selected_chapters=None, 
                                  selected_sections=None):
        """
        Save LGED hierarchy with support for:
        - Section headers (parents without rates, have children)
        - Leaf items (parents with rates, no children)
        - Child items (belong to section headers)
        """
        from datetime import date
        import json
        
        # Ensure schema is updated
        self.migrate_lged_tables_for_parent_types()
        
        conn = self.get_connection()
        cursor = conn.cursor()
        
        effective_date = effective_date or date.today()
        has_sections = selected_sections is not None and len(selected_sections) > 0
        
        try:
            # Create version record
            cursor.execute("""
                INSERT INTO rate_versions (source, version_name, edition_year, effective_from, 
                                        is_active, release_date, created_by, has_sections)
                VALUES ('LGED', ?, ?, ?, 1, ?, ?, ?)
            """, (version_name, edition_year, effective_date, datetime.now(), 
                'system', has_sections))
            
            version_id = cursor.lastrowid
            
            # Save section headers (parents without rates)
            section_headers = [p for p in hierarchy.get('section_headers', [])]
            leaf_items = [p for p in hierarchy.get('leaf_items', [])]
            children = hierarchy.get('children', [])
            
            parent_ids = {}
            
            # Save section headers (parents without rates, may have children)
            for header in section_headers:
                cursor.execute("""
                    INSERT INTO lged_parents (code, description, chapter_number, section_number, 
                                            parent_type, has_children, version_id)
                    VALUES (?, ?, ?, ?, 'section_header', ?, ?)
                """, (header['code'], header.get('description', ''), 
                    header.get('chapter_number', ''), header.get('section_number', ''),
                    1 if header.get('has_children') else 0, version_id))
                parent_ids[header['code']] = cursor.lastrowid
            
            # Save leaf items (parents with rates, no children)
            for leaf in leaf_items:
                cursor.execute("""
                    INSERT INTO lged_parents (code, description, chapter_number, section_number, 
                                            parent_type, has_children, unit, version_id)
                    VALUES (?, ?, ?, ?, 'leaf_item', 0, ?, ?)
                """, (leaf['code'], leaf.get('description', ''), 
                    leaf.get('chapter_number', ''), leaf.get('section_number', ''),
                    leaf.get('unit', ''), version_id))
                parent_id = cursor.lastrowid
                parent_ids[leaf['code']] = parent_id
                
                # Save rates for leaf items
                for zone, rate in leaf.get('rates', {}).items():
                    cursor.execute("""
                        INSERT INTO lged_zone_rates (parent_id, zone_name, unit_rate, version_id)
                        VALUES (?, ?, ?, ?)
                    """, (parent_id, zone, rate, version_id))
            
            # Save child items (belong to section headers)
            for child in children:
                parent_code = child.get('parent_code', '')
                cursor.execute("""
                    INSERT INTO lged_children (code, parent_code, description, unit, 
                                            chapter_number, section_number,
                                            zone_a, zone_b, zone_c, zone_d,
                                            edition_year, version_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (child['code'], parent_code, child.get('description', ''), 
                    child.get('unit', ''),
                    child.get('chapter_number', ''), child.get('section_number', ''),
                    child.get('zone_a'), child.get('zone_b'), 
                    child.get('zone_c'), child.get('zone_d'),
                    edition_year, version_id))
            
            # Update version record with statistics
            cursor.execute("""
                UPDATE rate_versions 
                SET total_parents = ?, total_children = ?, total_rates = ?
                WHERE id = ?
            """, (len(section_headers) + len(leaf_items), len(children), 
                len(leaf_items) + len(children), version_id))
            
            conn.commit()
            
            print(f"✅ Saved LGED hierarchy:")
            print(f"  - Section Headers: {len(section_headers)}")
            print(f"  - Leaf Items: {len(leaf_items)}")
            print(f"  - Child Items: {len(children)}")
            
            return version_id
            
        except Exception as e:
            conn.rollback()
            print(f"❌ Error saving LGED hierarchy: {e}")
            import traceback
            traceback.print_exc()
            raise e
        finally:
            conn.close()

    def save_lged_hierarchy(self, hierarchy, version_name, edition_year, effective_date=None):
        """Save LGED hierarchy to database"""
        from datetime import date
        
        self.init_lged_tables()
        
        conn = self.get_connection()
        cursor = conn.cursor()
        
        effective_date = effective_date or date.today()
        
        # Create version record with all required columns
        cursor.execute("""
            INSERT INTO rate_versions (source, version_name, edition_year, effective_from, is_active, release_date, created_by)
            VALUES ('LGED', ?, ?, ?, 1, ?, ?)
        """, (version_name, edition_year, effective_date, datetime.now(), 'system'))
        
        version_id = cursor.lastrowid
        
        # Insert parents
        parents_saved = 0
        for parent in hierarchy.get('parents', []):
            cursor.execute("""
                INSERT INTO lged_parents (code, description, chapter_number, version_id)
                VALUES (?, ?, ?, ?)
            """, (parent['code'], parent['description'], parent['chapter'], version_id))
            parents_saved += 1
        
        # Insert children
        children_saved = 0
        rates_saved = 0
        
        for child in hierarchy.get('children', []):
            cursor.execute("""
                INSERT INTO lged_children (code, parent_code, description, unit, edition_year, version_id)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (child['code'], child['parent_code'], child['description'], 
                child.get('unit', ''), edition_year, version_id))
            child_id = cursor.lastrowid
            children_saved += 1
            
            # Insert zone rates
            for zone, rate in child.get('rates', {}).items():
                cursor.execute("""
                    INSERT INTO lged_zone_rates (child_id, zone_name, unit_rate, version_id)
                    VALUES (?, ?, ?, ?)
                """, (child_id, zone, rate, version_id))
                rates_saved += 1
        
        # Update version record with statistics
        cursor.execute("""
            UPDATE rate_versions 
            SET total_parents = ?, total_children = ?, total_rates = ?
            WHERE id = ?
        """, (parents_saved, children_saved, rates_saved, version_id))
        
        conn.commit()
        conn.close()
        
        print(f"✅ Saved LGED hierarchy: {parents_saved} parents, {children_saved} children, {rates_saved} rates")
        
        return version_id
    # =========================================================
    # RATE VERSIONS METHODS
    # =========================================================

    def get_rate_versions(self, source: str = None):
        """Get rate versions as DataFrame"""
        with self.get_connection() as conn:
            if source:
                return pd.read_sql_query("""
                    SELECT id, source, version_name, edition_year, version_number,
                        effective_from, is_active, release_date, notes,
                        total_parents, total_children, total_rates
                    FROM rate_versions
                    WHERE source = ?
                    ORDER BY edition_year DESC, version_number DESC
                """, conn, params=[source])
            else:
                return pd.read_sql_query("""
                    SELECT id, source, version_name, edition_year, version_number,
                        effective_from, is_active, release_date, notes,
                        total_parents, total_children, total_rates
                    FROM rate_versions
                    ORDER BY source, edition_year DESC, version_number DESC
                """, conn)

    def get_rate_versions_dict(self, source: str = None):
        """Get rate versions as list of dictionaries"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            if source:
                cursor.execute("""
                    SELECT id, source, version_name, edition_year, version_number,
                        effective_from, is_active, release_date, notes,
                        total_parents, total_children, total_rates
                    FROM rate_versions
                    WHERE source = ?
                    ORDER BY edition_year DESC, version_number DESC
                """, (source,))
            else:
                cursor.execute("""
                    SELECT id, source, version_name, edition_year, version_number,
                        effective_from, is_active, release_date, notes,
                        total_parents, total_children, total_rates
                    FROM rate_versions
                    ORDER BY source, edition_year DESC, version_number DESC
                """)
            return [dict(row) for row in cursor.fetchall()]

    # database/crud_operations.py - Fixed calculate_win_probability

    
    def calculate_win_probability(self, tender_id: str, user_total_cost: float):
        """
        Evaluates historical competitor price ranges for the same tender 
        to calculate a statistical win probability percentage.
        
        Args:
            tender_id: The tender ID to analyze
            user_total_cost: User's total bid amount
        
        Returns:
            tuple: (probability_percentage, message)
        """
        try:
            with self.get_connection() as conn:
                # Use pandas to read the data
                import pandas as pd
                query = "SELECT total_bid_amount FROM competitor_bids WHERE tender_id = ?"
                df_comp = pd.read_sql_query(query, conn, params=(tender_id,))
            
            if df_comp.empty:
                return None, "No competitor bid data available for this tender"
            
            all_bids = df_comp['total_bid_amount'].tolist()
            all_bids.append(user_total_cost)
            all_bids.sort()
            
            # Find rank index positioning (Lower price is preferred in e-GP L1 procurement formats)
            rank = all_bids.index(user_total_cost) + 1
            total_bidders = len(all_bids)
            
            # Calculate probability: higher ranking position results in a stronger win margin
            probability = ((total_bidders - rank) / total_bidders) * 100
            return round(probability, 1), f"Rank {rank} of {total_bidders}"
            
        except Exception as e:
            logger.error(f"Error calculating win probability: {e}")
            return None, f"Error: {str(e)}"

    # ==================== BATCH 1: USER MANAGEMENT & AUTHENTICATION ====================

    def store_password_reset_token(self, email: str, token: str, expires_in_minutes: int = 60) -> bool:
        """Store password reset token"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            try:
                expires_at = (datetime.now() + timedelta(minutes=expires_in_minutes)).strftime('%Y-%m-%d %H:%M:%S')
                cursor.execute('''
                    INSERT OR REPLACE INTO password_reset_tokens 
                    (email, token, expires_at) 
                    VALUES (?, ?, ?)
                ''', (email, token, expires_at))
                return True
            except Exception as e:
                logger.error(f"Failed to store reset token: {e}")
                return False

    def verify_reset_token(self, token: str) -> Optional[str]:
        """Verify reset token and return email if valid"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            try:
                cursor.execute('''
                    SELECT email FROM password_reset_tokens 
                    WHERE token = ? AND expires_at > ?
                ''', (token, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                result = cursor.fetchone()
                return result.get('email') if result else None
            finally:
                pass

    def update_password(self, email: str, new_password: str) -> bool:
        """Update user password with bcrypt"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            try:
                hashed = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                cursor.execute("UPDATE users SET password = ? WHERE email = ?", (hashed, email))
                return cursor.rowcount > 0
            finally:
                pass

    def get_pending_users(self, company_id: int) -> List[Dict]:
        """Get all pending approval users for a company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                SELECT id, username, email, full_name, phone, role, created_at, created_by
                FROM users 
                WHERE company_id = ? AND is_approved = 0 AND (registration_complete = 0 OR registration_complete IS NULL) AND is_active = 1
                ORDER BY created_at ASC
            ''', (company_id,))
            rows = cursor.fetchall()
            return [dict(row) for row in rows]

    def approve_user(self, user_id: int, approved_by: int) -> bool:
        """Approve a pending user registration"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            try:
                cursor.execute('''
                    UPDATE users 
                    SET is_approved = 1, approved_by = ?, approved_at = ?
                    WHERE id = ?
                ''', (approved_by, datetime.now(), user_id))
                cursor.execute('''
                    UPDATE subscriptions 
                    SET status = 'active', analyses_limit = 5
                    WHERE user_id = ?
                ''', (user_id,))
                return True
            except Exception as e:
                logger.error(f"Error approving user: {e}")
                return False

    def reject_user(self, user_id: int, rejected_by: int) -> bool:
        """Reject a pending user registration"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                UPDATE users 
                SET is_active = 0, registration_complete = 0
                WHERE id = ?
            ''', (user_id,))
            return True

    def is_user_approved(self, user_id: int) -> bool:
        """Check if user is approved"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('SELECT is_approved, is_active FROM users WHERE id = ?', (user_id,))
            result = cursor.fetchone()
            if result:
                return result.get('is_approved', 0) == 1 and result.get('is_active', 0) == 1
            return False

    def get_all_companies(self) -> List[Dict]:
        """Fetch all companies from database"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                SELECT id, company_name, email, phone, division, district, created_at, status 
                FROM companies 
                ORDER BY company_name ASC
            ''')
            return [dict(row) for row in cursor.fetchall()]

    def get_company_by_id(self, company_id: int) -> Optional[Dict]:
        """Get company by ID"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, company_name, email, phone, division, district, 
                    is_individual, created_at, registration_number, vat_number
                FROM companies 
                WHERE id = ?
            """, (company_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_system_users(self):
        """Get all system-level users (company_id IS NULL)"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, username, email, full_name, phone, 
                    mobile_number, mobile_verified,  -- ✅ ADD THESE
                    role, is_active, created_at, last_login
                FROM users
                WHERE company_id IS NULL
                ORDER BY created_at DESC
            """)
            rows = cursor.fetchall()
            
            users = []
            for row in rows:
                users.append({
                    'id': row.get('id'),
                    'username': row.get('username'),
                    'email': row.get('email'),
                    'full_name': row.get('full_name'),
                    'phone': row.get('phone', ''),
                    'mobile_number': row.get('mobile_number', ''),  # ✅ ADDED
                    'mobile_verified': row.get('mobile_verified', 0),  # ✅ ADDED
                    'role': row.get('role', 'viewer'),
                    'is_active': row.get('is_active', 1),
                    'created_at': row.get('created_at'),
                    'last_login': row.get('last_login')
                })
            return users
    

    def get_company_users(self, company_id: int) -> List[Dict]:
        """Get all users belonging to a specific company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, username, email, full_name, phone, role, is_active, 
                    created_at, last_login
                FROM users
                WHERE company_id = ?
                ORDER BY created_at DESC
            """, (company_id,))
            return [dict(row) for row in cursor.fetchall()]

    # database/crud_operations.py - Update these methods

    def create_company_user(self, company_id: int, user_data: Dict, created_by: int) -> tuple:
        """Create a user under a specific company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Validate mobile number
            mobile = user_data.get('mobile_number', '')
            if not mobile:
                return False, "Mobile number is required"
            
            mobile = self.normalize_mobile(mobile)
            if not self.validate_bangladesh_mobile(mobile):
                return False, f"Invalid mobile number: {mobile}"
            
            # Check if mobile already exists
            cursor.execute("SELECT id FROM users WHERE mobile_number = ?", (mobile,))
            if cursor.fetchone():
                return False, f"Mobile number {mobile} is already registered"
            
            # Check if email already exists
            cursor.execute("SELECT id FROM users WHERE email = ?", (user_data['email'],))
            if cursor.fetchone():
                return False, f"Email {user_data['email']} is already registered"
            
            # Check if username already exists
            cursor.execute("SELECT id FROM users WHERE username = ?", (user_data['username'],))
            if cursor.fetchone():
                return False, f"Username {user_data['username']} is already taken"
            
            hashed_pass = bcrypt.hashpw(user_data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            try:
                cursor.execute('''
                    INSERT INTO users (
                        company_id, username, password, email, full_name, phone, mobile_number, role,
                        is_active, created_by, is_approved, account_type
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    company_id, 
                    user_data['username'], 
                    hashed_pass, 
                    user_data['email'],
                    user_data['full_name'], 
                    user_data.get('phone', ''), 
                    mobile,
                    user_data.get('role', 'viewer'),
                    1, 
                    created_by, 
                    1,
                    'company'
                ))
                user_id = cursor.lastrowid
                cursor.execute('''
                    INSERT INTO subscriptions (user_id, plan, status, analyses_limit)
                    VALUES (?, ?, ?, ?)
                ''', (user_id, 'free', 'active', 5))
                return True, user_id
            except Exception as e:
                return False, str(e)


    def create_system_user(self, user_data: Dict, created_by: int) -> tuple:
        """Create a system-level user (no company)"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Validate mobile number
            mobile = user_data.get('mobile_number', '')
            if not mobile:
                return False, "Mobile number is required"
            
            mobile = self.normalize_mobile(mobile)
            if not self.validate_bangladesh_mobile(mobile):
                return False, f"Invalid mobile number: {mobile}"
            
            # Check if mobile already exists
            cursor.execute("SELECT id FROM users WHERE mobile_number = ?", (mobile,))
            if cursor.fetchone():
                return False, f"Mobile number {mobile} is already registered"
            
            # Check if email already exists
            cursor.execute("SELECT id FROM users WHERE email = ?", (user_data['email'],))
            if cursor.fetchone():
                return False, f"Email {user_data['email']} is already registered"
            
            # Check if username already exists
            cursor.execute("SELECT id FROM users WHERE username = ?", (user_data['username'],))
            if cursor.fetchone():
                return False, f"Username {user_data['username']} is already taken"
            
            hashed_pass = bcrypt.hashpw(user_data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            try:
                cursor.execute('''
                    INSERT INTO users (
                        company_id, username, password, email, full_name, phone, mobile_number, role,
                        is_active, created_by, is_approved, account_type
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    None, 
                    user_data['username'], 
                    hashed_pass, 
                    user_data['email'],
                    user_data['full_name'], 
                    user_data.get('phone', ''), 
                    mobile,
                    user_data.get('role', 'viewer'),
                    1, 
                    created_by, 
                    1,
                    'system'
                ))
                user_id = cursor.lastrowid
                cursor.execute('''
                    INSERT INTO subscriptions (user_id, plan, status, analyses_limit)
                    VALUES (?, ?, ?, ?)
                ''', (user_id, 'professional', 'active', -1))
                return True, user_id
            except Exception as e:
                return False, str(e)
    def _hash_password(self, password: str) -> str:
        """Hash a password using bcrypt"""
        return bcrypt.hashpw(
            password.encode('utf-8'), 
            bcrypt.gensalt()
        ).decode('utf-8')

    # 2. Verify password
    def _verify_password(self, password: str, hashed: str) -> bool:
        """Verify a password against its hash"""
        try:
            return bcrypt.checkpw(
                password.encode('utf-8'), 
                hashed.encode('utf-8')
            )
        except bcrypt.InvalidHashError:
            return False

    def change_user_password(self, user_id: int, current_password: str, 
                         new_password: str) -> Tuple[bool, str]:
        """
        Change user password with verification of current password.
        
        Args:
            user_id: User ID
            current_password: Current password for verification
            new_password: New password to set
        
        Returns:
            Tuple[bool, str]: (success, message)
        """
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                # Get current password hash
                cursor.execute("SELECT password FROM users WHERE id = ?", (user_id,))
                row = cursor.fetchone()
                
                if not row:
                    return False, "User not found"
                
                # Extract password hash
                if isinstance(row, dict):
                    stored_hash = row.get('password')
                else:
                    stored_hash = row[0]
                
                # Verify current password
                if not bcrypt.checkpw(
                    current_password.encode('utf-8'), 
                    stored_hash.encode('utf-8')
                ):
                    return False, "Current password is incorrect"
                
                # Hash new password
                hashed = bcrypt.hashpw(
                    new_password.encode('utf-8'), 
                    bcrypt.gensalt()
                ).decode('utf-8')
                
                # Update password
                cursor.execute(
                    "UPDATE users SET password = ? WHERE id = ?",
                    (hashed, user_id)
                )
                conn.commit()
                
                # Log the activity
                self.log_user_activity(
                    user_id, 
                    'password_change', 
                    'Password changed successfully'
                )
                
                return True, "Password changed successfully"
                
            except bcrypt.InvalidHashError:
                logger.error(f"Invalid hash error for user {user_id}")
                return False, "Password verification failed"
            except Exception as e:
                logger.error(f"Password change failed for user {user_id}: {e}")
                return False, f"Failed to change password: {str(e)}"

    def reset_user_password(self, user_id: int, new_password: str = None) -> Tuple[bool, str]:
        """
        Reset user password. If new_password not provided, generate random.
        
        Args:
            user_id: User ID to reset password for
            new_password: Optional new password (if None, auto-generate)
        
        Returns:
            Tuple[bool, str]: (success, password_or_error_message)
        """
        import string
        import secrets
        
        try:
            # Generate password if not provided
            if not new_password:
                alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
                new_password = ''.join(secrets.choice(alphabet) for _ in range(12))
            
            # Hash the password
            hashed = bcrypt.hashpw(
                new_password.encode('utf-8'), 
                bcrypt.gensalt()
            ).decode('utf-8')
            
            # Update in database
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute(
                    "UPDATE users SET password = ? WHERE id = ?", 
                    (hashed, user_id)
                )
                conn.commit()
                
                if cursor.rowcount == 0:
                    return False, "User not found"
                
                # Log the activity
                self.log_user_activity(
                    user_id, 
                    'password_reset', 
                    'Password was reset by admin'
                )
                
                return True, new_password
                
        except bcrypt.InvalidHashError:
            logger.error(f"Invalid hash error for user {user_id}")
            return False, "Password hashing failed"
        except Exception as e:
            logger.error(f"Password reset failed for user {user_id}: {e}")
            return False, f"Failed to reset password: {str(e)}"


    def delete_user(self, user_id: int) -> bool:
        """Hard delete user (only allowed for non-admin users)"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            try:
                cursor.execute("SELECT role FROM users WHERE id = ?", (user_id,))
                role = cursor.fetchone()
                if role and role.get('role') == 'admin':
                    return False
                cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
                cursor.execute("DELETE FROM subscriptions WHERE user_id = ?", (user_id,))
                return True
            except Exception as e:
                logger.error(f"Delete user failed: {e}")
                return False

    def get_user_by_username(self, username: str) -> Optional[Dict]:
        """Get user by username"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("SELECT id, username, email FROM users WHERE username = ?", (username,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_all_pending_users(self) -> List[Dict]:
        """Get all pending user registrations across all companies (for system admin)"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                SELECT u.id, u.username, u.email, u.full_name, u.phone, u.role, 
                    u.created_at, u.created_by, c.company_name, u.is_approved, u.is_active
                FROM users u
                LEFT JOIN companies c ON u.company_id = c.id
                WHERE u.is_approved = 0
                ORDER BY u.created_at ASC
            ''')
            return [dict(row) for row in cursor.fetchall()]

    def update_user_role(self, user_id: int, new_role: str, updated_by: int) -> bool:
        """Update user role"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('UPDATE users SET role = ? WHERE id = ?', (new_role, user_id))
            return True

    def update_user_status(self, user_id: int, is_active: bool) -> bool:
        """Activate/deactivate user"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('UPDATE users SET is_active = ? WHERE id = ?', (1 if is_active else 0, user_id))
            return True                

    # ==================== BATCH 2: SUBSCRIPTION & RATE LIMIT METHODS ====================

    def get_effective_subscription(self, user_id: int, company_id: Optional[int] = None) -> Dict[str, Any]:
        """Resolve subscription with NULL-safe queries"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                # Priority 1: Company subscription (latest active)
                if company_id:
                    cursor.execute('''
                        SELECT plan, status, analyses_used, analyses_limit,
                            max_boq_generations, max_bid_optimizations,
                            can_export_data, can_edit_rates, can_delete_rates,
                            can_create_versions, can_manage_team,
                            end_date
                        FROM subscriptions 
                        WHERE company_id = ? AND status = 'active'
                        ORDER BY id DESC
                        LIMIT 1
                    ''', (company_id,))
                    row = cursor.fetchone()
                    if row and row.get('status') == 'active':
                        return {
                            'owner_type': 'company',
                            'owner_id': company_id,
                            'plan': row.get('plan', 'free'),
                            'status': 'active',
                            'analyses_used': row.get('analyses_used', 0),
                            'analyses_limit': row.get('analyses_limit', 5),
                            'max_boq_generations': row.get('max_boq_generations', 5),
                            'max_bid_optimizations': row.get('max_bid_optimizations', 5),
                            'can_export_data': row.get('can_export_data', False),
                            'can_edit_rates': row.get('can_edit_rates', False),
                            'can_delete_rates': row.get('can_delete_rates', False),
                            'can_create_versions': row.get('can_create_versions', False),
                            'can_manage_team': row.get('can_manage_team', False),
                            'end_date': row.get('end_date')
                        }
                
                # Priority 2: User subscription
                cursor.execute('''
                    SELECT plan, status, analyses_used, analyses_limit,
                        max_boq_generations, max_bid_optimizations,
                        can_export_data, can_edit_rates, can_delete_rates,
                        can_create_versions, can_manage_team,
                        end_date
                    FROM subscriptions 
                    WHERE user_id = ? AND status = 'active'
                    ORDER BY id DESC
                    LIMIT 1
                ''', (user_id,))
                row = cursor.fetchone()
                if row and row.get('status') == 'active':
                    return {
                        'owner_type': 'user',
                        'owner_id': user_id,
                        'plan': row.get('plan', 'free'),
                        'status': 'active',
                        'analyses_used': row.get('analyses_used', 0),
                        'analyses_limit': row.get('analyses_limit', 5),
                        'max_boq_generations': row.get('max_boq_generations', 5),
                        'max_bid_optimizations': row.get('max_bid_optimizations', 5),
                        'can_export_data': row.get('can_export_data', False),
                        'can_edit_rates': row.get('can_edit_rates', False),
                        'can_delete_rates': row.get('can_delete_rates', False),
                        'can_create_versions': row.get('can_create_versions', False),
                        'can_manage_team': row.get('can_manage_team', False),
                        'end_date': row.get('end_date')
                    }
                    
            return {
                'owner_type': 'free',
                'plan': 'free',
                'status': 'active',
                'analyses_used': 0,
                'analyses_limit': 5,
                'max_boq_generations': 5,
                'max_bid_optimizations': 5,
                'can_export_data': False,
                'can_edit_rates': False,
                'can_delete_rates': False,
                'can_create_versions': False,
                'can_manage_team': False
            }
                    
        except Exception as e:
            logger.error(f"Subscription lookup error: {e}")
            return {
                'owner_type': 'free',
                'plan': 'free',
                'analyses_used': 0,
                'analyses_limit': 5,
                'max_boq_generations': 5,
                'max_bid_optimizations': 5,
                'can_export_data': False,
                'can_edit_rates': False,
                'can_delete_rates': False,
                'can_create_versions': False,
                'can_manage_team': False
            }
    

    def get_all_subscriptions(self):
        """Get all subscriptions for admin - returns TUPLES for backward compatibility"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                SELECT 
                    s.id, s.user_id, s.plan, s.status, s.start_date, s.end_date, 
                    s.analyses_used, s.analyses_limit, s.company_id, s.created_at,
                    COALESCE(u.username, 'N/A') as username,
                    COALESCE(u.email, 'N/A') as email,
                    COALESCE(u.full_name, 'N/A') as full_name,
                    COALESCE(c.company_name, 'N/A') as company_name
                FROM subscriptions s
                LEFT JOIN users u ON s.user_id = u.id
                LEFT JOIN companies c ON s.company_id = c.id
                ORDER BY s.updated_at DESC
            ''')
            rows = cursor.fetchall()
            
            # Return as TUPLES
            return [(
                row['id'],           # 0
                row['user_id'],      # 1
                row['plan'],         # 2
                row['status'],       # 3
                row['start_date'],   # 4
                row['end_date'],     # 5
                row['analyses_used'],# 6
                row['analyses_limit'],# 7
                row['company_id'],   # 8
                row['created_at'],   # 9
                row['username'],     # 10
                row['email'],        # 11
                row['full_name'],    # 12
                row['company_name']  # 13
            ) for row in rows]

    def can_perform_analysis(self, user_id: int) -> tuple:
        """Check if user can perform an analysis"""
        sub = self.get_user_subscription(user_id)
        if sub.get('analyses_limit', 5) == -1:
            return True, "Unlimited"
        remaining = sub.get('analyses_limit', 5) - sub.get('analyses_used', 0)
        return remaining > 0, remaining

    def increment_analysis_usage(self, user_id: int, company_id: Optional[int] = None) -> bool:
        """Increment usage on the active subscription (company first, then personal)"""
        try:
            sub = self.get_effective_subscription(user_id, company_id)
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                if sub['owner_type'] == 'company':
                    cursor.execute('''
                        UPDATE subscriptions SET analyses_used = analyses_used + 1, updated_at = CURRENT_TIMESTAMP
                        WHERE company_id = ?
                    ''', (sub['owner_id'],))
                elif sub['owner_type'] == 'user':
                    cursor.execute('''
                        UPDATE subscriptions SET analyses_used = analyses_used + 1, updated_at = CURRENT_TIMESTAMP
                        WHERE user_id = ?
                    ''', (sub['owner_id'],))
            return True
        except Exception as e:
            logger.error(f"Usage increment error: {e}")
            return False    

    # ==================== BATCH 3: COMPETITOR & HISTORICAL TENDER METHODS ====================



    def delete_competitor_bak(self, competitor_id: int) -> bool:
        """Soft delete competitor (mark inactive)"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('UPDATE competitor_master SET is_active = 0 WHERE id = ?', (competitor_id,))
            return True

    def update_competitor_stats_from_bid_bak(self, company_id: int, competitor_name: str, 
                                        bid_ratio: float, was_winner: bool) -> bool:
        """Update competitor statistics from a bid"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                SELECT id, total_bids, total_wins, avg_bid_ratio 
                FROM competitor_master 
                WHERE company_id = ? AND competitor_name = ?
            ''', (company_id, competitor_name))
            existing = cursor.fetchone()
            
            if existing:
                comp_id = existing['id']
                total_bids = existing['total_bids']
                total_wins = existing['total_wins']
                avg_ratio = existing['avg_bid_ratio']
                new_total_bids = total_bids + 1
                new_total_wins = total_wins + (1 if was_winner else 0)
                new_avg_ratio = (avg_ratio * total_bids + bid_ratio) / new_total_bids if new_total_bids > 0 else bid_ratio
                
                cursor.execute('''
                    UPDATE competitor_master 
                    SET total_bids = ?, total_wins = ?, avg_bid_ratio = ?,
                        last_seen = ?, updated_at = ?
                    WHERE id = ?
                ''', (new_total_bids, new_total_wins, new_avg_ratio, 
                    datetime.now().date(), datetime.now(), comp_id))
            else:
                cursor.execute('''
                    INSERT INTO competitor_master (
                        company_id, competitor_name, first_seen, last_seen,
                        total_bids, total_wins, avg_bid_ratio
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (company_id, competitor_name, datetime.now().date(), datetime.now().date(),
                    1, 1 if was_winner else 0, bid_ratio))
            return True

    def get_competitor_performance_against_us(self, company_id: int, competitor_name: str = None) -> List[Dict]:
        """Get competitor performance statistics against our company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            query = '''
                SELECT 
                    winning_competitor,
                    COUNT(*) as times_won,
                    AVG(awarded_price / official_estimate) as avg_winning_ratio
                FROM historical_tenders 
                WHERE company_id = ? 
                AND winning_company_type = 'Competitor'
            '''
            params = [company_id]
            if competitor_name:
                query += " AND winning_competitor = ?"
                params.append(competitor_name)
            query += " GROUP BY winning_competitor ORDER BY times_won DESC"
            cursor.execute(query, params)
            rows = cursor.fetchall()
            return [{'competitor': row['winning_competitor'], 'wins': row['times_won'], 
                    'avg_ratio': row['avg_winning_ratio']} for row in rows]

    def get_historical_tenders_with_winner(self, company_id: int, procurement_type: str = None, 
                                            winner_type: str = None, limit: int = 100) -> pd.DataFrame:
        """Get historical tenders with winner filtering"""
        import pandas as pd
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            query = 'SELECT * FROM historical_tenders WHERE company_id = ?'
            params = [company_id]
            if procurement_type:
                query += " AND procurement_type = ?"
                params.append(procurement_type)
            if winner_type and winner_type != "All":
                query += " AND winning_company_type = ?"
                params.append(winner_type)
            query += " ORDER BY award_date DESC LIMIT ?"
            params.append(limit)
            cursor.execute(query, params)
            rows = cursor.fetchall()
            if rows:
                df = pd.DataFrame([dict(row) for row in rows])
                return df
            return pd.DataFrame()

    def get_winner_trends(self, company_id: int, procurement_type: str = None, months: int = 12) -> pd.DataFrame:
        """Get winner trends over time"""
        import pandas as pd
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            query = '''
                SELECT 
                    strftime('%Y-%m', award_date) as month,
                    COUNT(*) as total,
                    SUM(CASE WHEN winning_company_type = 'Our Company' THEN 1 ELSE 0 END) as our_wins,
                    SUM(CASE WHEN winning_company_type = 'Competitor' THEN 1 ELSE 0 END) as competitor_wins
                FROM historical_tenders 
                WHERE company_id = ?
                AND award_date >= date('now', ?)
            '''
            params = [company_id, f'-{months} months']
            if procurement_type:
                query += " AND procurement_type = ?"
                params.append(procurement_type)
            query += " GROUP BY strftime('%Y-%m', award_date) ORDER BY month DESC"
            cursor.execute(query, params)
            rows = cursor.fetchall()
            if rows:
                return pd.DataFrame([dict(row) for row in rows])
            return pd.DataFrame()

    def update_historical_tender_winner(self, tender_id: int, winner_name: str, 
                                        winner_type: str, winning_price: float) -> bool:
        """Update winner information for an existing historical tender"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                UPDATE historical_tenders 
                SET winning_competitor = ?, winning_company_type = ?, awarded_price = ?
                WHERE id = ?
            ''', (winner_name, winner_type, winning_price, tender_id))
            return True

    def get_nppi_for_company(self, company_id: int, procurement_type: str = 'goods') -> Optional[Dict]:
        """Get the latest NPPI for a company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                SELECT nppi_factor, calculation_date, data_points
                FROM company_nppi
                WHERE company_id = ? AND procurement_type = ?
                ORDER BY calculation_date DESC LIMIT 1
            ''', (company_id, procurement_type))
            row = cursor.fetchone()
            return dict(row) if row else None

    def save_company_nppi(self, company_id: int, procurement_type: str, nppi_factor: float, data_points: int) -> bool:
        """Save calculated NPPI for a company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                INSERT INTO company_nppi (company_id, procurement_type, nppi_factor, data_points)
                VALUES (?, ?, ?, ?)
            ''', (company_id, procurement_type, nppi_factor, data_points))
            return True    
        
    # ==================== BATCH 4: CONTACT, ACTIVITY LOG & CONSULTANT METHODS ====================

    def save_contact_message(self, name: str, email: str, subject: str, message: str) -> bool:
        """Save contact message"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                INSERT INTO contact_messages (name, email, subject, message)
                VALUES (?, ?, ?, ?)
            ''', (name, email, subject, message))
            return True

    def log_team_activity(self, company_id: int, actor_user_id: int, 
                        action_type: str, target_type: str, target_id: str,
                        details: str = None) -> bool:
        """Log team management activity for audit trail."""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute('''
                    INSERT INTO activity_logs (
                        company_id, actor_user_id, action_type, target_type, 
                        target_id, details, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    company_id,
                    actor_user_id,
                    action_type,
                    target_type,
                    target_id,
                    details,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))
            return True
        except Exception as e:
            logger.error(f"Failed to log activity: {e}")
            return False

    def add_consultant_client(self, consultant_id: int, client_company_id: int, role: str = 'manager') -> bool:
        """Link a consultant to a client company"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute('''
                    INSERT OR IGNORE INTO consultant_clients (consultant_user_id, client_company_id, role)
                    VALUES (?, ?, ?)
                ''', (consultant_id, client_company_id, role))
            return True
        except Exception as e:
            logger.error(f"Client relationship error: {e}")
            return False

    def get_consultant_clients(self, consultant_id: int) -> List[Dict]:
        """Fetch all client companies linked to a consultant"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute('''
                    SELECT c.id, c.company_name, c.email, cc.role, cc.created_at
                    FROM consultant_clients cc
                    JOIN companies c ON cc.client_company_id = c.id
                    WHERE cc.consultant_user_id = ?
                ''', (consultant_id,))
                rows = cursor.fetchall()
                return [{'id': row['id'], 'company_name': row['company_name'], 
                        'email': row['email'], 'role': row['role']} for row in rows]
        except Exception as e:
            logger.error(f"Fetch consultant clients error: {e}")
            return []

    def create_individual_user(self, user_data: Dict) -> tuple:
        """Create an individual user (no company)"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                personal_company_name = f"{user_data['full_name']} (Individual)"
                
                cursor.execute('''
                    INSERT INTO companies (company_name, email, phone, division, is_active)
                    VALUES (?, ?, ?, ?, ?)
                ''', (personal_company_name, user_data['email'], user_data.get('phone', ''), 'Dhaka', 1))
                company_id = cursor.lastrowid
                
                hashed = bcrypt.hashpw(user_data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                cursor.execute('''
                    INSERT INTO users (company_id, username, password, email, full_name, phone, 
                                    role, account_type, status, is_approved, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    company_id,
                    user_data['username'],
                    hashed,
                    user_data['email'],
                    user_data['full_name'],
                    user_data.get('phone', ''),
                    user_data.get('role', 'individual'),
                    'individual',
                    'active',
                    1,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))
                user_id = cursor.lastrowid
            return True, user_id
        except Exception as e:
            logger.error(f"Individual user creation failed: {e}")
            return False, str(e)

    def generate_random_password(self, length: int = 12) -> str:
        """Generate a secure random password"""
        import string
        import secrets
        alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
        return ''.join(secrets.choice(alphabet) for _ in range(length))
    
    # ==================== BATCH 5: TENDER & COMPANY CRUD METHODS ====================

    def get_tender_analyses_by_company(self, company_id: int, tender_id: str = None) -> List[Dict]:
        """Get all analyses for a company, optionally for a specific tender"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            if tender_id:
                cursor.execute('''
                    SELECT id, tender_title, recommended_bid, success_probability, 
                        risk_level, analysis_date, user_id
                    FROM tender_analyses
                    WHERE company_id = ? AND tender_id = ?
                    ORDER BY analysis_date DESC
                ''', (company_id, tender_id))
            else:
                cursor.execute('''
                    SELECT id, tender_id, tender_title, recommended_bid, success_probability,
                        risk_level, analysis_date
                    FROM tender_analyses
                    WHERE company_id = ?
                    ORDER BY analysis_date DESC
                    LIMIT 20
                ''', (company_id,))
            rows = cursor.fetchall()
            return [dict(row) for row in rows]

    def add_tender_lot(self, tender_id: int, lot_data: Dict) -> bool:
        """Add lot information for a tender"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('''
                INSERT INTO tender_lots (tender_id, lot_no, lot_description, location, 
                                        security_amount, estimated_value, start_date, completion_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (tender_id, lot_data.get('lot_no'), lot_data.get('description'),
                lot_data.get('location'), lot_data.get('security_amount', 0),
                lot_data.get('estimated_value', 0), lot_data.get('start_date'),
                lot_data.get('completion_date')))
            return True

    def update_tender(self, tender_id: int, tender_data: Dict, user_id: int) -> bool:
        """Update an existing tender with full e-GP fields"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)

            cursor.execute("""
                UPDATE company_tenders SET
                    tender_id = ?,
                    tender_title = ?,
                    procuring_entity = ?,
                    division = ?,
                    district = ?,
                    thana = ?,
                    country = ?,
                    procurement_type = ?,
                    official_estimate = ?,
                    submission_deadline = ?,
                    tender_security = ?,
                    document_fee = ?,
                    evaluation_type = ?,
                    mode_of_payment = ?,
                    eligibility_criteria = ?,
                    invitation_ref_no = ?,
                    package_no = ?,
                    project_code = ?,
                    project_name = ?,
                    inviting_official_name = ?,
                    inviting_official_designation = ?,
                    inviting_official_phone = ?,
                    inviting_official_email = ?,
                    inviting_official_address = ?,
                    inviting_official_city = ?,
                    inviting_official_thana = ?,
                    inviting_official_district = ?,
                    notes = ?,
                    app_id = ?,
                    procuring_entity_code = ?,
                    procurement_nature = ?,
                    event_type = ?,
                    budget_type = ?,
                    source_of_funds = ?,
                    category = ?,
                    tender_publication_date = ?,
                    document_selling_end_date = ?,
                    pre_bid_meeting_start = ?,
                    pre_bid_meeting_end = ?,
                    bid_opening_date = ?,
                    security_submission_deadline = ?,
                    security_valid_upto = ?,
                    tender_valid_upto = ?,
                    updated_at = CURRENT_TIMESTAMP,
                    created_by = ?
                WHERE id = ?
            """, (
                tender_data.get('tender_id'),           # 1
                tender_data.get('tender_title'),        # 2
                tender_data.get('procuring_entity'),    # 3
                tender_data.get('division'),            # 4
                tender_data.get('district'),            # 5
                tender_data.get('thana'),               # 6
                tender_data.get('country', 'Bangladesh'), # 7
                tender_data.get('procurement_type'),    # 8
                tender_data.get('official_estimate'),   # 9
                tender_data.get('submission_deadline'), # 10
                tender_data.get('tender_security'),     # 11
                tender_data.get('document_fee'),        # 12
                tender_data.get('evaluation_type'),     # 13
                tender_data.get('mode_of_payment'),     # 14
                tender_data.get('eligibility_criteria'), # 15
                tender_data.get('invitation_ref_no'),   # 16
                tender_data.get('package_no'),          # 17
                tender_data.get('project_code'),        # 18
                tender_data.get('project_name'),        # 19
                tender_data.get('inviting_official_name'), # 20
                tender_data.get('inviting_official_designation'), # 21
                tender_data.get('inviting_official_phone'), # 22
                tender_data.get('inviting_official_email'), # 23
                tender_data.get('inviting_official_address'), # 24
                tender_data.get('inviting_official_city'), # 25
                tender_data.get('inviting_official_thana'), # 26
                tender_data.get('inviting_official_district'), # 27
                tender_data.get('notes'),               # 28
                tender_data.get('app_id'),              # 29
                tender_data.get('procuring_entity_code'), # 30
                tender_data.get('procurement_nature'),  # 31
                tender_data.get('event_type'),          # 32
                tender_data.get('budget_type'),         # 33
                tender_data.get('source_of_funds'),     # 34
                tender_data.get('category'),            # 35
                tender_data.get('tender_publication_date'), # 36
                tender_data.get('document_selling_end_date'), # 37
                tender_data.get('pre_bid_meeting_start'), # 38
                tender_data.get('pre_bid_meeting_end'), # 39
                tender_data.get('bid_opening_date'),    # 40
                tender_data.get('security_submission_deadline'), # 41
                tender_data.get('security_valid_upto'), # 42
                tender_data.get('tender_valid_upto'),   # 43
                user_id,                                # 44 (created_by)
                tender_id                               # 45 (WHERE clause)
            ))

            return cursor.rowcount > 0

    def update_tender_bak(self, tender_id: int, tender_data: Dict, user_id: int) -> bool:
        """
        Update an existing tender
        
        Args:
            tender_id: The database ID of the tender to update
            tender_data: Dictionary containing tender fields to update
            user_id: The ID of the user performing the update
        
        Returns:
            bool: True if update was successful, False otherwise
        """
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                # Define all fields that can be updated
                allowed_fields = [
                    'tender_id', 'tender_title', 'procuring_entity', 'division',
                    'procurement_type', 'official_estimate', 'submission_deadline',
                    'tender_security', 'document_fee', 'project_code', 'project_name',
                    'package_no', 'budget_type', 'notes', 'is_active',
                    'district', 'thana', 'evaluation_type', 'mode_of_payment',
                    'invitation_ref_no', 'winning_competitor', 'winning_bid_amount',
                    'total_bidders', 'evaluation_status'
                ]
                
                # Build update query
                update_fields = []
                update_values = []
                
                for field in allowed_fields:
                    if field in tender_data:
                        update_fields.append(f"{field} = ?")
                        update_values.append(tender_data[field])
                
                if not update_fields:
                    return False
                
                # Add updated_at and updated_by
                update_fields.append("updated_at = ?")
                update_values.append(datetime.now())
                update_fields.append("updated_by = ?")
                update_values.append(user_id)
                
                # Add tender_id for WHERE clause
                update_values.append(tender_id)
                
                query = f"""
                    UPDATE company_tenders 
                    SET {', '.join(update_fields)}
                    WHERE id = ?
                """
                
                cursor.execute(query, update_values)
                conn.commit()
                
                # Log the update
                logger.info(f"Tender {tender_id} updated with fields: {list(tender_data.keys())}")
                
                return cursor.rowcount > 0
                
        except Exception as e:
            logger.error(f"Error updating tender {tender_id}: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
  
    def update_tender_lock_status(self, tender_id: int, locked: bool) -> bool:
        """Update the lock status of a tender"""
        import streamlit as st
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute('''
                    UPDATE company_tenders 
                    SET is_locked = ?, locked_at = ?, locked_by = ?
                    WHERE id = ?
                ''', (
                    1 if locked else 0,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S') if locked else None,
                    st.session_state.get('user_id') if locked else None,
                    tender_id
                ))
            return True
        except Exception as e:
            logger.error(f"Failed to update tender lock status: {e}")
            return False

    def create_tender_copy(self, original_tender_id: int, created_by: int) -> Optional[int]:
        """Create a backup copy of a locked tender"""
        import streamlit as st
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute('SELECT * FROM company_tenders WHERE id = ?', (original_tender_id,))
                original = cursor.fetchone()
                if not original:
                    return None
                
                cursor.execute('''
                    INSERT INTO company_tenders (
                        company_id, tender_id, tender_title, procuring_entity, official_estimate,
                        submission_deadline, procurement_type, division, district, thana,
                        tender_security, document_fee, evaluation_type, created_at,
                        is_locked, is_copy, original_tender_id, created_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    original.get('company_id'),
                    f"{original.get('tender_id')}_COPY",
                    f"{original.get('tender_title')} (Backup Copy)",
                    original.get('procuring_entity'), original.get('official_estimate'),
                    original.get('submission_deadline'), original.get('procurement_type'),
                    original.get('division'), original.get('district'), original.get('thana'),
                    original.get('tender_security'), original.get('document_fee'),
                    original.get('evaluation_type'), datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    0, 1, original_tender_id, created_by
                ))
                return cursor.lastrowid
        except Exception as e:
            logger.error(f"Failed to create tender copy: {e}")
            return None

    def delete_tender(self, tender_id: int) -> bool:
        """Soft delete a tender (mark as inactive)"""
        import streamlit as st
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute('''
                    UPDATE company_tenders 
                    SET is_active = 0, deleted_at = ?, deleted_by = ?
                    WHERE id = ?
                ''', (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), st.session_state.get('user_id'), tender_id))
            return True
        except Exception as e:
            logger.error(f"Failed to delete tender: {e}")
            return False

    def get_company_stats_by_id(self, company_id: int) -> Dict:
        """Get statistics for a specific company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                # Get total users
                cursor.execute('SELECT COUNT(*) as total FROM users WHERE company_id = ?', (company_id,))
                row = cursor.fetchone()
                total_users = row['total'] if row else 0
                print(f"📊 total_users: {total_users}")
                
                # Get total analyses
                cursor.execute('SELECT COUNT(*) as total FROM tender_analyses WHERE company_id = ?', (company_id,))
                row = cursor.fetchone()
                total_analyses = row['total'] if row else 0
                print(f"📊 total_analyses: {total_analyses}")
                
                # Get won tenders
                cursor.execute('''
                    SELECT COUNT(*) as total FROM tender_analyses 
                    WHERE company_id = ? AND bid_status = 'Won'
                ''', (company_id,))
                row = cursor.fetchone()
                won_tenders = row['total'] if row else 0
                print(f"📊 won_tenders: {won_tenders}")
                
                win_rate = (won_tenders / total_analyses * 100) if total_analyses > 0 else 0
                print(f"📊 win_rate: {win_rate:.1f}%")
                
                return {
                    'total_users': total_users,
                    'total_analyses': total_analyses,
                    'won_tenders': won_tenders,
                    'win_rate': win_rate
                }
                
            except Exception as e:
                print(f"❌ Error in get_company_stats_by_id: {e}")
                import traceback
                traceback.print_exc()
                return {
                    'total_users': 0,
                    'total_analyses': 0,
                    'won_tenders': 0,
                    'win_rate': 0
                }

    def get_company_profile(self, company_id: int) -> Optional[Dict]:
        """Get company profile information"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id, company_name, registration_number, vat_number, address,
                    district, division, phone, email, website, created_at, is_active
                FROM companies 
                WHERE id = ?
            """, (company_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def update_company_profile(self, company_id: int, profile_data: Dict) -> bool:
        """Update company profile information"""
        allowed_fields = ['company_name', 'registration_number', 'vat_number', 'address',
                        'district', 'division', 'phone', 'email', 'website']
        updates = []
        values = []
        for key, value in profile_data.items():
            if key in allowed_fields and value is not None:
                updates.append(f"{key} = ?")
                values.append(value)
        if not updates:
            return False
        values.append(company_id)
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute(f"UPDATE companies SET {', '.join(updates)} WHERE id = ?", values)
            return True

    # =========================================================
    # COMPANY PROFILE MANAGEMENT
    # =========================================================

    def save_company_profile(self, company_id: int, profile_data: Dict) -> bool:
        """Save or update company profile"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                cursor.execute("""
                    INSERT OR REPLACE INTO company_profile (
                        company_id, legal_name, trade_name, registration_number,
                        date_of_incorporation, business_nature, business_category,
                        registered_address, corporate_address, phone_primary,
                        phone_secondary, email_primary, email_secondary, website,
                        fax, division, district, upazila, post_code, status,
                        updated_by, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (
                    company_id,
                    profile_data.get('legal_name'),
                    profile_data.get('trade_name'),
                    profile_data.get('registration_number'),
                    profile_data.get('date_of_incorporation'),
                    profile_data.get('business_nature'),
                    profile_data.get('business_category'),
                    profile_data.get('registered_address'),
                    profile_data.get('corporate_address'),
                    profile_data.get('phone_primary'),
                    profile_data.get('phone_secondary'),
                    profile_data.get('email_primary'),
                    profile_data.get('email_secondary'),
                    profile_data.get('website'),
                    profile_data.get('fax'),
                    profile_data.get('division'),
                    profile_data.get('district'),
                    profile_data.get('upazila'),
                    profile_data.get('post_code'),
                    profile_data.get('status', 'active'),
                    profile_data.get('updated_by')
                ))
                return True
            except Exception as e:
                logger.error(f"Error saving company profile: {e}")
                return False


    # =========================================================
    # DOCUMENT MANAGEMENT WITH VERSIONING
    # =========================================================

    def add_document(self, company_id: int, document_data: Dict, file_content: bytes = None) -> Optional[int]:
        """
        Add a new document with version control
        """
        import uuid
        import hashlib
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                doc_uuid = str(uuid.uuid4())
                file_hash = None
                extracted_text = document_data.get('extracted_text', '')
                
                if file_content:
                    file_hash = hashlib.sha256(file_content).hexdigest()
                
                cursor.execute("""
                    INSERT INTO document_registry (
                        company_id, document_uuid, document_name, document_type,
                        reference_id, reference_table, version_number, is_latest_version,
                        file_path, file_name, file_size, file_hash, mime_type,
                        extracted_text, description, tags, category, language,
                        document_date, expiry_date, effective_date, verification_status,
                        uploaded_by, uploaded_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (
                    company_id, doc_uuid, document_data.get('document_name'),
                    document_data.get('document_type'), document_data.get('reference_id'),
                    document_data.get('reference_table'), 1, 1,
                    document_data.get('file_path'), document_data.get('file_name'),
                    document_data.get('file_size'), file_hash, document_data.get('mime_type'),
                    extracted_text, document_data.get('description'),
                    json.dumps(document_data.get('tags', [])), document_data.get('category'),
                    document_data.get('language', 'en'), document_data.get('document_date'),
                    document_data.get('expiry_date'), document_data.get('effective_date'),
                    document_data.get('verification_status', 'pending'),
                    document_data.get('uploaded_by')
                ))
                
                doc_id = cursor.lastrowid
                
                # Index for search
                self._index_document_for_search(doc_id, extracted_text, document_data, company_id)
                
                return doc_id
                
            except Exception as e:
                logger.error(f"Error adding document: {e}")
                return None

    def update_document(self, document_id: int, new_version_data: Dict, file_content: bytes = None) -> Optional[int]:
        """Create a new version of an existing document"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                # Get current document
                cursor.execute("SELECT * FROM document_registry WHERE id = ?", (document_id,))
                current = cursor.fetchone()
                
                if not current:
                    return None
                
                current_dict = dict(current)
                
                # Mark current as not latest
                cursor.execute("UPDATE document_registry SET is_latest_version = 0 WHERE id = ?", (document_id,))
                
                # Create new version
                new_doc_id = self.add_document(
                    current_dict.get('company_id'),
                    {
                        'document_name': new_version_data.get('document_name', current_dict.get('document_name')),
                        'document_type': current_dict.get('document_type'),
                        'reference_id': current_dict.get('reference_id'),
                        'reference_table': current_dict.get('reference_table'),
                        'file_path': new_version_data.get('file_path', current_dict.get('file_path')),
                        'file_name': new_version_data.get('file_name', current_dict.get('file_name')),
                        'file_size': new_version_data.get('file_size', current_dict.get('file_size')),
                        'mime_type': new_version_data.get('mime_type', current_dict.get('mime_type')),
                        'description': new_version_data.get('description', current_dict.get('description')),
                        'tags': new_version_data.get('tags', json.loads(current_dict.get('tags') or '[]')),
                        'category': new_version_data.get('category', current_dict.get('category')),
                        'document_date': new_version_data.get('document_date', current_dict.get('document_date')),
                        'expiry_date': new_version_data.get('expiry_date', current_dict.get('expiry_date')),
                        'effective_date': new_version_data.get('effective_date', current_dict.get('effective_date')),
                        'uploaded_by': new_version_data.get('uploaded_by', current_dict.get('uploaded_by')),
                        'extracted_text': new_version_data.get('extracted_text', current_dict.get('extracted_text'))
                    },
                    file_content
                )
                
                # Link versions
                if new_doc_id:
                    cursor.execute("UPDATE document_registry SET previous_version_id = ? WHERE id = ?", (document_id, new_doc_id))
                
                return new_doc_id
                
            except Exception as e:
                logger.error(f"Error updating document: {e}")
                return None

    def _index_document_for_search(self, doc_id: int, content: str, metadata: Dict, company_id: int):
        """Index document for full-text search"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute("""
                    INSERT OR REPLACE INTO fts_documents (
                        company_id, document_uuid, entity_type, entity_id,
                        field_name, content, metadata
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    company_id,
                    metadata.get('document_uuid'),
                    'document',
                    doc_id,
                    'full_text',
                    content[:5000] if content else '',
                    json.dumps({'document_name': metadata.get('document_name'), 'tags': metadata.get('tags')})
                ))
        except Exception as e:
            logger.error(f"Error indexing document: {e}")


    # =========================================================
    # EXPERIENCE & PROJECT MANAGEMENT
    # =========================================================

    def add_experience(self, company_id: int, experience_data: Dict) -> Optional[int]:
        """Add experience record"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                cursor.execute("""
                    INSERT INTO experience_record (
                        company_id, project_name, project_location, client_name,
                        client_type, procuring_entity, contract_number, contract_date,
                        completion_date, contract_value, currency, nature_of_work,
                        scope_of_work, key_deliverables, is_completed, is_running,
                        completion_percentage, quality_rating, safety_rating,
                        client_satisfaction, project_manager, site_engineer,
                        contract_document_path, completion_certificate_path,
                        performance_certificate_path, created_by, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (
                    company_id,
                    experience_data.get('project_name'),
                    experience_data.get('project_location'),
                    experience_data.get('client_name'),
                    experience_data.get('client_type'),
                    experience_data.get('procuring_entity'),
                    experience_data.get('contract_number'),
                    experience_data.get('contract_date'),
                    experience_data.get('completion_date'),
                    experience_data.get('contract_value'),
                    experience_data.get('currency', 'BDT'),
                    experience_data.get('nature_of_work'),
                    experience_data.get('scope_of_work'),
                    experience_data.get('key_deliverables'),
                    1 if experience_data.get('is_completed') else 0,
                    1 if experience_data.get('is_running') else 0,
                    experience_data.get('completion_percentage', 0),
                    experience_data.get('quality_rating'),
                    experience_data.get('safety_rating'),
                    experience_data.get('client_satisfaction'),
                    experience_data.get('project_manager'),
                    experience_data.get('site_engineer'),
                    experience_data.get('contract_document_path'),
                    experience_data.get('completion_certificate_path'),
                    experience_data.get('performance_certificate_path'),
                    experience_data.get('created_by')
                ))
                
                return cursor.lastrowid
                
            except Exception as e:
                logger.error(f"Error adding experience: {e}")
                return None

    def get_experiences(self, company_id: int, limit: int = 100) -> pd.DataFrame:
        """Get all experiences for a company"""
        import pandas as pd
        with self.get_connection() as conn:
            try:
                df = pd.read_sql_query("""
                    SELECT * FROM experience_record 
                    WHERE company_id = ? 
                    ORDER BY completion_date DESC
                    LIMIT ?
                """, conn, params=[company_id, limit])
                return df
            except Exception as e:
                logger.error(f"Error getting experiences: {e}")
                return pd.DataFrame()

    def get_experience_by_id(self, experience_id: int) -> Optional[Dict]:
        """Get experience by ID"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("SELECT * FROM experience_record WHERE id = ?", (experience_id,))
            row = cursor.fetchone()
            return dict(row) if row else None


    # =========================================================
    # PERSONNEL MANAGEMENT
    # =========================================================

    def add_personnel(self, company_id: int, personnel_data: Dict) -> Optional[int]:
        """Add personnel record"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                cursor.execute("""
                    INSERT INTO personnel (
                        company_id, full_name, father_name, mother_name, spouse_name,
                        date_of_birth, nationality, nid_number, passport_number,
                        birth_certificate_number, personal_phone, personal_email,
                        present_address, permanent_address, designation, department,
                        employee_id, joining_date, confirmation_date,
                        educational_qualification, professional_certifications,
                        skills, languages, cv_path, photo_path, employment_status,
                        is_key_personnel, created_by, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (
                    company_id,
                    personnel_data.get('full_name'),
                    personnel_data.get('father_name'),
                    personnel_data.get('mother_name'),
                    personnel_data.get('spouse_name'),
                    personnel_data.get('date_of_birth'),
                    personnel_data.get('nationality', 'Bangladeshi'),
                    personnel_data.get('nid_number'),
                    personnel_data.get('passport_number'),
                    personnel_data.get('birth_certificate_number'),
                    personnel_data.get('personal_phone'),
                    personnel_data.get('personal_email'),
                    personnel_data.get('present_address'),
                    personnel_data.get('permanent_address'),
                    personnel_data.get('designation'),
                    personnel_data.get('department'),
                    personnel_data.get('employee_id'),
                    personnel_data.get('joining_date'),
                    personnel_data.get('confirmation_date'),
                    personnel_data.get('educational_qualification'),
                    json.dumps(personnel_data.get('professional_certifications', [])),
                    json.dumps(personnel_data.get('skills', [])),
                    json.dumps(personnel_data.get('languages', [])),
                    personnel_data.get('cv_path'),
                    personnel_data.get('photo_path'),
                    personnel_data.get('employment_status', 'active'),
                    1 if personnel_data.get('is_key_personnel') else 0,
                    personnel_data.get('created_by')
                ))
                
                return cursor.lastrowid
                
            except Exception as e:
                logger.error(f"Error adding personnel: {e}")
                return None

    def get_personnel(self, company_id: int, limit: int = 100) -> pd.DataFrame:
        """Get all personnel for a company"""
        import pandas as pd
        with self.get_connection() as conn:
            try:
                df = pd.read_sql_query("""
                    SELECT * FROM personnel 
                    WHERE company_id = ? AND employment_status = 'active'
                    ORDER BY full_name
                    LIMIT ?
                """, conn, params=[company_id, limit])
                return df
            except Exception as e:
                logger.error(f"Error getting personnel: {e}")
                return pd.DataFrame()

    def get_personnel_by_id(self, personnel_id: int) -> Optional[Dict]:
        """Get personnel by ID"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("SELECT * FROM personnel WHERE id = ?", (personnel_id,))
            row = cursor.fetchone()
            return dict(row) if row else None


    # =========================================================
    # EQUIPMENT MANAGEMENT
    # =========================================================

    def add_equipment(self, company_id: int, equipment_data: Dict) -> Optional[int]:
        """Add equipment record"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                cursor.execute("""
                    INSERT INTO equipment (
                        company_id, equipment_name, equipment_type, model,
                        manufacturer, serial_number, capacity, power_rating,
                        fuel_type, year_of_manufacture, country_of_origin,
                        ownership_type, owner_name, registration_number,
                        chassis_number, engine_number, purchase_date,
                        purchase_cost, currency, supplier_name, invoice_number,
                        current_status, location, operator_name, operating_hours,
                        last_maintenance_date, next_maintenance_date, created_by, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (
                    company_id,
                    equipment_data.get('equipment_name'),
                    equipment_data.get('equipment_type'),
                    equipment_data.get('model'),
                    equipment_data.get('manufacturer'),
                    equipment_data.get('serial_number'),
                    equipment_data.get('capacity'),
                    equipment_data.get('power_rating'),
                    equipment_data.get('fuel_type'),
                    equipment_data.get('year_of_manufacture'),
                    equipment_data.get('country_of_origin'),
                    equipment_data.get('ownership_type'),
                    equipment_data.get('owner_name'),
                    equipment_data.get('registration_number'),
                    equipment_data.get('chassis_number'),
                    equipment_data.get('engine_number'),
                    equipment_data.get('purchase_date'),
                    equipment_data.get('purchase_cost'),
                    equipment_data.get('currency', 'BDT'),
                    equipment_data.get('supplier_name'),
                    equipment_data.get('invoice_number'),
                    equipment_data.get('current_status', 'available'),
                    equipment_data.get('location'),
                    equipment_data.get('operator_name'),
                    equipment_data.get('operating_hours', 0),
                    equipment_data.get('last_maintenance_date'),
                    equipment_data.get('next_maintenance_date'),
                    equipment_data.get('created_by')
                ))
                
                return cursor.lastrowid
                
            except Exception as e:
                logger.error(f"Error adding equipment: {e}")
                return None

    def get_equipment(self, company_id: int, limit: int = 100) -> pd.DataFrame:
        """Get all equipment for a company"""
        import pandas as pd
        with self.get_connection() as conn:
            try:
                df = pd.read_sql_query("""
                    SELECT * FROM equipment 
                    WHERE company_id = ?
                    ORDER BY equipment_name
                    LIMIT ?
                """, conn, params=[company_id, limit])
                return df
            except Exception as e:
                logger.error(f"Error getting equipment: {e}")
                return pd.DataFrame()

    def get_equipment_by_id(self, equipment_id: int) -> Optional[Dict]:
        """Get equipment by ID"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("SELECT * FROM equipment WHERE id = ?", (equipment_id,))
            row = cursor.fetchone()
            return dict(row) if row else None


    # =========================================================
    # FINANCIAL CAPACITY MANAGEMENT
    # =========================================================

    def add_financial_capacity(self, company_id: int, financial_data: Dict) -> Optional[int]:
        """Add financial capacity record"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                cursor.execute("""
                    INSERT INTO financial_capacity (
                        company_id, fiscal_year, annual_turnover, construction_turnover,
                        export_turnover, total_assets, current_assets, fixed_assets,
                        total_liabilities, current_liabilities, net_worth,
                        liquid_assets, cash_and_bank, working_capital,
                        current_ratio, quick_ratio, debt_to_equity_ratio,
                        profit_margin, credit_limit, bank_guarantee_limit,
                        overdraft_limit, letter_of_credit_limit, audited_by,
                        audit_firm, audit_report_path, audit_date, is_audited, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (
                    company_id,
                    financial_data.get('fiscal_year'),
                    financial_data.get('annual_turnover'),
                    financial_data.get('construction_turnover'),
                    financial_data.get('export_turnover'),
                    financial_data.get('total_assets'),
                    financial_data.get('current_assets'),
                    financial_data.get('fixed_assets'),
                    financial_data.get('total_liabilities'),
                    financial_data.get('current_liabilities'),
                    financial_data.get('net_worth'),
                    financial_data.get('liquid_assets'),
                    financial_data.get('cash_and_bank'),
                    financial_data.get('working_capital'),
                    financial_data.get('current_ratio'),
                    financial_data.get('quick_ratio'),
                    financial_data.get('debt_to_equity_ratio'),
                    financial_data.get('profit_margin'),
                    financial_data.get('credit_limit'),
                    financial_data.get('bank_guarantee_limit'),
                    financial_data.get('overdraft_limit'),
                    financial_data.get('letter_of_credit_limit'),
                    financial_data.get('audited_by'),
                    financial_data.get('audit_firm'),
                    financial_data.get('audit_report_path'),
                    financial_data.get('audit_date'),
                    1 if financial_data.get('is_audited') else 0
                ))
                
                return cursor.lastrowid
                
            except Exception as e:
                logger.error(f"Error adding financial capacity: {e}")
                return None

    def get_financial_records(self, company_id: int, limit: int = 10) -> pd.DataFrame:
        """Get financial records for a company"""
        import pandas as pd
        with self.get_connection() as conn:
            try:
                df = pd.read_sql_query("""
                    SELECT * FROM financial_capacity 
                    WHERE company_id = ?
                    ORDER BY fiscal_year DESC
                    LIMIT ?
                """, conn, params=[company_id, limit])
                return df
            except Exception as e:
                logger.error(f"Error getting financial records: {e}")
                return pd.DataFrame()


    # =========================================================
    # KEYWORD SEARCH METHOD
    # =========================================================

    def keyword_search(self, company_id: int, query: str, 
                    entity_types: List[str] = None, limit: int = 20) -> List[Dict]:
        """Keyword search across knowledge base"""
        results = []
        search_term = f"%{query}%"
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                # Search personnel
                if not entity_types or 'personnel' in entity_types:
                    cursor.execute("""
                        SELECT 'personnel' as entity_type, id, full_name as name,
                            designation, personal_phone, personal_email
                        FROM personnel
                        WHERE company_id = ? AND employment_status = 'active'
                        AND (full_name LIKE ? OR designation LIKE ? OR skills LIKE ?)
                        LIMIT ?
                    """, (company_id, search_term, search_term, search_term, limit))
                    
                    for row in cursor.fetchall():
                        results.append({
                            'entity_type': row[0],
                            'entity_id': row[1],
                            'content': f"{row[2]} - {row[3]}",
                            'relevance': 0.8
                        })
                
                # Search equipment
                if not entity_types or 'equipment' in entity_types:
                    cursor.execute("""
                        SELECT 'equipment' as entity_type, id, equipment_name as name,
                            equipment_type, model, current_status
                        FROM equipment
                        WHERE company_id = ?
                        AND (equipment_name LIKE ? OR model LIKE ? OR serial_number LIKE ?)
                        LIMIT ?
                    """, (company_id, search_term, search_term, search_term, limit))
                    
                    for row in cursor.fetchall():
                        results.append({
                            'entity_type': row[0],
                            'entity_id': row[1],
                            'content': f"{row[2]} - {row[3]}",
                            'relevance': 0.8
                        })
                
                # Search experiences
                if not entity_types or 'experience' in entity_types:
                    cursor.execute("""
                        SELECT 'experience' as entity_type, id, project_name as name,
                            client_name, nature_of_work
                        FROM experience_record
                        WHERE company_id = ?
                        AND (project_name LIKE ? OR client_name LIKE ? OR nature_of_work LIKE ?)
                        ORDER BY completion_date DESC
                        LIMIT ?
                    """, (company_id, search_term, search_term, search_term, limit))
                    
                    for row in cursor.fetchall():
                        results.append({
                            'entity_type': row[0],
                            'entity_id': row[1],
                            'content': f"{row[2]} - {row[3]}",
                            'relevance': 0.8
                        })
                
                return results[:limit]
                
            except Exception as e:
                logger.error(f"Error in keyword search: {e}")
                return []


    # =========================================================
    # DOCUMENT RETRIEVAL METHODS
    # =========================================================

    def get_documents(self, company_id: int, doc_type: str = None, 
                    show_expired: bool = False, limit: int = 50) -> List[Dict]:
        """Get documents for a company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            query = """
                SELECT * FROM document_registry 
                WHERE company_id = ? AND is_latest_version = 1
            """
            params = [company_id]
            
            if doc_type:
                query += " AND document_type = ?"
                params.append(doc_type)
            
            if not show_expired:
                query += " AND (expiry_date IS NULL OR expiry_date >= date('now'))"
            
            query += " ORDER BY uploaded_at DESC LIMIT ?"
            params.append(limit)
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            return [dict(row) for row in rows]

    def delete_document(self, document_id: int, company_id: int) -> bool:
        """Soft delete a document"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                UPDATE document_registry 
                SET is_latest_version = 0, verification_status = 'deleted'
                WHERE id = ? AND company_id = ?
            """, (document_id, company_id))
            return cursor.rowcount > 0


    # =========================================================
    # AUTO-FILL METHODS
    # =========================================================

    def get_auto_fill_data(self, company_id: int, data_type: str, 
                        search_term: str = None) -> Dict:
        """Get data for auto-filling forms"""
        result = {}
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                if data_type == 'personnel':
                    cursor.execute("""
                        SELECT id, full_name, designation, employee_id,
                            educational_qualification, skills
                        FROM personnel
                        WHERE company_id = ? AND employment_status = 'active'
                        AND is_key_personnel = 1
                        ORDER BY full_name
                    """, (company_id,))
                    
                    personnel = []
                    for row in cursor.fetchall():
                        row_dict = dict(row)
                        personnel.append({
                            'id': row_dict.get('id'),
                            'name': row_dict.get('full_name'),
                            'designation': row_dict.get('designation'),
                            'employee_id': row_dict.get('employee_id'),
                            'qualification': row_dict.get('educational_qualification'),
                            'skills': json.loads(row_dict.get('skills') or '[]')
                        })
                    result['personnel'] = personnel
                    
                elif data_type == 'equipment':
                    cursor.execute("""
                        SELECT id, equipment_name, equipment_type, model,
                            capacity, current_status, ownership_type
                        FROM equipment
                        WHERE company_id = ? AND current_status = 'available'
                        ORDER BY equipment_name
                    """, (company_id,))
                    
                    equipment = []
                    for row in cursor.fetchall():
                        row_dict = dict(row)
                        equipment.append({
                            'id': row_dict.get('id'),
                            'name': row_dict.get('equipment_name'),
                            'type': row_dict.get('equipment_type'),
                            'model': row_dict.get('model'),
                            'capacity': row_dict.get('capacity'),
                            'status': row_dict.get('current_status'),
                            'ownership': row_dict.get('ownership_type')
                        })
                    result['equipment'] = equipment
                    
                elif data_type == 'experience':
                    cursor.execute("""
                        SELECT id, project_name, client_name, contract_value,
                            completion_date, nature_of_work
                        FROM experience_record
                        WHERE company_id = ? AND is_completed = 1
                        ORDER BY completion_date DESC
                        LIMIT 20
                    """, (company_id,))
                    
                    experiences = []
                    for row in cursor.fetchall():
                        row_dict = dict(row)
                        experiences.append({
                            'id': row_dict.get('id'),
                            'project': row_dict.get('project_name'),
                            'client': row_dict.get('client_name'),
                            'value': row_dict.get('contract_value'),
                            'completion_date': row_dict.get('completion_date'),
                            'nature_of_work': row_dict.get('nature_of_work')
                        })
                    result['experiences'] = experiences
                    
                elif data_type == 'financial':
                    cursor.execute("""
                        SELECT fiscal_year, annual_turnover, net_worth,
                            working_capital, credit_limit, bank_guarantee_limit
                        FROM financial_capacity
                        WHERE company_id = ?
                        ORDER BY fiscal_year DESC
                        LIMIT 3
                    """, (company_id,))
                    
                    financial = []
                    for row in cursor.fetchall():
                        row_dict = dict(row)
                        financial.append({
                            'year': row_dict.get('fiscal_year'),
                            'turnover': row_dict.get('annual_turnover'),
                            'net_worth': row_dict.get('net_worth'),
                            'working_capital': row_dict.get('working_capital'),
                            'credit_limit': row_dict.get('credit_limit'),
                            'bg_limit': row_dict.get('bank_guarantee_limit')
                        })
                    result['financial'] = financial
                    
                return result
                
            except Exception as e:
                logger.error(f"Error getting auto-fill data: {e}")
                return {}

    def search_knowledge_base(self, company_id: int, query: str, 
                            categories: List[str] = None) -> List[Dict]:
        """Unified search across all knowledge base entities"""
        results = []
        query_like = f"%{query}%"
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                if not categories or 'personnel' in categories:
                    cursor.execute("""
                        SELECT 'personnel' as source, id, full_name as name, 
                            designation, personal_phone as phone, personal_email as email
                        FROM personnel
                        WHERE company_id = ? AND employment_status = 'active'
                        AND (full_name LIKE ? OR designation LIKE ? OR employee_id LIKE ?)
                        LIMIT 10
                    """, (company_id, query_like, query_like, query_like))
                    
                    for row in cursor.fetchall():
                        row_dict = dict(row)
                        results.append({
                            'source': 'personnel',
                            'id': row_dict.get('id'),
                            'name': row_dict.get('name'),
                            'designation': row_dict.get('designation'),
                            'phone': row_dict.get('phone'),
                            'email': row_dict.get('email'),
                            'relevance': 1.0
                        })
                
                if not categories or 'equipment' in categories:
                    cursor.execute("""
                        SELECT 'equipment' as source, id, equipment_name as name,
                            equipment_type, model, capacity, current_status
                        FROM equipment
                        WHERE company_id = ? AND current_status != 'scrapped'
                        AND (equipment_name LIKE ? OR model LIKE ? OR serial_number LIKE ?)
                        LIMIT 10
                    """, (company_id, query_like, query_like, query_like))
                    
                    for row in cursor.fetchall():
                        row_dict = dict(row)
                        results.append({
                            'source': 'equipment',
                            'id': row_dict.get('id'),
                            'name': row_dict.get('name'),
                            'type': row_dict.get('equipment_type'),
                            'model': row_dict.get('model'),
                            'capacity': row_dict.get('capacity'),
                            'status': row_dict.get('current_status'),
                            'relevance': 1.0
                        })
                
                if not categories or 'experience' in categories:
                    cursor.execute("""
                        SELECT 'experience' as source, id, project_name as name,
                            client_name, contract_value, completion_date, nature_of_work
                        FROM experience_record
                        WHERE company_id = ? AND is_completed = 1
                        AND (project_name LIKE ? OR client_name LIKE ? OR nature_of_work LIKE ?)
                        ORDER BY completion_date DESC
                        LIMIT 10
                    """, (company_id, query_like, query_like, query_like))
                    
                    for row in cursor.fetchall():
                        row_dict = dict(row)
                        results.append({
                            'source': 'experience',
                            'id': row_dict.get('id'),
                            'name': row_dict.get('name'),
                            'client': row_dict.get('client_name'),
                            'value': row_dict.get('contract_value'),
                            'date': row_dict.get('completion_date'),
                            'nature': row_dict.get('nature_of_work'),
                            'relevance': 1.0
                        })
                
                return results
                
            except Exception as e:
                logger.error(f"Error searching knowledge base: {e}")
                return []


    # =========================================================
    # EXTENSION USAGE TRACKING METHODS
    # =========================================================
    def get_extension_fill_usage(self, company_id: int) -> Dict:
        """Get extension fill usage for a company"""
        try:
            # Get plan limit from subscription_plans via subscriptions
            usage = self.get_extension_usage_metrics(company_id)
            
            limit = usage.get('extension_auto_fills', 5)
            used = usage.get('analyses_used', 0)
            
            # Get actual extension log count
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute("""
                    SELECT COUNT(*) FROM extension_auto_fill_log
                    WHERE company_id = ? 
                    AND filled_at >= date('now', 'start of month')
                """, (company_id,))
                row = cursor.fetchone()
                if row:
                    if hasattr(row, 'keys'):
                        actual_used = row.get('COUNT(*)', 0)
                    else:
                        actual_used = row[0] if row else 0
                else:
                    actual_used = 0
            
            return {
                'used': actual_used,
                'limit': limit,
                'remaining': -1 if limit == -1 else max(0, limit - actual_used),
                'is_unlimited': limit == -1,
                'plan': usage.get('plan', 'free'),
                'can_export_data': usage.get('can_export_data', False),
                'can_edit_rates': usage.get('can_edit_rates', False),
                'can_delete_rates': usage.get('can_delete_rates', False),
                'can_create_versions': usage.get('can_create_versions', False),
                'can_manage_team': usage.get('can_manage_team', False)
            }
            
        except Exception as e:
            print(f"Error getting extension fill usage: {e}")
            return {
                'used': 0,
                'limit': 5,
                'remaining': 5,
                'is_unlimited': False,
                'plan': 'free',
                'can_export_data': False,
                'can_edit_rates': False,
                'can_delete_rates': False,
                'can_create_versions': False,
                'can_manage_team': False
            }    
    def get_extension_usage_metrics(self, company_id: int) -> Dict:
        """Get extension usage metrics for a company including plan limits"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                # Join subscriptions with subscription_plans
                cursor.execute("""
                    SELECT 
                        s.plan,
                        s.analyses_used,
                        s.analyses_limit,
                        s.max_boq_generations,
                        s.boq_used,
                        s.max_bid_optimizations,
                        s.bid_optimizations_used,
                        sp.extension_auto_fills,
                        sp.max_users,
                        sp.can_export_data,
                        sp.can_edit_rates,
                        sp.can_delete_rates,
                        sp.can_create_versions,
                        sp.can_manage_team,
                        sp.max_tender_analyses,
                        s.status
                    FROM subscriptions s
                    LEFT JOIN subscription_plans sp ON s.plan = sp.plan_name
                    WHERE s.company_id = ? AND s.status = 'active'
                    ORDER BY s.id DESC LIMIT 1
                """, (company_id,))
                
                result = cursor.fetchone()
                
                if result:
                    return dict(result)
                
                # Default values if no subscription found
                return {
                    'plan': 'free',
                    'analyses_used': 0,
                    'analyses_limit': 5,
                    'max_boq_generations': 5,
                    'boq_used': 0,
                    'max_bid_optimizations': 5,
                    'bid_optimizations_used': 0,
                    'extension_auto_fills': 5,
                    'max_users': 1,
                    'max_tender_analyses': 5,
                    'can_export_data': False,
                    'can_edit_rates': False,
                    'can_delete_rates': False,
                    'can_create_versions': False,
                    'can_manage_team': False,
                    'status': 'active'
                }
                
        except Exception as e:
            print(f"Error getting extension usage: {e}")
            return {
                'plan': 'free',
                'analyses_used': 0,
                'analyses_limit': 5,
                'max_boq_generations': 5,
                'boq_used': 0,
                'max_bid_optimizations': 5,
                'bid_optimizations_used': 0,
                'extension_auto_fills': 5,
                'max_users': 1,
                'max_tender_analyses': 5,
                'can_export_data': False,
                'can_edit_rates': False,
                'can_delete_rates': False,
                'can_create_versions': False,
                'can_manage_team': False,
                'status': 'active'
            }    
     
    def log_extension_fill(self, company_id: int, user_id: int, field_label: str, 
                        confidence: float, page_url: str) -> bool:
        """Log an auto-fill action from the extension"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                # Create table if it doesn't exist
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS extension_auto_fill_log (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        company_id INTEGER NOT NULL,
                        user_id INTEGER NOT NULL,
                        field_label TEXT,
                        confidence_score REAL,
                        page_url TEXT,
                        filled_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                cursor.execute("""
                    INSERT INTO extension_auto_fill_log (
                        company_id, user_id, field_label, confidence_score, page_url
                    ) VALUES (?, ?, ?, ?, ?)
                """, (company_id, user_id, field_label, confidence, page_url))
                return True
            except Exception as e:
                logger.error(f"Error logging extension fill: {e}")
                return False

    def can_use_extension_fill(self, company_id: int) -> Tuple[bool, str, int]:
        """Check if company can use extension auto-fill"""
        usage = self.get_extension_fill_usage(company_id)
        
        if usage['is_unlimited']:
            return True, "Unlimited fills available", -1
        
        if usage['remaining'] > 0:
            return True, f"{usage['remaining']} fills remaining this month", usage['remaining']
        else:
            return False, f"You've used all {usage['limit']} auto-fills this month. Upgrade to continue.", 0

    # =========================================================
    # BOQ MANAGEMENT METHODS
    # =========================================================

    def create_boq(self, tender_id: int, company_id: int, 
                rate_source: str, zone: str, notes: str = None) -> Tuple[Optional[int], str]:
        """Create a new BOQ"""
        import streamlit as st
        
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                # Get tender details
                cursor.execute("""
                    SELECT tender_id, tender_title, procuring_entity, official_estimate
                    FROM company_tenders
                    WHERE id = ? AND company_id = ?
                """, (tender_id, company_id))
                tender = cursor.fetchone()
                
                if not tender:
                    return None, "Tender not found"
                
                tender_dict = dict(tender) if tender else {}
                
                # Get rate version
                cursor.execute("""
                    SELECT edition_year FROM rate_versions 
                    WHERE source = ? AND is_active = 1
                    LIMIT 1
                """, (rate_source,))
                version = cursor.fetchone()
                edition_year = version[0] if version else 2025
                
                cursor.execute("""
                    INSERT INTO boq_generation_history (
                        user_id, company_id, tender_id, tender_title, procuring_entity,
                        selected_zone, rate_source, edition_year, status, notes, generated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'draft', ?, CURRENT_TIMESTAMP)
                """, (
                    st.session_state.get('user_id', 0), company_id, tender_dict.get('tender_id'), 
                    tender_dict.get('tender_title'), tender_dict.get('procuring_entity'),
                    zone, rate_source, edition_year, notes
                ))
                
                boq_id = cursor.lastrowid
                self._log_boq_activity(boq_id, 'create', "BOQ created")
                
                return boq_id, "BOQ created successfully"
                
        except Exception as e:
            logger.error(f"Error creating BOQ: {e}")
            return None, str(e)

    def add_boq_item(self, boq_id: int, item_code: str, description: str, 
                    unit: str, quantity: float, unit_rate: float, 
                    is_custom: bool = False, notes: str = None) -> Tuple[bool, str]:
        """Add item to BOQ"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                total = quantity * unit_rate
                
                cursor.execute("""
                    INSERT INTO boq_items (
                        boq_id, item_code, description, unit, quantity, unit_rate, total, is_custom, notes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (boq_id, item_code, description, unit, quantity, unit_rate, total, is_custom, notes))
                
                # Update BOQ totals
                cursor.execute("""
                    UPDATE boq_generation_history 
                    SET item_count = (SELECT COUNT(*) FROM boq_items WHERE boq_id = ?),
                        total_estimated_cost = (SELECT SUM(total) FROM boq_items WHERE boq_id = ?)
                    WHERE id = ?
                """, (boq_id, boq_id, boq_id))
                
                return True, "Item added successfully"
                
        except Exception as e:
            logger.error(f"Error adding BOQ item: {e}")
            return False, str(e)

    def update_boq_item(self, item_id: int, quantity: float, unit_rate: float = None) -> Tuple[bool, str]:
        """Update BOQ item quantity or rate"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                if unit_rate:
                    cursor.execute("""
                        UPDATE boq_items 
                        SET quantity = ?, unit_rate = ?, total = quantity * ?
                        WHERE id = ?
                    """, (quantity, unit_rate, unit_rate, item_id))
                else:
                    cursor.execute("""
                        UPDATE boq_items 
                        SET quantity = ?, total = quantity * unit_rate
                        WHERE id = ?
                    """, (quantity, item_id))
                
                # Get boq_id to update totals
                cursor.execute("SELECT boq_id FROM boq_items WHERE id = ?", (item_id,))
                result = cursor.fetchone()
                if result:
                    boq_id = result[0]
                    cursor.execute("""
                        UPDATE boq_generation_history 
                        SET total_estimated_cost = (SELECT SUM(total) FROM boq_items WHERE boq_id = ?)
                        WHERE id = ?
                    """, (boq_id, boq_id))
                
                return True, "Item updated successfully"
                
        except Exception as e:
            logger.error(f"Error updating BOQ item: {e}")
            return False, str(e)

    def delete_boq_item(self, item_id: int) -> Tuple[bool, str]:
        """Delete item from BOQ"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                cursor.execute("SELECT boq_id FROM boq_items WHERE id = ?", (item_id,))
                result = cursor.fetchone()
                if result:
                    boq_id = result[0]
                    cursor.execute("DELETE FROM boq_items WHERE id = ?", (item_id,))
                    
                    cursor.execute("""
                        UPDATE boq_generation_history 
                        SET item_count = (SELECT COUNT(*) FROM boq_items WHERE boq_id = ?),
                            total_estimated_cost = (SELECT SUM(total) FROM boq_items WHERE boq_id = ?)
                        WHERE id = ?
                    """, (boq_id, boq_id, boq_id))
                
                return True, "Item deleted successfully"
                
        except Exception as e:
            logger.error(f"Error deleting BOQ item: {e}")
            return False, str(e)

    def get_boq(self, boq_id: int) -> Optional[Dict]:
        """Get complete BOQ with items"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                # Get BOQ header
                cursor.execute("""
                    SELECT b.*, c.company_name, u.username as created_by_name
                    FROM boq_generation_history b
                    LEFT JOIN companies c ON b.company_id = c.id
                    LEFT JOIN users u ON b.user_id = u.id
                    WHERE b.id = ?
                """, (boq_id,))
                
                row = cursor.fetchone()
                if not row:
                    return None
                
                boq = dict(row)
                
                # Get BOQ items
                import pandas as pd
                items = pd.read_sql_query("""
                    SELECT * FROM boq_items
                    WHERE boq_id = ?
                    ORDER BY id
                """, conn, params=[boq_id])
                
                # Get approval history
                history = pd.read_sql_query("""
                    SELECT * FROM boq_approval_history
                    WHERE boq_id = ?
                    ORDER BY created_at DESC
                """, conn, params=[boq_id])
                
                return {
                    'boq': boq,
                    'items': items,
                    'history': history
                }
                
        except Exception as e:
            logger.error(f"Error getting BOQ: {e}")
            return None

    def get_boq_list(self, company_id: int = None, status: str = None, limit: int = 100) -> pd.DataFrame:
        """Get list of BOQs with filters"""
        import streamlit as st
        import pandas as pd
        
        try:
            with self.get_connection() as conn:
                user_role = st.session_state.get('user_role', 'viewer')
                
                if user_role in ['admin', 'system_admin']:
                    query = """
                        SELECT b.*, c.company_name, u.username as created_by_name
                        FROM boq_generation_history b
                        LEFT JOIN companies c ON b.company_id = c.id
                        LEFT JOIN users u ON b.user_id = u.id
                        WHERE 1=1
                    """
                    params = []
                else:
                    if not company_id:
                        company_id = st.session_state.get('company_id')
                    query = """
                        SELECT b.*, u.username as created_by_name
                        FROM boq_generation_history b
                        LEFT JOIN users u ON b.user_id = u.id
                        WHERE b.company_id = ?
                    """
                    params = [company_id]
                
                if status:
                    query += " AND b.status = ?"
                    params.append(status)
                
                query += " ORDER BY b.generated_at DESC LIMIT ?"
                params.append(limit)
                
                return pd.read_sql_query(query, conn, params=params)
                
        except Exception as e:
            logger.error(f"Error getting BOQ list: {e}")
            return pd.DataFrame()


    # =========================================================
    # BOQ WORKFLOW METHODS
    # =========================================================

    def submit_boq(self, boq_id: int, comment: str = None) -> Tuple[bool, str]:
        """Submit BOQ for approval"""
        return self._update_boq_status(boq_id, 'submitted', comment, 'submitted_for_review')

    def approve_boq(self, boq_id: int, comment: str = None) -> Tuple[bool, str]:
        """Approve BOQ"""
        return self._update_boq_status(boq_id, 'approved', comment, 'approved')

    def reject_boq(self, boq_id: int, comment: str = None) -> Tuple[bool, str]:
        """Reject BOQ with comments"""
        return self._update_boq_status(boq_id, 'rejected', comment, 'rejected')

    def _update_boq_status(self, boq_id: int, status: str, comment: str, action: str) -> Tuple[bool, str]:
        """Internal method to update BOQ status"""
        import streamlit as st
        from datetime import datetime
        
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                cursor.execute("""
                    UPDATE boq_generation_history 
                    SET status = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (status, boq_id))
                
                # Log approval history
                cursor.execute("""
                    INSERT INTO boq_approval_history (boq_id, action, comment, user_id, username, user_role, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (boq_id, action, comment, 
                    st.session_state.get('user_id', 0),
                    st.session_state.get('username', 'system'),
                    st.session_state.get('user_role', 'viewer')))
                
                self._log_boq_activity(boq_id, action, comment)
                
                return True, f"BOQ {status} successfully"
                
        except Exception as e:
            logger.error(f"Error updating BOQ status: {e}")
            return False, str(e)

    def delete_boq(self, boq_id: int) -> Tuple[bool, str]:
        """Delete BOQ"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                cursor.execute("DELETE FROM boq_items WHERE boq_id = ?", (boq_id,))
                cursor.execute("DELETE FROM boq_approval_history WHERE boq_id = ?", (boq_id,))
                cursor.execute("DELETE FROM boq_activity_log WHERE boq_id = ?", (boq_id,))
                cursor.execute("DELETE FROM boq_generation_history WHERE id = ?", (boq_id,))
                
                return True, "BOQ deleted successfully"
                
        except Exception as e:
            logger.error(f"Error deleting BOQ: {e}")
            return False, str(e)

    def copy_boq(self, boq_id: int) -> Tuple[Optional[int], str]:
        """Create a copy of existing BOQ"""
        try:
            original = self.get_boq(boq_id)
            if not original:
                return None, "Original BOQ not found"
            
            # Create new BOQ
            new_boq_id, msg = self.create_boq(
                tender_id=original['boq'].get('tender_id'),
                company_id=original['boq'].get('company_id'),
                rate_source=original['boq'].get('rate_source'),
                zone=original['boq'].get('selected_zone'),
                notes=f"Copy of BOQ #{boq_id}"
            )
            
            if not new_boq_id:
                return None, msg
            
            # Copy items
            for _, item in original['items'].iterrows():
                self.add_boq_item(
                    boq_id=new_boq_id,
                    item_code=item.get('item_code'),
                    description=item.get('description'),
                    unit=item.get('unit'),
                    quantity=item.get('quantity', 0),
                    unit_rate=item.get('unit_rate', 0),
                    is_custom=item.get('is_custom', False),
                    notes=item.get('notes')
                )
            
            return new_boq_id, "BOQ copied successfully"
            
        except Exception as e:
            logger.error(f"Error copying BOQ: {e}")
            return None, str(e)

    def lock_boq(self, boq_id: int) -> Tuple[bool, str]:
        """Lock BOQ after bid submission (prevents further edits)"""
        import streamlit as st
        
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                cursor.execute("""
                    UPDATE boq_generation_history 
                    SET is_locked = 1, locked_at = CURRENT_TIMESTAMP, locked_by = ?
                    WHERE id = ?
                """, (st.session_state.get('user_id', 0), boq_id))
                
                self._log_boq_activity(boq_id, 'lock', "BOQ locked after bid submission")
                return True, "BOQ locked successfully"
                
        except Exception as e:
            logger.error(f"Error locking BOQ: {e}")
            return False, str(e)

    def unlock_boq(self, boq_id: int) -> Tuple[bool, str]:
        """Unlock BOQ (admin only)"""
        import streamlit as st
        
        try:
            user_role = st.session_state.get('user_role', 'viewer')
            if user_role not in ['admin', 'system_admin']:
                return False, "Only administrators can unlock BOQs"
            
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                cursor.execute("""
                    UPDATE boq_generation_history 
                    SET is_locked = 0, locked_at = NULL, locked_by = NULL
                    WHERE id = ?
                """, (boq_id,))
                
                self._log_boq_activity(boq_id, 'unlock', "BOQ unlocked by admin")
                return True, "BOQ unlocked successfully"
                
        except Exception as e:
            logger.error(f"Error unlocking BOQ: {e}")
            return False, str(e)


    # =========================================================
    # BOQ HELPER METHODS
    # =========================================================

    def _log_boq_activity(self, boq_id: int, action: str, details: str):
        """Log BOQ activity"""
        import streamlit as st
        
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                # Create table if not exists
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS boq_activity_log (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        boq_id INTEGER,
                        action TEXT,
                        details TEXT,
                        user_id INTEGER,
                        username TEXT,
                        user_role TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                cursor.execute("""
                    INSERT INTO boq_activity_log (boq_id, action, details, user_id, username, user_role, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (boq_id, action, details,
                    st.session_state.get('user_id', 0),
                    st.session_state.get('username', 'system'),
                    st.session_state.get('user_role', 'viewer')))
                
        except Exception as e:
            logger.error(f"Error logging BOQ activity: {e}")


    # =========================================================
    # BOQ EXPORT METHODS
    # =========================================================

    def export_boq_to_excel(self, boq_id: int) -> bytes:
        """Export BOQ to Excel format"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        boq_data = self.get_boq(boq_id)
        if not boq_data:
            return None
        
        wb = Workbook()
        
        # Main BOQ sheet
        ws = wb.active
        ws.title = "BOQ"
        
        # Header styling
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        
        # Company header
        ws.merge_cells('A1:F1')
        ws['A1'] = f"BOQ Report - {boq_data['boq'].get('tender_title', 'N/A')}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Tender details
        ws['A3'] = "Tender ID:"
        ws['B3'] = boq_data['boq'].get('tender_id', 'N/A')
        ws['A4'] = "Rate Source:"
        ws['B4'] = boq_data['boq'].get('rate_source', 'N/A')
        ws['A5'] = "Zone:"
        ws['B5'] = boq_data['boq'].get('selected_zone', 'N/A')
        ws['A6'] = "Status:"
        ws['B6'] = boq_data['boq'].get('status', 'N/A').upper()
        
        # Items header
        headers = ['Sl No', 'Item Code', 'Description', 'Unit', 'Quantity', 'Unit Rate (BDT)', 'Total (BDT)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Items data
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        total = 0
        items_df = boq_data['items']
        if not items_df.empty:
            for idx, (_, item) in enumerate(items_df.iterrows(), 1):
                row = 8 + idx
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=item.get('item_code', ''))
                ws.cell(row=row, column=3, value=(item.get('description', '') or '')[:100])
                ws.cell(row=row, column=4, value=item.get('unit', ''))
                ws.cell(row=row, column=5, value=float(item.get('quantity', 0)))
                ws.cell(row=row, column=6, value=round(float(item.get('unit_rate', 0)), 2))
                ws.cell(row=row, column=7, value=round(float(item.get('total', 0)), 2))
                total += float(item.get('total', 0))
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
        
        # Total row
        total_row = 8 + len(items_df) + 1
        ws.merge_cells(f'A{total_row}:F{total_row}')
        ws.cell(row=total_row, column=1, value="GRAND TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=round(total, 2))
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()    
    # =========================================================
    # BATCH 1: CHAPTER MANAGEMENT METHODS
    # =========================================================

    def add_lged_chapter(self, chapter_number: str, chapter_name: str, description: str = ""):
        """Add a new LGED chapter"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                INSERT OR REPLACE INTO lged_chapters (chapter_number, chapter_name, description)
                VALUES (?, ?, ?)
            """, (chapter_number, chapter_name, description))

    def add_pwd_chapter(self, chapter_number: str, chapter_name: str, description: str = ""):
        """Add a new PWD chapter"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                INSERT OR REPLACE INTO pwd_chapters (chapter_number, chapter_name, description)
                VALUES (?, ?, ?)
            """, (chapter_number, chapter_name, description))

    def get_lged_sections_by_chapter(self, chapter_number: str):
        """Get LGED sections for a specific chapter"""
        import pandas as pd
        with self.get_connection() as conn:
            # First get the chapter ID
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT id FROM rate_chapters 
                WHERE source = 'LGED' AND chapter_number = ?
                ORDER BY id DESC LIMIT 1
            """, (chapter_number,))
            
            chapter_row = cursor.fetchone()
            
            if chapter_row:
                chapter_id = chapter_row['id']
                df = pd.read_sql_query("""
                    SELECT section_number, section_name, description
                    FROM rate_sections 
                    WHERE source = 'LGED' AND chapter_id = ?
                    ORDER BY section_number
                """, conn, params=[chapter_id])
                return df
            return pd.DataFrame()

    def get_pwd_stats(self) -> Dict:
        """Get PWD statistics"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("SELECT COUNT(*) FROM pwd_parents")
            total_parents = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM pwd_children")
            total_children = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM pwd_rates")
            total_rates = cursor.fetchone()[0]
            
            return {
                'total_parents': total_parents,
                'total_children': total_children,
                'total_rates': total_rates,
                'total_items': total_parents + total_children
            }    
    # =========================================================
    # BATCH 2: VERSION MANAGEMENT METHODS
    # =========================================================

    def get_version_history(self, source: str, edition_year: int):
        """Get version history for a specific source and edition year"""
        import pandas as pd
        with self.get_connection() as conn:
            query = """
                SELECT 
                    id, version_name, version_number, is_active,
                    COALESCE(created_at, release_date, datetime('now')) as created_at,
                    COALESCE(updated_at, release_date, datetime('now')) as updated_at,
                    COALESCE(total_parents, 0) as total_parents,
                    COALESCE(total_children, 0) as total_children,
                    COALESCE(total_rates, 0) as total_rates,
                    COALESCE(notes, '') as notes
                FROM rate_versions
                WHERE source = ? AND edition_year = ?
                ORDER BY version_number DESC
            """
            df = pd.read_sql_query(query, conn, params=(source, edition_year))
            
            if not df.empty:
                df['total_items'] = df['total_parents'] + df['total_children']
            
            return df

    def clear_lged_version_data(self, version_id: int) -> bool:
        """Clear all LGED data for a specific version"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute("DELETE FROM lged_zone_rates WHERE version_id = ?", (version_id,))
                cursor.execute("DELETE FROM lged_children WHERE version_id = ?", (version_id,))
                cursor.execute("DELETE FROM lged_parents WHERE version_id = ?", (version_id,))
            print(f"✅ Cleared LGED data for version ID: {version_id}")
            return True
        except Exception as e:
            print(f"Error clearing LGED version data: {e}")
            return False

    def clear_pwd_version_data(self, version_id: int) -> bool:
        """Clear all PWD data for a specific version"""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute("DELETE FROM pwd_rates WHERE version_id = ?", (version_id,))
                cursor.execute("DELETE FROM pwd_children WHERE version_id = ?", (version_id,))
                cursor.execute("DELETE FROM pwd_parents WHERE version_id = ?", (version_id,))
            print(f"✅ Cleared PWD data for version ID: {version_id}")
            return True
        except Exception as e:
            print(f"Error clearing PWD version data: {e}")
            return False

    
    # =========================================================
    # BATCH 3: TABLE INITIALIZATION METHODS
    # =========================================================

    def init_chapters_tables(self):
        """Initialize chapters tables for both PWD and LGED"""
        import sqlite3
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # PWD Chapters table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS pwd_chapters (
                    chapter_number TEXT PRIMARY KEY,
                    chapter_name TEXT NOT NULL,
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # LGED Chapters table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS lged_chapters (
                    chapter_number TEXT PRIMARY KEY,
                    chapter_name TEXT NOT NULL,
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Insert default PWD chapters
            default_pwd_chapters = [
                ("01", "General, Site Facilities and Safety"),
                ("02", "Excavation, Filling & Site Development"),
                ("03", "Brick Works, Patent Stone and Fancy Screen"),
                ("04", "Reinforced Cement Concrete (RCC) Works"),
                ("05", "Mosaic Works"),
                ("06", "Tiles, Marble and Granite Stone Works"),
                ("07", "Wood Works"),
                ("08", "Window Grill, Verandah Grill & Netting"),
                ("09", "Pile Works and Pile Test"),
                ("10", "Structural Steel Works"),
                ("11", "Wood Works in Door and Window Frame"),
                ("12", "Collapsible Gate, M.S. Gate, Rolling Shutter"),
                ("13", "Aluminium Door, Window and Partition"),
                ("14", "Cement Plaster, Fair-Face Plaster"),
                ("15", "Painting and Polishing"),
                ("16", "Water Proofing and Heat Proofing"),
                ("17", "False Ceiling and Wall Panelling"),
                ("18", "Sanitary and Plumbing Works"),
                ("19", "Gas Pipe Line Installation"),
                ("20", "Sub-Soil Investigation"),
                ("21", "Repair Works")
            ]
            
            for chapter_num, chapter_name in default_pwd_chapters:
                cursor.execute("INSERT OR IGNORE INTO pwd_chapters (chapter_number, chapter_name) VALUES (?, ?)", 
                            (chapter_num, chapter_name))
            
            # Insert default LGED chapters
            default_lged_chapters = [
                ("1", "General, Site Facilities and Safety"),
                ("2", "Earth Works in Road Embankment"),
                ("3", "Sub-Base and Base Course"),
                ("4", "Bituminous Pavement"),
                ("5", "Bridge and Culvert Works"),
                ("6", "Building Works"),
                ("7", "Sanitary and Water Supply"),
                ("8", "Electrical Works"),
                ("9", "Landscaping and Development")
            ]
            
            for chapter_num, chapter_name in default_lged_chapters:
                cursor.execute("INSERT OR IGNORE INTO lged_chapters (chapter_number, chapter_name) VALUES (?, ?)", 
                            (chapter_num, chapter_name))
            
            print("✅ Chapters tables initialized")

    def init_lged_tables(self):
        """Initialize LGED-specific tables"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # LGED Versions table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS rate_versions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source TEXT DEFAULT 'LGED',
                    version_name TEXT NOT NULL,
                    edition_year INTEGER NOT NULL,
                    effective_from DATE,
                    is_active BOOLEAN DEFAULT 0,
                    released_by TEXT,
                    release_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    notes TEXT,
                    total_parents INTEGER DEFAULT 0,
                    total_children INTEGER DEFAULT 0,
                    total_rates INTEGER DEFAULT 0,
                    created_by TEXT
                )
            """)
            
            # LGED Parents table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS lged_parents (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    code TEXT NOT NULL,
                    description TEXT,
                    chapter_number TEXT,
                    section_number TEXT,
                    parent_type TEXT DEFAULT 'section_header',
                    has_children BOOLEAN DEFAULT 0,
                    unit TEXT,
                    version_id INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (version_id) REFERENCES rate_versions(id)
                )
            """)
            
            # LGED Children table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS lged_children (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    code TEXT NOT NULL,
                    parent_code TEXT NOT NULL,
                    description TEXT,
                    unit TEXT,
                    chapter_number TEXT,
                    section_number TEXT,
                    zone_a REAL,
                    zone_b REAL,
                    zone_c REAL,
                    zone_d REAL,
                    edition_year INTEGER,
                    version_id INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (version_id) REFERENCES rate_versions(id)
                )
            """)
            
            # LGED Zone Rates table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS lged_zone_rates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    parent_id INTEGER,
                    zone_name TEXT NOT NULL,
                    unit_rate REAL NOT NULL,
                    version_id INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (parent_id) REFERENCES lged_parents(id),
                    FOREIGN KEY (version_id) REFERENCES rate_versions(id)
                )
            """)
            
            # LGED Sections table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS lged_sections (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    chapter_number TEXT NOT NULL,
                    section_number TEXT NOT NULL,
                    section_name TEXT NOT NULL,
                    description TEXT,
                    display_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(chapter_number, section_number)
                )
            """)
            
            # Zone mapping table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS lged_zone_mapping (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    zone_code TEXT PRIMARY KEY,
                    zone_name TEXT,
                    divisions TEXT,
                    accessibility_bonus REAL DEFAULT 0.05,
                    description TEXT
                )
            """)
            
            # Insert default zone mapping
            cursor.execute("""
                INSERT OR IGNORE INTO lged_zone_mapping (zone_code, zone_name, divisions, accessibility_bonus)
                VALUES 
                ('A', 'Dhaka & Mymensingh Division', '["Dhaka","Mymensingh"]', 0.05),
                ('B', 'Chattogram & Sylhet Division', '["Chattogram","Sylhet"]', 0.05),
                ('C', 'Rajshahi & Rangpur Division', '["Rajshahi","Rangpur"]', 0.05),
                ('D', 'Khulna & Barishal Division', '["Khulna","Barishal"]', 0.05)
            """)
            
            print("✅ LGED tables initialized")

    def init_pwd_hierarchical_tables(self):
        """Initialize hierarchical PWD tables"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS rate_versions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source TEXT DEFAULT 'PWD',
                    version_name TEXT NOT NULL,
                    edition_year INTEGER NOT NULL,
                    effective_from DATE,
                    is_active BOOLEAN DEFAULT 0,
                    released_by TEXT,
                    release_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    notes TEXT,
                    total_parents INTEGER DEFAULT 0,
                    total_children INTEGER DEFAULT 0,
                    total_rates INTEGER DEFAULT 0,
                    created_by TEXT
                )
            """)
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS pwd_parents (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    pwd_code TEXT PRIMARY KEY,
                    description TEXT,
                    chapter_number TEXT,
                    version_id INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS pwd_children (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    pwd_code TEXT PRIMARY KEY,
                    parent_code TEXT NOT NULL,
                    description TEXT,
                    unit TEXT,
                    edition_year INTEGER,
                    version_id INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (parent_code) REFERENCES pwd_parents(pwd_code) ON DELETE CASCADE
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS pwd_rates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    pwd_code TEXT NOT NULL,
                    zone_name TEXT NOT NULL,
                    unit_rate REAL NOT NULL,
                    edition_year INTEGER NOT NULL,
                    version_id INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(pwd_code, zone_name, edition_year),
                    FOREIGN KEY (pwd_code) REFERENCES pwd_children(pwd_code) ON DELETE CASCADE
                )
            ''')
            
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_pwd_parent_chapter ON pwd_parents(chapter_number)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_pwd_child_parent ON pwd_children(parent_code)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_pwd_rates_code ON pwd_rates(pwd_code)')
            
            print("✅ PWD hierarchical tables initialized")

    def search_pwd_items(self, search_term: str, limit: int = 50):
        """Search PWD items by code or description"""
        import pandas as pd
        with self.get_connection() as conn:
            return pd.read_sql_query("""
                SELECT pwd_code, description, unit 
                FROM pwd_children 
                WHERE pwd_code LIKE ? OR description LIKE ?
                LIMIT ?
            """, conn, params=(f"%{search_term}%", f"%{search_term}%", limit))    

    # =========================================================
    # COMPETITOR METHODS - KEEP THESE (CLEAN VERSIONS)
    # =========================================================

    def get_competitor_master_list(self, company_id: int, active_only: bool = True) -> List[Dict]:
        """Get competitor master list for a company"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            query = "SELECT * FROM competitor_master WHERE company_id = ?"
            params = [company_id]
            
            if active_only:
                query += " AND is_active = 1"
            
            query += " ORDER BY competitor_name"
            
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]

    def get_competitor_by_id(self, competitor_id: int) -> Optional[Dict]:
        """Get competitor details by ID"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('SELECT * FROM competitor_master WHERE id = ?', (competitor_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def add_competitor_to_master(self, company_id: int, competitor_data: Dict) -> int:
        """Add competitor to master list"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Check if exists
            cursor.execute(
                "SELECT id FROM competitor_master WHERE company_id = ? AND competitor_name = ?",
                (company_id, competitor_data['competitor_name'])
            )
            existing = cursor.fetchone()
            
            if existing:
                # Update existing
                cursor.execute("""
                    UPDATE competitor_master
                    SET business_type = ?, contact_person = ?, phone = ?, email = ?,
                        address = ?, notes = ?, preferred_strategy = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (
                    competitor_data.get('business_type', ''),
                    competitor_data.get('contact_person', ''),
                    competitor_data.get('phone', ''),
                    competitor_data.get('email', ''),
                    competitor_data.get('address', ''),
                    competitor_data.get('notes', ''),
                    competitor_data.get('preferred_strategy', 'Unknown'),
                    existing['id']
                ))
                return existing['id']
            else:
                # Insert new
                cursor.execute("""
                    INSERT INTO competitor_master (
                        company_id, competitor_name, business_type, contact_person,
                        phone, email, address, notes, preferred_strategy,
                        first_seen, last_seen, is_active
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    company_id,
                    competitor_data['competitor_name'],
                    competitor_data.get('business_type', ''),
                    competitor_data.get('contact_person', ''),
                    competitor_data.get('phone', ''),
                    competitor_data.get('email', ''),
                    competitor_data.get('address', ''),
                    competitor_data.get('notes', ''),
                    competitor_data.get('preferred_strategy', 'Unknown'),
                    datetime.now().date(),
                    datetime.now().date(),
                    1
                ))
                return cursor.lastrowid

    def update_competitor_master(self, competitor_id: int, update_data: Dict) -> bool:
        """Update competitor information"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            allowed_fields = ['business_type', 'contact_person', 'phone', 'email', 
                            'address', 'notes', 'preferred_strategy', 'is_active']
            
            updates = []
            values = []
            
            for key, value in update_data.items():
                if key in allowed_fields:
                    updates.append(f"{key} = ?")
                    values.append(value)
            
            if not updates:
                return False
            
            values.append(competitor_id)
            cursor.execute(f"""
                UPDATE competitor_master 
                SET {', '.join(updates)}, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, values)
            
            return cursor.rowcount > 0

    def delete_competitor(self, competitor_id: int) -> bool:
        """Soft delete competitor (mark inactive)"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute('UPDATE competitor_master SET is_active = 0, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (competitor_id,))
            return True

    def update_competitor_stats_from_bid(self, company_id: int, competitor_name: str, 
                                        bid_ratio: float, was_winner: bool) -> bool:
        """Update competitor statistics from a bid"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Ensure bid_ratio is a valid number
            if bid_ratio is None or bid_ratio <= 0:
                bid_ratio = 0.95  # Default fallback
            
            cursor.execute("""
                SELECT id, total_bids, total_wins, avg_bid_ratio 
                FROM competitor_master 
                WHERE company_id = ? AND competitor_name = ?
            """, (company_id, competitor_name))
            
            existing = cursor.fetchone()
            
            if existing:
                comp_id = existing['id']
                total_bids = existing['total_bids'] or 0
                total_wins = existing['total_wins'] or 0
                avg_ratio = existing['avg_bid_ratio'] or 0.90  # Default if None
                
                new_total_bids = total_bids + 1
                new_total_wins = total_wins + (1 if was_winner else 0)
                
                # Calculate new average safely
                if new_total_bids > 0:
                    # Safely calculate new average
                    if avg_ratio is None:
                        avg_ratio = 0.90
                    if bid_ratio is None:
                        bid_ratio = 0.95
                        
                    new_avg_ratio = ((avg_ratio * total_bids) + bid_ratio) / new_total_bids
                else:
                    new_avg_ratio = bid_ratio if bid_ratio else 0.95
                
                cursor.execute("""
                    UPDATE competitor_master 
                    SET total_bids = ?, total_wins = ?, avg_bid_ratio = ?,
                        last_seen = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (new_total_bids, new_total_wins, new_avg_ratio, datetime.now().date(), comp_id))
                
                # Also update competitor_profiles
                self._update_competitor_profile(company_id, competitor_name, bid_ratio, was_winner)
                
                return True
            else:
                # Create new competitor
                cursor.execute("""
                    INSERT INTO competitor_master (
                        company_id, competitor_name, total_bids, total_wins, 
                        avg_bid_ratio, first_seen, last_seen
                    ) VALUES (?, ?, 1, ?, ?, ?, ?)
                """, (company_id, competitor_name, 1 if was_winner else 0, 
                    bid_ratio if bid_ratio else 0.95, 
                    datetime.now().date(), datetime.now().date()))
                
                self._update_competitor_profile(company_id, competitor_name, bid_ratio, was_winner)
                return True

    def _update_competitor_profile(self, company_id: int, competitor_name: str, 
                                bid_ratio: float, was_winner: bool) -> bool:
        """Update competitor profile for analytics"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Ensure bid_ratio is a valid number
            if bid_ratio is None or bid_ratio <= 0:
                bid_ratio = 0.95
            
            cursor.execute("""
                SELECT id, total_appearances, wins_count, avg_bid_ratio, bid_std_dev
                FROM competitor_profiles 
                WHERE company_id = ? AND competitor_name = ?
            """, (company_id, competitor_name))
            
            existing = cursor.fetchone()
            
            if existing:
                comp_id, total, wins, avg_ratio, std_dev = existing
                
                # Handle None values
                total = total or 0
                wins = wins or 0
                avg_ratio = avg_ratio or 0.90
                
                new_total = total + 1
                new_wins = wins + (1 if was_winner else 0)
                new_avg_ratio = ((avg_ratio * total) + bid_ratio) / new_total if new_total > 0 else bid_ratio
                
                if new_avg_ratio < 0.88:
                    strategy = "Aggressive"
                elif new_avg_ratio < 0.93:
                    strategy = "Moderate"
                else:
                    strategy = "Conservative"
                
                cursor.execute("""
                    UPDATE competitor_profiles 
                    SET total_appearances = ?, wins_count = ?, avg_bid_ratio = ?,
                        strategy = ?, last_seen = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (new_total, new_wins, new_avg_ratio, strategy, datetime.now().date(), comp_id))
            else:
                strategy = "Aggressive" if bid_ratio < 0.88 else "Moderate" if bid_ratio < 0.93 else "Conservative"
                cursor.execute("""
                    INSERT INTO competitor_profiles (
                        company_id, competitor_name, first_seen, last_seen,
                        total_appearances, wins_count, avg_bid_ratio, strategy
                    ) VALUES (?, ?, ?, ?, 1, ?, ?, ?)
                """, (company_id, competitor_name, datetime.now().date(), datetime.now().date(),
                    1 if was_winner else 0, bid_ratio, strategy))
            
            return True
    def update_competitor_stats_from_bid_bak2(self, company_id: int, competitor_name: str, 
                                        bid_ratio: float, was_winner: bool) -> bool:
        """Update competitor statistics from a bid"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                SELECT id, total_bids, total_wins, avg_bid_ratio 
                FROM competitor_master 
                WHERE company_id = ? AND competitor_name = ?
            """, (company_id, competitor_name))
            
            existing = cursor.fetchone()
            
            if existing:
                comp_id = existing['id']
                total_bids = existing['total_bids']
                total_wins = existing['total_wins']
                avg_ratio = existing['avg_bid_ratio']
                
                new_total_bids = total_bids + 1
                new_total_wins = total_wins + (1 if was_winner else 0)
                new_avg_ratio = ((avg_ratio * total_bids) + bid_ratio) / new_total_bids if new_total_bids > 0 else bid_ratio
                
                cursor.execute("""
                    UPDATE competitor_master 
                    SET total_bids = ?, total_wins = ?, avg_bid_ratio = ?,
                        last_seen = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (new_total_bids, new_total_wins, new_avg_ratio, datetime.now().date(), comp_id))
                
                # Also update competitor_profiles
                self._update_competitor_profile(company_id, competitor_name, bid_ratio, was_winner)
                
                return True
            else:
                # Create new competitor
                cursor.execute("""
                    INSERT INTO competitor_master (
                        company_id, competitor_name, total_bids, total_wins, 
                        avg_bid_ratio, first_seen, last_seen
                    ) VALUES (?, ?, 1, ?, ?, ?, ?)
                """, (company_id, competitor_name, 1 if was_winner else 0, bid_ratio, 
                    datetime.now().date(), datetime.now().date()))
                
                self._update_competitor_profile(company_id, competitor_name, bid_ratio, was_winner)
                return True

    def _update_competitor_profile_bak(self, company_id: int, competitor_name: str, 
                                bid_ratio: float, was_winner: bool) -> bool:
        """Update competitor profile for analytics"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                SELECT id, total_appearances, wins_count, avg_bid_ratio, bid_std_dev
                FROM competitor_profiles 
                WHERE company_id = ? AND competitor_name = ?
            """, (company_id, competitor_name))
            
            existing = cursor.fetchone()
            
            if existing:
                comp_id, total, wins, avg_ratio, std_dev = existing
                new_total = total + 1
                new_wins = wins + (1 if was_winner else 0)
                new_avg_ratio = ((avg_ratio * total) + bid_ratio) / new_total if new_total > 0 else bid_ratio
                
                if new_avg_ratio < 0.88:
                    strategy = "Aggressive"
                elif new_avg_ratio < 0.93:
                    strategy = "Moderate"
                else:
                    strategy = "Conservative"
                
                cursor.execute("""
                    UPDATE competitor_profiles 
                    SET total_appearances = ?, wins_count = ?, avg_bid_ratio = ?,
                        strategy = ?, last_seen = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (new_total, new_wins, new_avg_ratio, strategy, datetime.now().date(), comp_id))
            else:
                strategy = "Aggressive" if bid_ratio < 0.88 else "Moderate" if bid_ratio < 0.93 else "Conservative"
                cursor.execute("""
                    INSERT INTO competitor_profiles (
                        company_id, competitor_name, first_seen, last_seen,
                        total_appearances, wins_count, avg_bid_ratio, strategy
                    ) VALUES (?, ?, ?, ?, 1, ?, ?, ?)
                """, (company_id, competitor_name, datetime.now().date(), datetime.now().date(),
                    1 if was_winner else 0, bid_ratio, strategy))
            
            return True

    def get_competitor_profiles(self, company_id: int) -> List[Dict]:
        """Get all competitor profiles for analytics"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            cursor.execute("""
                SELECT * FROM competitor_profiles
                WHERE company_id = ?
                ORDER BY total_appearances DESC
            """, (company_id,))
            return [dict(row) for row in cursor.fetchall()]

    def get_competitor_bid_history(self, company_id: int, competitor_name: str = None,
                                limit: int = 100) -> List[Dict]:
        """Get competitor bid history"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            if competitor_name:
                cursor.execute("""
                    SELECT * FROM competitor_bid_history
                    WHERE company_id = ? AND competitor_name = ?
                    ORDER BY bid_date DESC
                    LIMIT ?
                """, (company_id, competitor_name, limit))
            else:
                cursor.execute("""
                    SELECT * FROM competitor_bid_history
                    WHERE company_id = ?
                    ORDER BY bid_date DESC
                    LIMIT ?
                """, (company_id, limit))
            
            return [dict(row) for row in cursor.fetchall()]
    
    def update_competitor_bid(self, tender_id: str, competitor_name: str, 
                         bid_amount: float, was_winner: bool = False) -> bool:
        """
        Force update bid amount and winner status.
        Always updates to ensure winner deselection works.
        """
        try:
            company_id = st.session_state.get('company_id')
            if not company_id:
                return False

            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                # Calculate bid ratio
                official_estimate = 1.0
                cursor.execute("SELECT official_estimate FROM company_tenders WHERE tender_id = ?", (tender_id,))
                row = cursor.fetchone()
                if row and row['official_estimate']:
                    official_estimate = float(row['official_estimate'])
                
                bid_ratio = bid_amount / official_estimate if official_estimate > 0 else 0.95

                # === FORCE UPDATE (always write) ===
                cursor.execute("""
                    UPDATE competitor_bid_history 
                    SET bid_amount = ?,
                        was_winner = ?,
                        bid_ratio = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE tender_id = ? AND competitor_name = ?
                """, (bid_amount, 1 if was_winner else 0, bid_ratio, tender_id, competitor_name))
                
                if cursor.rowcount == 0:
                    # Insert if record doesn't exist
                    cursor.execute("""
                        INSERT INTO competitor_bid_history 
                        (company_id, competitor_name, tender_id, bid_amount, official_estimate, 
                        bid_ratio, was_winner, bid_date, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    """, (
                        company_id, competitor_name, tender_id, bid_amount, official_estimate,
                        bid_ratio, 1 if was_winner else 0, datetime.now().date()
                    ))
                
                conn.commit()
                
                # Update competitor statistics
                self.update_competitor_stats_from_bid(
                    company_id=company_id,
                    competitor_name=competitor_name,
                    bid_ratio=bid_ratio,
                    was_winner=was_winner
                )
                
                return True
                
        except Exception as e:
            print(f"Error updating bid for {competitor_name}: {e}")
            return False



    def add_competitor_bid_history(self, company_id: int, competitor_name: str,
                                tender_id: str, bid_amount: float,
                                official_estimate: float, was_winner: bool = False,
                                bid_date: str = None) -> bool:
        """Add a competitor bid to history"""
        if not bid_date:
            bid_date = datetime.now().date()
        
        bid_ratio = bid_amount / official_estimate if official_estimate > 0 else 1.0
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                INSERT INTO competitor_bid_history (
                    company_id, competitor_name, tender_id, bid_amount,
                    official_estimate, bid_ratio, was_winner, bid_date
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (company_id, competitor_name, tender_id, bid_amount,
                official_estimate, bid_ratio, 1 if was_winner else 0, bid_date))
            
            self.update_competitor_stats_from_bid(company_id, competitor_name, bid_ratio, was_winner)
            
            return True

    def get_competitor_insights(self, company_id: int) -> Dict:
        """Get aggregated competitor insights"""
        profiles = self.get_competitor_profiles(company_id)
        
        if not profiles:
            return {
                'total_competitors': 0,
                'aggressive_count': 0,
                'moderate_count': 0,
                'conservative_count': 0,
                'avg_market_ratio': 0.90
            }
        
        strategies = [p.get('strategy', 'Moderate') for p in profiles]
        
        return {
            'total_competitors': len(profiles),
            'aggressive_count': strategies.count('Aggressive'),
            'moderate_count': strategies.count('Moderate'),
            'conservative_count': strategies.count('Conservative'),
            'avg_market_ratio': sum(p.get('avg_bid_ratio', 0.90) for p in profiles) / len(profiles)
        }

    def predict_competitor_bid_ratio(self, company_id: int, competitor_name: str) -> Dict:
        """Predict competitor's bid ratio based on historical data"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                SELECT avg_bid_ratio, bid_std_dev, strategy, total_appearances
                FROM competitor_profiles 
                WHERE company_id = ? AND competitor_name = ?
            """, (company_id, competitor_name))
            
            profile = cursor.fetchone()
            
            if profile:
                return {
                    'avg_bid_ratio': profile['avg_bid_ratio'],
                    'bid_std_dev': profile['bid_std_dev'],
                    'strategy': profile['strategy'],
                    'appearances': profile['total_appearances'],
                    'confidence': min(0.95, 0.50 + (profile['total_appearances'] * 0.03))
                }
            else:
                return {
                    'avg_bid_ratio': 0.92,
                    'bid_std_dev': 0.03,
                    'strategy': 'Unknown',
                    'appearances': 0,
                    'confidence': 0.40
                }
    # =========================================================
    # BATCH 4: UPDATE METHODS
    # =========================================================


    def update_historical_tender_schema(self):
        """Add new columns for winner tracking if not exists"""
        import sqlite3
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            try:
                cursor.execute("ALTER TABLE historical_tenders ADD COLUMN winning_company_type TEXT")
                print("✓ Added winning_company_type column")
            except sqlite3.OperationalError:
                pass
            
            try:
                cursor.execute("ALTER TABLE historical_tenders ADD COLUMN our_awarded_price REAL")
                print("✓ Added our_awarded_price column")
            except sqlite3.OperationalError:
                pass

    def update_tender_lock_status(self, tender_id: int, locked: bool) -> bool:
        """Update the lock status of a tender"""
        import streamlit as st
        from datetime import datetime
        
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute("""
                    UPDATE company_tenders 
                    SET is_locked = ?, locked_at = ?, locked_by = ?
                    WHERE id = ?
                """, (
                    1 if locked else 0,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S') if locked else None,
                    st.session_state.get('user_id') if locked else None,
                    tender_id
                ))
            return True
        except Exception as e:
            logger.error(f"Failed to update tender lock status: {e}")
            return False

    def update_role_permissions_for_rates(self):
        """Add rate management permissions to role_permissions table"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            rate_permissions = {
                'admin': {
                    'manage_zones': True, 'manage_chapters': True, 'manage_parents': True,
                    'manage_children': True, 'manage_versions': True, 'view_rates': True,
                    'edit_rates': True, 'delete_rates': True
                },
                'system_admin': {
                    'manage_zones': True, 'manage_chapters': True, 'manage_parents': True,
                    'manage_children': True, 'manage_versions': True, 'view_rates': True,
                    'edit_rates': True, 'delete_rates': True
                },
                'company_admin': {
                    'manage_zones': False, 'manage_chapters': True, 'manage_parents': True,
                    'manage_children': True, 'manage_versions': True, 'view_rates': True,
                    'edit_rates': True, 'delete_rates': False
                },
                'manager': {
                    'manage_zones': False, 'manage_chapters': True, 'manage_parents': True,
                    'manage_children': True, 'manage_versions': False, 'view_rates': True,
                    'edit_rates': True, 'delete_rates': False
                },
                'analyst': {
                    'manage_zones': False, 'manage_chapters': False, 'manage_parents': False,
                    'manage_children': True, 'manage_versions': False, 'view_rates': True,
                    'edit_rates': True, 'delete_rates': False
                },
                'viewer': {
                    'manage_zones': False, 'manage_chapters': False, 'manage_parents': False,
                    'manage_children': False, 'manage_versions': False, 'view_rates': True,
                    'edit_rates': False, 'delete_rates': False
                }
            }
            
            for role, perms in rate_permissions.items():
                existing = self.get_role_permissions(role)
                existing.update(perms)
                self.update_role_permissions(role, existing)
            
            print("✅ Rate management permissions added to roles")

    # =========================================================
    # BATCH 5: HIERARCHY UPDATE METHODS
    # =========================================================

    def migrate_lged_tables_for_parent_types(self):
        """Migrate existing LGED tables to support parent types"""
        import sqlite3
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("PRAGMA table_info(lged_parents)")
            columns = [col[1] for col in cursor.fetchall()]
            
            if 'parent_type' not in columns:
                cursor.execute("ALTER TABLE lged_parents ADD COLUMN parent_type TEXT DEFAULT 'section_header'")
                print("✅ Added parent_type to lged_parents")
            
            if 'has_children' not in columns:
                cursor.execute("ALTER TABLE lged_parents ADD COLUMN has_children BOOLEAN DEFAULT 0")
                print("✅ Added has_children to lged_parents")
            
            if 'unit' not in columns:
                cursor.execute("ALTER TABLE lged_parents ADD COLUMN unit TEXT")
                print("✅ Added unit to lged_parents")
            
            cursor.execute("PRAGMA table_info(lged_children)")
            child_columns = [col[1] for col in cursor.fetchall()]
            
            if 'zone_a' not in child_columns:
                cursor.execute("ALTER TABLE lged_children ADD COLUMN zone_a REAL")
                cursor.execute("ALTER TABLE lged_children ADD COLUMN zone_b REAL")
                cursor.execute("ALTER TABLE lged_children ADD COLUMN zone_c REAL")
                cursor.execute("ALTER TABLE lged_children ADD COLUMN zone_d REAL")
                print("✅ Added zone rate columns to lged_children")

    def update_lged_hierarchy(self, hierarchy: Dict, edition_year: int, notes: str = None) -> Dict:
        """UPDATE existing active version (replaces data, no version increment)"""
        from datetime import datetime
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Get the active version
            cursor.execute("""
                SELECT id, version_number FROM rate_versions 
                WHERE source = 'LGED' AND edition_year = ? AND is_active = 1
            """, (edition_year,))
            
            result = cursor.fetchone()
            if not result:
                return {
                    'success': False,
                    'error': f"No active version found for LGED {edition_year}",
                    'message': "❌ No active version found"
                }
            
            version_id = result['id']
            current_version = result['version_number']
            
            # Clear existing data
            cursor.execute("DELETE FROM lged_zone_rates WHERE version_id = ?", (version_id,))
            cursor.execute("DELETE FROM lged_children WHERE version_id = ?", (version_id,))
            cursor.execute("DELETE FROM lged_parents WHERE version_id = ?", (version_id,))
            
            section_headers = hierarchy.get('section_headers', [])
            leaf_items = hierarchy.get('leaf_items', [])
            children = hierarchy.get('children', [])
            
            parent_ids = {}
            
            # Save section headers
            for header in section_headers:
                cursor.execute("""
                    INSERT INTO lged_parents (code, description, chapter_number, section_number, 
                                            parent_type, has_children, version_id)
                    VALUES (?, ?, ?, ?, 'section_header', ?, ?)
                """, (header['code'], header.get('description', '')[:500], 
                    header.get('chapter_number', ''), header.get('section_number', ''),
                    1 if header.get('has_children') else 0, version_id))
                parent_ids[header['code']] = cursor.lastrowid
            
            # Save leaf items
            for leaf in leaf_items:
                cursor.execute("""
                    INSERT INTO lged_parents (code, description, chapter_number, section_number, 
                                            parent_type, has_children, unit, version_id)
                    VALUES (?, ?, ?, ?, 'leaf_item', 0, ?, ?)
                """, (leaf['code'], leaf.get('description', '')[:500], 
                    leaf.get('chapter_number', ''), leaf.get('section_number', ''),
                    leaf.get('unit', ''), version_id))
                parent_id = cursor.lastrowid
                parent_ids[leaf['code']] = parent_id
                
                for zone, rate in leaf.get('rates', {}).items():
                    cursor.execute("""
                        INSERT INTO lged_zone_rates (parent_id, zone_name, unit_rate, version_id)
                        VALUES (?, ?, ?, ?)
                    """, (parent_id, zone, rate, version_id))
            
            # Save child items
            for child in children:
                cursor.execute("""
                    INSERT INTO lged_children (code, parent_code, description, unit, 
                                            chapter_number, section_number,
                                            zone_a, zone_b, zone_c, zone_d,
                                            edition_year, version_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (child['code'], child.get('parent_code', ''), child.get('description', '')[:500], 
                    child.get('unit', ''), child.get('chapter_number', ''), child.get('section_number', ''),
                    child.get('zone_a'), child.get('zone_b'), child.get('zone_c'), child.get('zone_d'),
                    edition_year, version_id))
            
            total_parents = len(section_headers) + len(leaf_items)
            total_children = len(children)
            total_rates = len(leaf_items) + len(children)
            
            cursor.execute("""
                UPDATE rate_versions 
                SET total_parents = ?, total_children = ?, total_rates = ?,
                    updated_at = CURRENT_TIMESTAMP, notes = ?
                WHERE id = ?
            """, (total_parents, total_children, total_rates, notes or "Updated via import", version_id))
            
            return {
                'success': True,
                'version_id': version_id,
                'version_number': current_version,
                'mode': 'update_existing',
                'message': f"✅ Updated existing version {current_version} for LGED {edition_year}"
            }

    def update_lged_chapter_section(self, hierarchy: Dict, version_id: int, edition_year: int,
                                    chapter_num: str, section_num: str = None, notes: str = None) -> Dict:
        """Update data for a specific chapter/section within an existing version"""
        from datetime import datetime
        
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("SELECT id, version_number FROM rate_versions WHERE id = ? AND source = 'LGED'", (version_id,))
            version = cursor.fetchone()
            
            if not version:
                return {'success': False, 'error': f"Version {version_id} not found", 'message': "Version not found"}
            
            version_number = version['version_number']
            
            section_headers = hierarchy.get('section_headers', [])
            rate_items = hierarchy.get('rate_items', [])
            
            # Delete existing data for this specific chapter/section
            if section_num:
                cursor.execute("DELETE FROM lged_children WHERE version_id = ? AND chapter_number = ? AND section_number = ?", 
                            (version_id, chapter_num, section_num))
                cursor.execute("DELETE FROM lged_parents WHERE version_id = ? AND chapter_number = ? AND section_number = ?", 
                            (version_id, chapter_num, section_num))
            else:
                cursor.execute("DELETE FROM lged_children WHERE version_id = ? AND chapter_number = ?", 
                            (version_id, chapter_num))
                cursor.execute("DELETE FROM lged_parents WHERE version_id = ? AND chapter_number = ?", 
                            (version_id, chapter_num))
            
            # Save section headers
            for header in section_headers:
                cursor.execute("""
                    INSERT INTO lged_parents (code, description, chapter_number, section_number, 
                                            parent_type, has_children, version_id)
                    VALUES (?, ?, ?, ?, 'section_header', ?, ?)
                """, (header['code'], header.get('description', '')[:500], 
                    header.get('chapter_number', chapter_num), 
                    header.get('section_number', section_num or ''),
                    1 if header.get('has_children') else 0, version_id))
            
            # Save rate items
            for item in rate_items:
                cursor.execute("""
                    INSERT INTO lged_children (code, parent_code, description, unit, 
                                            chapter_number, section_number,
                                            zone_a, zone_b, zone_c, zone_d,
                                            edition_year, version_id, is_parent)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (item['code'], item.get('parent_code'), item.get('description', '')[:500], 
                    item.get('unit', ''), item.get('chapter_number', chapter_num), 
                    item.get('section_number', section_num or ''),
                    item.get('zone_a'), item.get('zone_b'), item.get('zone_c'), item.get('zone_d'),
                    edition_year, version_id, 1 if item.get('is_parent') else 0))
            
            # Update version statistics
            cursor.execute("SELECT COUNT(*) FROM lged_parents WHERE version_id = ?", (version_id,))
            total_parents = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM lged_children WHERE version_id = ?", (version_id,))
            total_children = cursor.fetchone()[0]
            total_rates = total_children * 4
            
            cursor.execute("""
                UPDATE rate_versions 
                SET total_parents = ?, total_children = ?, total_rates = ?,
                    updated_at = CURRENT_TIMESTAMP, notes = ?
                WHERE id = ?
            """, (total_parents, total_children, total_rates, notes, version_id))
            
            return {
                'success': True,
                'version_id': version_id,
                'version_number': version_number,
                'chapter': chapter_num,
                'section': section_num,
                'mode': 'update_chapter_section',
                'message': f"✅ Updated Chapter {chapter_num}{' Section ' + section_num if section_num else ''} in Version {version_number}"
            }
    # =========================================================
    # BATCH 6: REMAINING METHODS
    # =========================================================

    # Add this method to UnifiedDatabaseManager class
    def get_tender_by_id(self, tender_id: str, company_id: Optional[int] = None) -> Optional[Dict]:
        """
        Get tender details by ID from company_tenders table
        
        Args:
            tender_id: The tender ID to look up
            company_id: Optional company ID to filter by. If not provided, will try to get from session.
        
        Returns:
            Dict: Tender data or None if not found
        """
        try:
            # If company_id not provided, try to get from session
            if company_id is None:
                import streamlit as st
                company_id = st.session_state.get('company_id')
            
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                
                if company_id:
                    # Query with company_id filter
                    cursor.execute("""
                        SELECT * FROM company_tenders 
                        WHERE tender_id = ? AND company_id = ? AND is_active = 1
                    """, (tender_id, company_id))
                    row = cursor.fetchone()
                    if row:
                        return dict(row)
                    
                    # If not found, log warning
                    print(f"⚠️ Tender {tender_id} not found for company {company_id}")
                    return None
                else:
                    # Fallback: query without company_id (not recommended)
                    cursor.execute("""
                        SELECT * FROM company_tenders 
                        WHERE tender_id = ? AND is_active = 1
                    """, (tender_id,))
                    row = cursor.fetchone()
                    if row:
                        print(f"⚠️ Warning: get_tender_by_id called without company_id for tender {tender_id}")
                        return dict(row)
                    return None
                    
        except Exception as e:
            print(f"❌ Error in get_tender_by_id: {str(e)}")
            return None
        
    def insert_historical_tender(self, company_id: int, historical_data: Dict) -> bool:
        """Insert historical tender record"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                INSERT INTO historical_tenders (
                    company_id, tender_id, tender_title, procuring_entity,
                    procurement_type, official_estimate, awarded_price,
                    num_competitors, total_bidders, winning_competitor,
                    winning_company_type, competitors_data, award_date, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                company_id,
                historical_data.get('tender_id'),
                historical_data.get('tender_title', ''),
                historical_data.get('procuring_entity', ''),
                historical_data.get('procurement_type', ''),
                historical_data.get('official_estimate', 0),
                historical_data.get('awarded_price', 0),
                historical_data.get('num_competitors', 0),
                historical_data.get('total_bidders', 0),
                historical_data.get('winning_competitor', ''),
                historical_data.get('winning_company_type', ''),
                historical_data.get('competitors_data', '[]'),
                historical_data.get('award_date'),
                historical_data.get('notes', '')
            ))
            
            return True
    
    def get_company_tenders(self, company_id: int, search_term: str = "", limit: int = 100) -> pd.DataFrame:
        """Get company tenders with optional search"""
        import pandas as pd  # ✅ Ensure pandas is imported
        
        try:
            # ✅ Ensure company_id is integer
            if company_id is None:
                print("⚠️ get_company_tenders: company_id is None")
                return pd.DataFrame()
            
            company_id = int(company_id)
            
            print(f"🔍 get_company_tenders: company_id={company_id}, search='{search_term}'")
            
            with self.get_connection() as conn:
                query = """
                    SELECT * FROM company_tenders 
                    WHERE company_id = ? AND is_active = 1
                """
                params = [company_id]
                
                if search_term:
                    query += " AND (tender_id LIKE ? OR tender_title LIKE ?)"
                    params.extend([f"%{search_term}%", f"%{search_term}%"])
                
                query += " ORDER BY created_at DESC LIMIT ?"
                params.append(limit)
                
                print(f"   Query: {query}")
                print(f"   Params: {params}")
                
                return pd.read_sql_query(query, conn, params=params)
                
        except Exception as e:
            print(f"❌ Error in get_company_tenders: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def create_tender(self, company_id: int, tender_data: Dict, created_by: int) -> Optional[int]:
        """Create a new tender with full e-GP fields"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)

            cursor.execute("""
                INSERT INTO company_tenders (
                    company_id, tender_id, tender_title, procuring_entity,
                    division, district, thana, country, procurement_type,
                    official_estimate, submission_deadline, tender_security,
                    document_fee, evaluation_type, mode_of_payment,
                    eligibility_criteria, invitation_ref_no, package_no,
                    project_code, project_name, inviting_official_name,
                    inviting_official_designation, inviting_official_phone,
                    inviting_official_email, inviting_official_address,
                    inviting_official_city, inviting_official_thana,
                    inviting_official_district, notes, app_id,
                    procuring_entity_code, procurement_nature, event_type,
                    budget_type, source_of_funds, category,
                    tender_publication_date, document_selling_end_date,
                    pre_bid_meeting_start, pre_bid_meeting_end,
                    bid_opening_date, security_submission_deadline,
                    security_valid_upto, tender_valid_upto,
                    created_by, created_at
                ) VALUES (
                    ?, ?, ?, ?,
                    ?, ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,                    
                    ?, CURRENT_TIMESTAMP
                )
            """, (
                company_id,
                tender_data.get('tender_id'),
                tender_data.get('tender_title'),
                tender_data.get('procuring_entity'),
                tender_data.get('division'),
                tender_data.get('district'),
                tender_data.get('thana'),
                tender_data.get('country', 'Bangladesh'),
                tender_data.get('procurement_type'),
                tender_data.get('official_estimate'),
                tender_data.get('submission_deadline'),
                tender_data.get('tender_security'),
                tender_data.get('document_fee'),
                tender_data.get('evaluation_type'),
                tender_data.get('mode_of_payment'),
                tender_data.get('eligibility_criteria'),
                tender_data.get('invitation_ref_no'),
                tender_data.get('package_no'),
                tender_data.get('project_code'),
                tender_data.get('project_name'),
                tender_data.get('inviting_official_name'),
                tender_data.get('inviting_official_designation'),
                tender_data.get('inviting_official_phone'),
                tender_data.get('inviting_official_email'),
                tender_data.get('inviting_official_address'),
                tender_data.get('inviting_official_city'),
                tender_data.get('inviting_official_thana'),
                tender_data.get('inviting_official_district'),
                tender_data.get('notes'),
                tender_data.get('app_id'),
                tender_data.get('procuring_entity_code'),
                tender_data.get('procurement_nature'),
                tender_data.get('event_type'),
                tender_data.get('budget_type'),
                tender_data.get('source_of_funds'),
                tender_data.get('category'),
                tender_data.get('tender_publication_date'),
                tender_data.get('document_selling_end_date'),
                tender_data.get('pre_bid_meeting_start'),
                tender_data.get('pre_bid_meeting_end'),
                tender_data.get('bid_opening_date'),
                tender_data.get('security_submission_deadline'),
                tender_data.get('security_valid_upto'),
                tender_data.get('tender_valid_upto'),
                created_by,
            ))

            return cursor.lastrowid

    def create_tender_bak(self, company_id: int, tender_data: Dict) -> Optional[int]:
        """Create a new tender"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                INSERT INTO company_tenders (
                    company_id, tender_id, tender_title, procuring_entity,
                    official_estimate, submission_deadline, procurement_type,
                    division, district, created_by, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (
                company_id,
                tender_data.get('tender_id'),
                tender_data.get('tender_title'),
                tender_data.get('procuring_entity'),
                tender_data.get('official_estimate'),
                tender_data.get('submission_deadline'),
                tender_data.get('procurement_type'),
                tender_data.get('division'),
                tender_data.get('district'),
                tender_data.get('created_by')
            ))
            
            return cursor.lastrowid

    def get_tender_team(self, tender_id: int):
        """Get team members assigned to a tender"""
        import pandas as pd
        with self.get_connection() as conn:
            return pd.read_sql_query("""
                SELECT tta.*, u.username, u.full_name, u.email
                FROM tender_team_assignments tta
                JOIN users u ON tta.user_id = u.id
                WHERE tta.tender_id = ? AND tta.is_active = 1
            """, conn, params=[tender_id])

    def assign_team_member(self, tender_id: int, user_id: int, role: str) -> bool:
        """Assign a team member to a tender"""
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                INSERT OR REPLACE INTO tender_team_assignments (tender_id, user_id, role, assigned_at, is_active)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP, 1)
            """, (tender_id, user_id, role))
            
            return True    
    # database/crud_operations.py - Add company config methods

    def get_company_config(self, company_id: int, key: str, default: Any = None) -> Any:
        """Get company-specific configuration value."""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute("""
                    SELECT config_value, config_type FROM company_config 
                    WHERE company_id = ? AND config_key = ?
                """, (company_id, key))
                row = cursor.fetchone()
                
                if row:
                    value = row['config_value']
                    config_type = row['config_type']
                    
                    # Parse based on type
                    if config_type == 'boolean':
                        return value.lower() == 'true'
                    elif config_type == 'number':
                        try:
                            return float(value) if '.' in value else int(value)
                        except:
                            return value
                    elif config_type == 'json':
                        try:
                            return json.loads(value)
                        except:
                            return value
                    else:
                        return value
                return default
        except Exception as e:
            logger.error(f"Error getting company config {company_id}:{key}: {e}")
            return default

    def set_company_config(self, company_id: int, key: str, value: Any, 
                        config_type: str = 'string', 
                        description: str = None,
                        user_id: int = None) -> bool:
        """Set company-specific configuration value."""
        try:
            # Determine config type
            if isinstance(value, bool):
                config_type = 'boolean'
                value = 'true' if value else 'false'
            elif isinstance(value, (int, float)):
                config_type = 'number'
                value = str(value)
            elif isinstance(value, (dict, list)):
                config_type = 'json'
                value = json.dumps(value)
            else:
                value = str(value)
            
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute("""
                    INSERT OR REPLACE INTO company_config 
                    (company_id, config_key, config_value, config_type, description, updated_by, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (company_id, key, value, config_type, description, user_id))
                return True
        except Exception as e:
            logger.error(f"Error setting company config {company_id}:{key}: {e}")
            return False

    def delete_company_config(self, company_id: int, key: str) -> bool:
        """Delete company-specific configuration."""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute("""
                    DELETE FROM company_config 
                    WHERE company_id = ? AND config_key = ?
                """, (company_id, key))
                return True
        except Exception as e:
            logger.error(f"Error deleting company config {company_id}:{key}: {e}")
            return False

    def get_all_company_configs(self, company_id: int) -> Dict[str, Any]:
        """Get all company-specific configurations."""
        try:
            with self.get_connection() as conn:
                cursor = self.db_conn.get_cursor(conn)
                cursor.execute("""
                    SELECT config_key, config_value, config_type 
                    FROM company_config 
                    WHERE company_id = ?
                """, (company_id,))
                rows = cursor.fetchall()
                
                result = {}
                for row in rows:
                    key = row['config_key']
                    value = row['config_value']
                    config_type = row['config_type']
                    
                    if config_type == 'boolean':
                        result[key] = value.lower() == 'true'
                    elif config_type == 'number':
                        try:
                            result[key] = float(value) if '.' in value else int(value)
                        except:
                            result[key] = value
                    elif config_type == 'json':
                        try:
                            result[key] = json.loads(value)
                        except:
                            result[key] = value
                    else:
                        result[key] = value
                
                return result
        except Exception as e:
            logger.error(f"Error getting all company configs for {company_id}: {e}")
            return {}
    
    
    # database/crud_operations.py - Add this method    
        
    # =========================================================
    # COMPETITOR INTELLIGENCE METHODS - ADD THESE
    # =========================================================

    def get_competitor_with_stats(self, competitor_id: int, company_id: int) -> Optional[Dict]:
        """
        Get competitor details with calculated statistics
        
        Args:
            competitor_id: The competitor's ID
            company_id: Company ID for tenant isolation
        
        Returns:
            Competitor record with stats or None
        """
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                SELECT 
                    cm.*,
                    ROUND(
                        CASE 
                            WHEN cm.total_bids > 0 
                            THEN (cm.total_wins * 100.0 / cm.total_bids) 
                            ELSE 0 
                        END, 
                        2
                    ) as win_percentage,
                    (
                        SELECT COUNT(*) 
                        FROM competitor_bid_history cbh
                        WHERE cbh.competitor_name = cm.competitor_name 
                        AND cbh.company_id = cm.company_id
                    ) as actual_total_bids
                FROM competitor_master cm
                WHERE cm.id = ? AND cm.company_id = ?
            """, (competitor_id, company_id))
            
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_competitor_bid_history_with_details(self, competitor_name: str, company_id: int) -> List[Dict]:
        """
        Get complete bid history with tender details and calculated fields
        
        Args:
            competitor_name: Name of the competitor
            company_id: Company ID for tenant isolation
        
        Returns:
            List of bid records with calculated fields
        """
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                SELECT 
                    cbh.tender_id,
                    ct.tender_title,
                    ct.procuring_entity,
                    ct.submission_deadline as closing_date,
                    ct.official_estimate as oce,
                    cbh.bid_amount,
                    cbh.bid_ratio,
                    cbh.was_winner,
                    cbh.bid_date,
                    -- Calculate discount from OCE
                    ROUND(
                        ((ct.official_estimate - cbh.bid_amount) / NULLIF(ct.official_estimate, 0)) * 100, 
                        2
                    ) as discount_from_oce,
                    -- Calculate rank (lower bid = better rank)
                    (
                        SELECT COUNT(*) + 1 
                        FROM competitor_bid_history cbh2 
                        WHERE cbh2.tender_id = cbh.tender_id 
                        AND cbh2.bid_amount < cbh.bid_amount
                        AND cbh2.company_id = cbh.company_id
                    ) as rank,
                    -- Get winning bid amount
                    (
                        SELECT MIN(cbh2.bid_amount) 
                        FROM competitor_bid_history cbh2 
                        WHERE cbh2.tender_id = cbh.tender_id 
                        AND cbh2.was_winner = 1
                        AND cbh2.company_id = cbh.company_id
                    ) as winning_bid,
                    -- Calculate NPPI (Normalized Price Performance Index)
                    ROUND(cbh.bid_amount / NULLIF(ct.official_estimate, 0), 4) as nppi,
                    -- Difference from winning bid
                    ROUND(
                        cbh.bid_amount - (
                            SELECT MIN(cbh2.bid_amount) 
                            FROM competitor_bid_history cbh2 
                            WHERE cbh2.tender_id = cbh.tender_id 
                            AND cbh2.was_winner = 1
                            AND cbh2.company_id = cbh.company_id
                        ), 
                        2
                    ) as diff_from_winning_bid
                FROM competitor_bid_history cbh
                JOIN company_tenders ct ON cbh.tender_id = ct.tender_id
                WHERE cbh.competitor_name = ? 
                AND cbh.company_id = ?
                AND ct.official_estimate > 0
                ORDER BY cbh.bid_date DESC
            """, (competitor_name, company_id))
            
            rows = cursor.fetchall()
            return [dict(row) for row in rows]

    def get_competitor_analytics(self, competitor_name: str, company_id: int) -> Dict:
        """
        Get comprehensive analytics for a competitor
        
        Args:
            competitor_name: Name of the competitor
            company_id: Company ID for tenant isolation
        
        Returns:
            Dictionary with analytics metrics
        """
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_bids,
                    SUM(CASE WHEN was_winner = 1 THEN 1 ELSE 0 END) as total_wins,
                    MIN(bid_date) as first_active,
                    MAX(bid_date) as last_active,
                    ROUND(AVG(bid_amount), 2) as avg_bid,
                    ROUND(AVG(bid_ratio), 4) as avg_bid_ratio,
                    ROUND(AVG(official_estimate), 2) as avg_oce,
                    ROUND(AVG(bid_amount / NULLIF(official_estimate, 0)), 4) as avg_nppi,
                    ROUND(
                        AVG(
                            ((official_estimate - bid_amount) / NULLIF(official_estimate, 0)) * 100
                        ), 
                        2
                    ) as avg_discount,
                    -- Additional analytics
                    ROUND(
                        AVG(
                            ((official_estimate - bid_amount) / NULLIF(official_estimate, 0)) * 100
                        ), 
                        2
                    ) as avg_discount,
                    ROUND(
                        STDEV(
                            ((official_estimate - bid_amount) / NULLIF(official_estimate, 0)) * 100
                        ), 
                        2
                    ) as std_discount,
                    ROUND(
                        MIN(
                            ((official_estimate - bid_amount) / NULLIF(official_estimate, 0)) * 100
                        ), 
                        2
                    ) as min_discount,
                    ROUND(
                        MAX(
                            ((official_estimate - bid_amount) / NULLIF(official_estimate, 0)) * 100
                        ), 
                        2
                    ) as max_discount,
                    (
                        SELECT COUNT(DISTINCT tender_id) 
                        FROM competitor_bid_history cbh2
                        WHERE cbh2.competitor_name = cbh.competitor_name 
                        AND cbh2.company_id = cbh.company_id
                    ) as total_tenders
                FROM competitor_bid_history cbh
                JOIN company_tenders ct ON cbh.tender_id = ct.tender_id
                WHERE cbh.competitor_name = ? 
                AND cbh.company_id = ?
                AND ct.official_estimate > 0
                GROUP BY cbh.competitor_name, cbh.company_id
            """, (competitor_name, company_id))
            
            row = cursor.fetchone()
            
            if not row:
                return {
                    'total_bids': 0,
                    'total_wins': 0,
                    'first_active': None,
                    'last_active': None,
                    'avg_bid': 0.0,
                    'avg_bid_ratio': 0.0,
                    'avg_oce': 0.0,
                    'avg_nppi': 0.0,
                    'avg_discount': 0.0,
                    'std_discount': 0.0,
                    'min_discount': 0.0,
                    'max_discount': 0.0,
                    'total_tenders': 0
                }
            
            result = dict(row)
            
            # Calculate derived metrics
            total_bids = result.get('total_bids', 0)
            total_wins = result.get('total_wins', 0)
            result['win_rate'] = round((total_wins / total_bids * 100), 2) if total_bids > 0 else 0.0
            
            # Calculate competition level
            if result.get('total_tenders', 0) > 0:
                result['avg_competition_level'] = round(
                    total_bids / result['total_tenders'], 2
                )
            else:
                result['avg_competition_level'] = 0.0
            
            # Calculate active months
            if result.get('first_active') and result.get('last_active'):
                first_date = pd.to_datetime(result['first_active'])
                last_date = pd.to_datetime(result['last_active'])
                result['active_months'] = round(((last_date - first_date).days / 30.44), 1)
            else:
                result['active_months'] = 0
            
            return result

    def get_competitor_activity_insights(self, competitor_name: str, company_id: int) -> List[str]:
        """
        Generate activity-based insights for a competitor
        
        Args:
            competitor_name: Name of the competitor
            company_id: Company ID for tenant isolation
        
        Returns:
            List of insight strings
        """
        insights = []
        
        # Get analytics
        analytics = self.get_competitor_analytics(competitor_name, company_id)
        
        if analytics.get('total_bids', 0) == 0:
            insights.append("No historical data available for this competitor")
            return insights
        
        # 1. Activity status based on first appearance
        active_months = analytics.get('active_months', 0)
        if active_months < 3:
            insights.append("🆕 New entrant: First appeared recently (less than 3 months ago)")
        elif active_months < 12:
            insights.append("📈 Active competitor: Participating regularly for under 1 year")
        elif active_months < 24:
            insights.append("🏛️ Established competitor: Active for 1-2 years")
        else:
            insights.append("🏗️ Veteran competitor: Active for over 2 years")
        
        # 2. Recent activity status
        last_active = analytics.get('last_active')
        if last_active:
            days_since_last = (datetime.now().date() - last_active).days
            if days_since_last > 365:
                insights.append(f"⚠️ Inactive: No bids recorded for over 1 year")
            elif days_since_last > 180:
                insights.append(f"⏸️ Less active: Last bid was {days_since_last} days ago")
            elif days_since_last > 90:
                insights.append(f"📉 Decreasing activity: Last bid was {days_since_last} days ago")
            elif days_since_last > 30:
                insights.append(f"📊 Moderately active: Last bid was {days_since_last} days ago")
            else:
                insights.append(f"⚡ Highly active: Recently bid {days_since_last} days ago")
        
        # 3. Bidding frequency
        total_bids = analytics.get('total_bids', 0)
        if active_months > 0:
            bids_per_month = total_bids / active_months
            if bids_per_month > 3:
                insights.append(f"🔥 High frequency: ~{bids_per_month:.1f} bids per month")
            elif bids_per_month > 1:
                insights.append(f"📊 Moderate frequency: ~{bids_per_month:.1f} bids per month")
            else:
                insights.append(f"🐢 Low frequency: ~{bids_per_month:.1f} bids per month")
        
        return insights

    def get_competitor_behavioral_insights(self, competitor_name: str, company_id: int) -> List[str]:
        """
        Generate behavioral insights based on bidding patterns
        
        Args:
            competitor_name: Name of the competitor
            company_id: Company ID for tenant isolation
        
        Returns:
            List of insight strings
        """
        insights = []
        
        analytics = self.get_competitor_analytics(competitor_name, company_id)
        
        if analytics.get('total_bids', 0) == 0:
            return insights
        
        # 1. Bidding consistency based on discount volatility
        volatility = analytics.get('std_discount', 0)
        if volatility < 3:
            insights.append("🎯 Highly consistent bidder (very low variance in discounts)")
        elif volatility < 8:
            insights.append("📊 Consistent bidder (moderate variance in discounts)")
        elif volatility < 15:
            insights.append("🔄 Variable bidder (significant variance in discounts)")
        else:
            insights.append("🎲 Aggressive bidder (highly variable discount patterns)")
        
        # 2. Win rate analysis
        win_rate = analytics.get('win_rate', 0)
        if win_rate > 40:
            insights.append("🏆 Strong performer: Win rate exceeds 40%")
        elif win_rate > 25:
            insights.append("📈 Competitive performer: Win rate between 25-40%")
        elif win_rate > 10:
            insights.append("📊 Developing performer: Win rate between 10-25%")
        else:
            insights.append("📉 Needs improvement: Win rate below 10%")
        
        # 3. Discount strategy
        avg_discount = analytics.get('avg_discount', 0)
        if avg_discount > 20:
            insights.append("💪 Aggressive pricing: Average discount above 20%")
        elif avg_discount > 10:
            insights.append("📊 Balanced pricing: Average discount between 10-20%")
        elif avg_discount > 5:
            insights.append("💰 Conservative pricing: Average discount between 5-10%")
        else:
            insights.append("🛡️ Premium positioning: Average discount below 5%")
        
        return insights

    def get_competitor_chart_data(self, competitor_name: str, company_id: int) -> Dict:
        """
        Get data for charts in a format ready for Plotly
        
        Args:
            competitor_name: Name of the competitor
            company_id: Company ID for tenant isolation
        
        Returns:
            Dictionary with chart data
        """
        history = self.get_competitor_bid_history_with_details(competitor_name, company_id)
        
        if not history:
            return {}
        
        df = pd.DataFrame(history)
        df['bid_date'] = pd.to_datetime(df['bid_date'])
        
        return {
            'discount_vs_time': df[['bid_date', 'discount_from_oce']].dropna().sort_values('bid_date'),
            'bid_vs_oce': df[['oce', 'bid_amount']].dropna(),
            'win_rate_trend': df.sort_values('bid_date')['was_winner'].rolling(5, min_periods=1).mean() * 100,
            'participation_timeline': df.groupby('bid_date').size().reset_index(name='count'),
            'bid_distribution': df['bid_amount'].dropna(),
            'rank_distribution': df['rank'].dropna()
        }

    def get_paginated_competitors(self, company_id: int, limit: int = 20, offset: int = 0, 
                                  search: str = None, sort_by: str = 'competitor_name') -> Dict:
        """
        Get paginated list of competitors with stats
        
        Args:
            company_id: Company ID for tenant isolation
            limit: Number of records to return
            offset: Offset for pagination
            search: Search term for competitor name
            sort_by: Column to sort by
        
        Returns:
            Dictionary with 'competitors' list and 'total' count
        """
        with self.get_connection() as conn:
            cursor = self.db_conn.get_cursor(conn)
            
            # Build query with calculated win percentage
            base_sql = """
                SELECT 
                    cm.*,
                    ROUND(
                        CASE 
                            WHEN cm.total_bids > 0 
                            THEN (cm.total_wins * 100.0 / cm.total_bids) 
                            ELSE 0 
                        END, 
                        2
                    ) as win_percentage
                FROM competitor_master cm
                WHERE cm.company_id = ?
                AND cm.is_active = 1
            """
            
            count_sql = """
                SELECT COUNT(*) as total
                FROM competitor_master cm
                WHERE cm.company_id = ?
                AND cm.is_active = 1
            """
            
            params = [company_id]
            count_params = [company_id]
            
            # Add search filter
            if search:
                base_sql += " AND cm.competitor_name LIKE ?"
                count_sql += " AND cm.competitor_name LIKE ?"
                search_param = f"%{search}%"
                params.append(search_param)
                count_params.append(search_param)
            
            # Add sorting
            valid_sort_columns = ['competitor_name', 'first_seen', 'last_seen', 'total_bids', 'win_percentage']
            if sort_by in valid_sort_columns:
                base_sql += f" ORDER BY cm.{sort_by} ASC"
            else:
                base_sql += " ORDER BY cm.competitor_name ASC"
            
            # Add pagination
            base_sql += " LIMIT ? OFFSET ?"
            params.extend([limit, offset])
            
            # Execute queries
            cursor.execute(base_sql, params)
            rows = cursor.fetchall()
            competitors = [dict(row) for row in rows]
            
            cursor.execute(count_sql, count_params)
            total_row = cursor.fetchone()
            total = total_row['total'] if total_row else 0
            
            return {
                'competitors': competitors,
                'total': total
            }